"""Diagnóstico (Fase 8, SP-Unificado): como um COMANDO (output) aparece no dado.

Bloqueia a spec de classificação-direção (D1-D4): sem saber como a direção é
sinalizada na entrada, qualquer regra é chute. Este script cruza, por endereço
DNP3, os sinais que o TDT REAL marca como comando (Direction ∈ {Write,
ReadWrite}) com o que o NOSSO pipeline atribuiu à mesma entrada, e dumpa o
texto bruto de entrada — para revelar o padrão determinístico (se houver).

Uso: PYTHONPATH=src python bench/diag_direcao_comando.py
Saída: bench/resultados/diag_direcao_comando.log

NÃO muda produção. Só instrumenta. A decisão de design (como classificar
direção) é do usuário, com este achado em mãos.
"""
from __future__ import annotations

import warnings
import logging
from collections import Counter
from pathlib import Path

warnings.simplefilter("ignore")
logging.disable(logging.CRITICAL)

import openpyxl

from tdt.config import Config
from tdt.dados.encoder import criar_encoder
from tdt.pipeline import executar

_INPUT = "docs/input_nao_homogeneo_1_GTA.xlsx"
_TEMPLATE = "docs/dnp3_template.xlsx"
_LISTA = "docs/Pontos Padrao ADMS_v2.xlsx"
_REAL = "docs/TDT/exportTDT_UTR_GTD_1_20260626.xlsx"
_SHEETS = ("DNP3_DiscreteSignals", "DNP3_AnalogSignals")
_COL_NOME, _COL_DIR, _COL_MM, _COL_INCOORDS = 0, 17, 19, 31
_OUT = Path("bench/resultados/diag_direcao_comando.log")

_COMANDO = {"Write", "ReadWrite"}


def _primeiro_addr(cell) -> int | None:
    if cell is None:
        return None
    txt = str(cell).split(";")[0].strip()
    return int(txt) if txt.lstrip("-").isdigit() else None


def carregar_real() -> dict[int, tuple[str, str, str]]:
    """addr -> (direction, nome, message_mapping)."""
    wb = openpyxl.load_workbook(_REAL, read_only=True, data_only=True)
    out: dict[int, tuple[str, str, str]] = {}
    for sn in _SHEETS:
        if sn not in wb.sheetnames:
            continue
        for r in wb[sn].iter_rows(min_row=5, values_only=True):
            addr = _primeiro_addr(r[_COL_INCOORDS] if len(r) > _COL_INCOORDS else None)
            if addr is None:
                continue
            direcao = str(r[_COL_DIR]) if len(r) > _COL_DIR and r[_COL_DIR] else "None"
            nome = str(r[_COL_NOME]) if r[_COL_NOME] else ""
            mm = str(r[_COL_MM]) if len(r) > _COL_MM and r[_COL_MM] else ""
            out[addr] = (direcao, nome, mm)
    wb.close()
    return out


def carregar_nosso() -> dict[int, tuple[str, str]]:
    """addr(input) -> (direcao_nossa, texto_bruto)."""
    cfg = Config()
    resultado, _wb = executar(
        _INPUT, _TEMPLATE, _LISTA,
        config=cfg, encoder=criar_encoder(cfg.modelo_embedding), subestacao="GTD",
    )
    out: dict[int, tuple[str, str]] = {}
    registros = list(resultado.lista.registros) + [ir.registro for ir in resultado.revisao]
    for rec in registros:
        for idx in rec.enderecamento.indices:
            out[idx] = (rec.tipo_sinal.direcao, rec.descricoes.bruta)
    return out


def main() -> None:
    real = carregar_real()
    nosso = carregar_nosso()

    comandos = {a: v for a, v in real.items() if v[0] in _COMANDO}
    pares: Counter[tuple[str, str]] = Counter()  # (dir_real, dir_nossa|ausente)
    amostras_erradas: list[str] = []

    for addr, (dir_real, nome, mm) in sorted(comandos.items()):
        nz = nosso.get(addr)
        dir_nossa = nz[0] if nz else "AUSENTE"
        pares[(dir_real, dir_nossa)] += 1
        acertou = dir_nossa in ("Output", "InputOutput")
        if not acertou and len(amostras_erradas) < 40:
            texto = nz[1] if nz else "—"
            amostras_erradas.append(
                f"  addr={addr:>5} real={dir_real:<9} nossa={dir_nossa:<10} "
                f"nome={nome:<28} texto={texto!r}\n    mm={mm}"
            )

    linhas: list[str] = []
    linhas.append("# diag_direcao_comando — SP-Unificado Fase 8 (2026-07-08)\n")
    linhas.append(f"comandos reais (Write/ReadWrite): {len(comandos)}\n")
    linhas.append("## cruzamento (direção real × direção nossa) por endereço:")
    for (dr, dn), n in pares.most_common():
        linhas.append(f"  {dr:<9} -> {dn:<10} : {n}")
    acertos = sum(n for (dr, dn), n in pares.items() if dn in ("Output", "InputOutput"))
    linhas.append(f"\nacerto de direção (Output/InputOutput): {acertos}/{len(comandos)}")
    linhas.append("\n## amostra de comandos reais que NÃO marcamos como comando:")
    linhas.extend(amostras_erradas)

    texto = "\n".join(linhas) + "\n"
    _OUT.write_text(texto, encoding="utf-8")
    print(texto)
    print(f"log em {_OUT}")


if __name__ == "__main__":
    main()
