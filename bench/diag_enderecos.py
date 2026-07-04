"""Diagnóstico de ruído na coluna de endereço DNP3 (SP-M, Fase 1 -- Levantamento).

Para cada lista/sheet de entrada real (nao GTD sintética/mockup): qual coluna o
DETECTOR REAL do pipeline escolhe como "indice" (endereço DNP3) e quais outras
colunas numéricas próximas poderiam, em tese, ser confundidas com ela (mesma
faixa de valores, monotonicidade parecida, etc.) -- e, quando há como checar
(cross-referência com TDT exportado conhecido), se o valor realmente lido pelo
pipeline bate com o endereço verdadeiro.

Critérios vigentes (lidos em `src/tdt/analise/analise_colunas.py` e
`src/tdt/inferencia_topologia.py`, ver docstring no fim do arquivo):

- Caminho HOMOGÊNEO (`estruturador_homogeneo.py`): sem heurística nenhuma --
  a coluna de endereço é a que tem o cabeçalho EXATO "INDEX DNP3" (comparação
  literal, sem acento/maiúsculas). Sem ambiguidade possível: ou o cabeçalho
  bate e usa aquela coluna, ou não bate e cai no caminho não-homogêneo.
- Caminho NÃO-HOMOGÊNEO (`analise_colunas._col_indice`, chamado por
  `analisar()`): heurística por CONTEÚDO, não por header. Para cada coluna,
  calcula `score = frac_inteiros * (0.5 + 0.5*monotonicidade_crescente)`
  onde `frac_inteiros` é a fração de células que casam `^-?\\d+$` e
  `monotonicidade` é a fração de pares consecutivos estritamente crescentes.
  Escolhe a coluna de MAIOR score (sem piso mínimo de score, sem checagem de
  faixa de valores, sem desempate por posição/proximidade de "Endereço DNP3"
  no header). Rodadas em TODAS as colunas da sheet, incluindo eventuais
  colunas de contagem/índice de linha, ANSI (79/67/51N), IDs de equipamento
  (52-1, 01Q0) etc. -- daí a hipótese de confusão que este script testa.
- `inferencia_topologia.py` (blocos de endereço contíguos) NÃO é usado hoje
  para detectar a coluna de endereço em si -- é usado só depois, em
  `subdividir_transformador_at_bt` (C2.4), para inferir o LADO (AT/BT) de um
  módulo Transformador a partir de blocos contíguos de endereços já
  extraídos. Não é (ainda) um sinal alternativo/de reforço para a escolha de
  coluna -- é o "modelo de blocos" citado no design como sinal FUTURO
  possível, não vigente.

Uso:
    PYTHONPATH=src python bench/diag_enderecos.py docs/*.xlsx
    PYTHONPATH=src python bench/diag_enderecos.py   # varre lista default (ver ARQUIVOS)

Saída: bench/resultados/diag_enderecos.log
"""
from __future__ import annotations

import glob
import re
import sys
import warnings
import logging
import zipfile
from pathlib import Path

warnings.simplefilter("ignore")
logging.disable(logging.CRITICAL)

import openpyxl

from tdt.analise.analise_colunas import (
    _col_indice,
    _header_por_densidade,
    _ncols,
    _norm,
    _valores_coluna,
    analisar,
    normalizar_emb,
)
from tdt.analise.identificador import classificar, ler_rows
from tdt.config import Config
from tdt.dados.encoder import criar_encoder
from tdt.dados.lista_padrao import ListaPadraoADMS
from tdt.normalizacao.estruturador_homogeneo import detectar_header
from tdt.normalizacao.normalizador import canonizar

_INT = re.compile(r"^-?\d+$")

# Listas reais tratadas como "lista de entrada de sinais" para este
# levantamento (ver critério de seleção no relatório -- excluídos: templates
# vazios dnp3_template/iec104_template, Pontos Padrao ADMS_v* (é a LISTA
# PADRÃO ADMS de referência, não uma lista de ENTRADA de UTR), TDT/* (é
# SAÍDA/ground-truth, não entrada), e arquivos não-.xlsx).
ARQUIVOS_PADRAO = [
    "docs/input_homogeneo_IMA.xlsx",
    "docs/input_nao_homogeneo_1_GTD.xlsx",
    "docs/input_nao_homogeneo_2_FWB.xlsx",
    "docs/input_nao_homogeneo_3_GPR.xlsx",
    "docs/input_nao_homogeneo_4_GAU.xlsx",
    "docs/RGE GAU 2026 - Lista de Pontos v09.xlsx",
    "docs/SAN2_LISTA_PADRONIZADA_PARA_TESTE.xlsx",
]


def _destyle_e_reabrir(path: str):
    """Alguns .xlsx reais têm <extLst> em estilos (ex: PatternFill) que o
    openpyxl 3.1.5 em Python 3.14 não consegue desserializar
    (`PatternFill.__init__() got an unexpected keyword argument 'extLst'`).
    Isso é uma limitação de ambiente (risco já registrado em memória do
    projeto), não um problema do detector de endereço -- então, ao invés de
    deixar o arquivo inteiro cair fora do levantamento, removemos os blocos
    <extLst> de xl/styles.xml (não mexe em dado/valor de célula, só em
    estilo) e reabrimos a partir de uma cópia temporária."""
    tmp = Path("bench/resultados") / f"_destyled_{Path(path).stem}.xlsx"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    zin = zipfile.ZipFile(path)
    try:
        styles = zin.read("xl/styles.xml").decode("utf-8")
    except KeyError:
        zin.close()
        raise
    styles2 = re.sub(r"<extLst>.*?</extLst>", "", styles, flags=re.S)
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            zout.writestr(item, styles2 if item.filename == "xl/styles.xml" else zin.read(item.filename))
    zin.close()
    return openpyxl.load_workbook(str(tmp), read_only=True, data_only=True)


def _abrir(path: str):
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except TypeError as e:
        if "extLst" in str(e):
            return _destyle_e_reabrir(path)
        raise


def _colunas_numericas_confundiveis(rows, inicio, ncols, col_escolhida):
    """Para cada coluna numérica (>=50% inteiros) que NÃO é a escolhida,
    reporta score bruto do MESMO critério do detector (frac*mono) + faixa de
    valores -- para julgar se é uma concorrente plausível (faixa/monotonia
    parecida com uma faixa de endereço DNP3 típica) ou claramente outra coisa
    (ANSI de 2 dígitos, contagem pequena, etc.)."""
    concorrentes = []
    for c in range(ncols):
        if c == col_escolhida:
            continue
        vals = _valores_coluna(rows, c, inicio)
        if not vals:
            continue
        ints = [int(v) for v in vals if _INT.match(v)]
        frac = len(ints) / len(vals)
        if frac < 0.5 or not ints:
            continue
        crescente = sum(1 for a, b in zip(ints, ints[1:]) if b > a)
        mono = (crescente / (len(ints) - 1)) if len(ints) > 1 else 0.0
        score = frac * (0.5 + 0.5 * mono)
        concorrentes.append({
            "col": c,
            "score": round(score, 3),
            "frac_inteiros": round(frac, 2),
            "mono": round(mono, 2),
            "min": min(ints),
            "max": max(ints),
            "n": len(ints),
        })
    concorrentes.sort(key=lambda d: -d["score"])
    return concorrentes


def _header_label(rows, header_row_1based, col):
    if col is None:
        return "?"
    hi = header_row_1based - 1
    if hi < 0 or hi >= len(rows) or col >= len(rows[hi]):
        return "?"
    return _norm(rows[hi][col])


def main(argv):
    arquivos = argv[1:] if len(argv) > 1 else ARQUIVOS_PADRAO
    resolvidos = []
    for pat in arquivos:
        resolvidos.extend(sorted(glob.glob(pat)) or [pat])

    cfg = Config()
    enc = criar_encoder(cfg.modelo_embedding)
    lp = ListaPadraoADMS.carregar("docs/Pontos Padrao ADMS_v1.xlsx")
    corpus = [(s.sigla, canonizar(s.descricao, cfg)) for s in lp.discretos if s.descricao]
    ref = normalizar_emb(enc([d for _, d in corpus]))

    linhas = ["# diag_enderecos -- SP-M Fase 1 (levantamento)", ""]

    for path in resolvidos:
        linhas.append(f"## {path}")
        try:
            wb = _abrir(path)
        except Exception as e:
            linhas.append(f"  !! não pôde ser lido: {e!r}")
            linhas.append("")
            continue

        try:
            rota = classificar(wb, override="auto", config=cfg)
        except Exception as e:
            linhas.append(f"  !! classificar() falhou: {e!r}")
            wb.close()
            linhas.append("")
            continue

        linhas.append(f"  homogeneo={rota.homogeneo} sheets_dados={rota.sheets_dados}")

        for sn in rota.sheets_dados:
            rows = ler_rows(wb[sn])
            header_homog = detectar_header(rows) if rota.homogeneo else None
            if header_homog is not None:
                header = rows[header_homog]
                col_idx = next(
                    (i for i, v in enumerate(header) if _norm(v).replace("Í", "I") == "INDEX DNP3"),
                    None,
                )
                linhas.append(
                    f"  [{sn}] rota=HOMOGENEO header_row={header_homog + 1} "
                    f"col_indice={col_idx} (via header exato 'INDEX DNP3')"
                )
                continue

            try:
                mapa = analisar(rows, enc, ref, siglas_set=lp.siglas)
            except Exception as e:
                linhas.append(f"  [{sn}] !! analisar() falhou: {e!r}")
                continue

            h = _header_por_densidade(rows)
            inicio = h + 1
            ncols = _ncols(rows)
            col_idx = mapa.colunas.get("indice")
            label = _header_label(rows, mapa.header_row, col_idx)
            linhas.append(
                f"  [{sn}] rota=NAO-HOMOGENEO header_row={mapa.header_row} "
                f"col_indice={col_idx} label={label!r} cols={mapa.colunas}"
            )

            concorrentes = _colunas_numericas_confundiveis(rows, inicio, ncols, col_idx)
            if concorrentes:
                for c in concorrentes[:5]:
                    clabel = _header_label(rows, mapa.header_row, c["col"])
                    linhas.append(
                        f"      concorrente col={c['col']} label={clabel!r} "
                        f"score={c['score']} frac_int={c['frac_inteiros']} "
                        f"mono={c['mono']} faixa=[{c['min']},{c['max']}] n={c['n']}"
                    )
            else:
                linhas.append("      (nenhuma coluna concorrente >=50% inteiros)")

        wb.close()
        linhas.append("")

    out = Path("bench/resultados/diag_enderecos.log")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(linhas), encoding="utf-8")
    print(f"log em {out}")


if __name__ == "__main__":
    main(sys.argv)
