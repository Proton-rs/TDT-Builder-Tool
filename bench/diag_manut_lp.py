"""Diagnóstico 2E (SP-OBS-17JUL): mede se as sheets MANUT_DiscreteSignals /
Manut_AnalogSignals da lista padrão (~33 linhas sigla+descrição) valeriam
como 4ª fonte de catálogo -- SEM incluir nada ainda (condicional).

Uso: PYTHONPATH=src python bench/diag_manut_lp.py

Para cada lista real (mesmas 5 do diag_ancora_revisao / P5), roda o pipeline
e, para cada registro que hoje termina em revisão (status != "decidido") OU
tem aparência de falso-positivo (decidido mas com score do candidato
vencedor baixo, heurística <0.6), checa se alguma sigla MANUT bate por
âncora exata no texto normalizado OU fuzzy >90 na descrição MANUT contra a
descrição do registro. Imprime os casos e decide, pela regra do plano:
>=5 casos reais -> recomenda abrir task de inclusão (não implementa aqui);
<5 -> "medida, não incluída", grava a contagem em resultados/spOBS_2E.txt.
"""
from __future__ import annotations

import warnings
import logging
from pathlib import Path

warnings.simplefilter("ignore")
logging.disable(logging.CRITICAL)

import openpyxl
from rapidfuzz import fuzz

from tdt.config import Config
from tdt.dados.encoder import criar_encoder
from tdt.dados.lista_padrao import ListaPadraoADMS
from tdt.pipeline import executar

_TEMPLATE = "docs/dnp3_template.xlsx"
_LISTA_PADRAO = "docs/Pontos Padrao ADMS_v8.xlsx"

_LISTAS_REAIS = [
    ("GTD", "docs/input_nao_homogeneo_1_GTA.xlsx"),
    ("IMA", "docs/input_homogeneo_IMA.xlsx"),
    ("FWB", "docs/input_nao_homogeneo_2_FWB.xlsx"),
    ("GPR", "docs/input_nao_homogeneo_3_GPR.xlsx"),
    ("GAU", "docs/input_nao_homogeneo_4_GAU.xlsx"),
]

_OUT = Path("bench/resultados/spOBS_2E.txt")

_SIGLA_MANUT_SHEETS = ("MANUT_DiscreteSignals", "Manut_AnalogSignals")


def carregar_manut(path: str) -> list[tuple[str, str]]:
    """Lê (sigla, descrição) das duas sheets MANUT. Colunas SINAL / DESCRIÇÃO
    NOVA na posição 0/1 (mesma ordem das sheets DiscreteSignals/AnalogSignals
    -- header vem com encoding corrompido no arquivo fonte, por isso posição
    em vez de nome de coluna, igual ao padrão de ``lista_padrao._ler_sheet``
    para as demais sheets)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for nome in _SIGLA_MANUT_SHEETS:
            if nome not in wb.sheetnames:
                raise ValueError(f"sheet {nome} ausente em {path}")
        sinais: list[tuple[str, str]] = []
        for nome in _SIGLA_MANUT_SHEETS:
            ws = wb[nome]
            linhas = ws.iter_rows(values_only=True)
            next(linhas)  # header
            for row in linhas:
                sigla = str(row[0]).strip() if row and row[0] not in (None, "") else ""
                if not sigla:
                    continue
                desc = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else ""
                sinais.append((sigla, desc))
    finally:
        wb.close()
    return sinais


def _fp_heuristico(rec) -> bool:
    """Decidido mas com aparência de falso-positivo: candidato vencedor com
    score baixo (heurística <0.6, sem gate de produção correspondente --
    só para dar cobertura a decisões fracas além dos itens de revisão)."""
    if rec.status != "decidido" or not rec.candidatos:
        return False
    melhor = max(c.score for c in rec.candidatos)
    return melhor < 0.6


def _bate_manut(texto_normalizado: str, descricao_bruta: str, manut: list[tuple[str, str]]):
    """Devolve lista de (sigla_manut, tipo_match, valor) para as siglas MANUT
    que casam com o registro -- âncora exata (token) OU fuzzy >90 na
    descrição."""
    tokens = texto_normalizado.upper().split()
    achados = []
    for sigla, desc in manut:
        sigla_u = sigla.upper()
        if sigla_u in tokens:
            achados.append((sigla, "ancora_exata", sigla_u))
            continue
        if desc:
            score = fuzz.token_sort_ratio(desc.upper(), descricao_bruta.upper())
            if score > 90:
                achados.append((sigla, "fuzzy_desc", round(score, 1)))
    return achados


def main() -> None:
    manut = carregar_manut(_LISTA_PADRAO)
    catalogo_primario = ListaPadraoADMS.carregar(_LISTA_PADRAO).siglas
    antes = len(manut)
    manut = [(s, d) for s, d in manut if s.upper() not in catalogo_primario]
    excluidas = antes - len(manut)
    print(
        f"MANUT: {antes} sinais carregados de {_LISTA_PADRAO}; "
        f"{excluidas} ja presentes no catalogo primario (excluidas); "
        f"{len(manut)} exclusivas restantes\n"
    )

    cfg = Config()
    enc = criar_encoder(cfg.modelo_embedding)

    casos: list[dict] = []
    for nome, path in _LISTAS_REAIS:
        try:
            resultado, _wb = executar(
                path, _TEMPLATE, _LISTA_PADRAO,
                config=cfg, encoder=enc, subestacao=nome,
            )
        except Exception as e:
            print(f"!! {nome} ({path}): pipeline falhou: {e!r}")
            continue

        alvos = []
        for item in resultado.revisao:
            if item.registro.status != "decidido":
                alvos.append((item.registro, item.motivo))
        for rec in resultado.lista.registros:
            if _fp_heuristico(rec):
                alvos.append((rec, "fp_heuristico_score<0.6"))

        for rec, motivo in alvos:
            achados = _bate_manut(
                rec.descricoes.normalizada, rec.descricoes.bruta, manut,
            )
            if achados:
                casos.append({
                    "lista": nome,
                    "id": rec.id,
                    "motivo": motivo,
                    "descricao": rec.descricoes.bruta,
                    "achados": achados,
                })

    print(f"casos onde uma sigla/descrição MANUT bateria: {len(casos)}\n")
    for c in casos:
        print(
            f"[{c['lista']}] id={c['id']} motivo={c['motivo']}\n"
            f"  desc={c['descricao']!r}\n"
            f"  achados={c['achados']}\n"
        )

    n = len(casos)
    if n >= 5:
        conclusao = (
            f"2E: {n} casos reais (>=5) -> RECOMENDA abrir task de inclusão "
            f"MANUT como 4a fonte do catalogo (flag origem='manut', gate "
            f"individual), fora do escopo desta task -- so recomendacao."
        )
    else:
        conclusao = f"2E: {n} casos reais (<5) -> medida, NAO incluida (licao v5: mais candidatos = diluicao)."
    print(conclusao)

    _OUT.parent.mkdir(parents=True, exist_ok=True)
    _OUT.write_text(f"{conclusao}\ncontagem={n}\n", encoding="utf-8")
    print(f"\ncontagem gravada em {_OUT}")


if __name__ == "__main__":
    main()
