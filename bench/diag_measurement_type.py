"""Diagnóstico 2G (SP-OBS-17JUL): compara a sheet ``DMS Signal Explanation``
da lista padrão contra o mapa hardcoded ``_MEASUREMENT_TYPE_PT_EN``
(engine_tdt.py) -- condicional (só altera o mapa se houver tipo faltante
com uso real confirmado).

Uso: PYTHONPATH=src python bench/diag_measurement_type.py

Nota sobre a sheet: o header declara colunas ``MatchingString`` (0),
``DMS Matching template`` (1), `` signalType`` (2), ``measurementType`` (3),
``phaseCode & sideMeasured`` (4) -- mas os dados reais das 1141 linhas vêm
deslocados uma coluna para a direita em relação ao header (col0 sempre
vazia; o "MatchingString" de cada linha está em col1). O script lê pela
posição populada (col1..col4), não pelo nome do header.

A sheet documenta o enum COMPLETO de ``measurementType`` do
DMSMatchingTemplateInfo (39 valores em uso nas 1139 linhas de dados, ~40 no
domínio) -- é uma referência ampla de todo o ADMS (inclui clima, financeiro,
etc.), não o vocabulário específico desta lista padrão. Por isso o
diagnóstico cruza em DOIS níveis:

1. sheet (measurementType, universo ADMS) x valores (lado EN) do mapa
   hardcoded -- mostra o que a sheet cobre a mais (universo amplo).
2. AnalogSignals!TIPO DE MEDIÇÃO (lado PT real, a fonte de
   ``sp.tipo_medicao`` que o engine de fato consome) x chaves do mapa --
   mostra o que é ACIONÁVEL (vocabulário PT realmente usado na lista
   padrão).
"""
from __future__ import annotations

import sys
import warnings
import logging
from pathlib import Path

warnings.simplefilter("ignore")
logging.disable(logging.CRITICAL)

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import openpyxl

from tdt.engine_tdt import _MEASUREMENT_TYPE_PT_EN

_LISTA_PADRAO = "docs/Pontos Padrao ADMS_v8.xlsx"
_SHEET_EXPLANATION = "DMS Signal Explanation"
_SHEET_ANALOG = "AnalogSignals"


def ler_measurement_types_sheet(path: str) -> set[str]:
    """Valores únicos de measurementType (lado EN) usados nas 1139 linhas de
    dados da DMS Signal Explanation (col3, populada quando col1/MatchingString
    não é None)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_SHEET_EXPLANATION]
        tipos: set[str] = set()
        linhas = ws.iter_rows(values_only=True)
        next(linhas)  # header
        for row in linhas:
            if row[1] is not None and row[3]:
                tipos.add(str(row[3]).strip())
    finally:
        wb.close()
    return tipos


def ler_tipo_medicao_pt_real(path: str) -> set[str]:
    """Valores únicos (PT, upper) da coluna TIPO DE MEDIÇÃO em AnalogSignals
    -- fonte real de ``sp.tipo_medicao`` consumida pelo engine."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_SHEET_ANALOG]
        linhas = ws.iter_rows(values_only=True)
        header = next(linhas)
        idx = None
        for i, h in enumerate(header):
            if h and "MEDI" in str(h).upper():
                idx = i
                break
        if idx is None:
            raise ValueError("coluna TIPO DE MEDIÇÃO não encontrada em AnalogSignals")
        tipos: set[str] = set()
        for row in linhas:
            v = row[idx] if idx < len(row) else None
            if v:
                tipos.add(str(v).strip().upper())
    finally:
        wb.close()
    return tipos


def main() -> None:
    tipos_sheet_en = ler_measurement_types_sheet(_LISTA_PADRAO)
    tipos_pt_real = ler_tipo_medicao_pt_real(_LISTA_PADRAO)

    chaves_mapa = set(_MEASUREMENT_TYPE_PT_EN.keys())
    valores_mapa = set(_MEASUREMENT_TYPE_PT_EN.values())

    print(f"mapa hardcoded: {len(_MEASUREMENT_TYPE_PT_EN)} entradas")
    print(f"DMS Signal Explanation: {len(tipos_sheet_en)} measurementType (EN) distintos em uso")
    print(f"AnalogSignals!TIPO DE MEDIÇÃO: {len(tipos_pt_real)} valores PT distintos\n")

    print("=== (a) divergências: valor EN do mapa hardcoded ausente do universo da sheet ===")
    divergencias = valores_mapa - tipos_sheet_en
    if divergencias:
        for v in sorted(divergencias):
            chave = [k for k, val in _MEASUREMENT_TYPE_PT_EN.items() if val == v]
            print(f"  {v!r} (chave PT: {chave}) -- NAO existe no domínio measurementType da sheet")
    else:
        print("  nenhuma -- todos os 12 valores EN do mapa existem no domínio da sheet")

    print("\n=== (b) tipos EN da sheet sem entrada no mapa hardcoded (universo ADMS amplo) ===")
    faltantes_en = sorted(tipos_sheet_en - valores_mapa)
    print(f"  {len(faltantes_en)} tipos: {faltantes_en}")

    print("\n=== (b-real) tipos PT de AnalogSignals sem chave no mapa hardcoded (ACIONÁVEL) ===")
    faltantes_pt = sorted(tipos_pt_real - chaves_mapa)
    if faltantes_pt:
        print(f"  {len(faltantes_pt)} tipos PT SEM tradução: {faltantes_pt}")
    else:
        print("  nenhum -- os 12 valores PT usados de fato em AnalogSignals já têm chave no mapa")

    print("\n=== caso-teste KMDF (observações item 16) ===")
    print("  KMDF (DISTANCIA DEFEITO) -> tipo_medicao PT = 'COMPRIMENTO'")
    if "COMPRIMENTO" in chaves_mapa:
        print(f"  'COMPRIMENTO' já mapeado -> {_MEASUREMENT_TYPE_PT_EN['COMPRIMENTO']!r} (resolvido em SP-Pendencias-09jul)")
    else:
        print("  'COMPRIMENTO' AUSENTE do mapa -- ação necessária")

    if faltantes_pt:
        conclusao = f"2G: {len(faltantes_pt)} tipo(s) PT real(is) sem tradução -> ADICIONAR ao mapa."
    else:
        conclusao = "2G: mapa hardcoded já cobre 100% do vocabulário PT real (AnalogSignals); sheet é só referência ampla do ADMS -- nada a adicionar."
    print(f"\n{conclusao}")


if __name__ == "__main__":
    main()
