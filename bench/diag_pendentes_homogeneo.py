"""Lista as siglas do input homogeneo real que caem em pendentes (lp.por_sigla == None),
classifica cada uma, e checa variantes de grafia contra a lista padrao.

Uso: PYTHONPATH=src python bench/diag_pendentes_homogeneo.py [input.xlsx] [lista_padrao.xlsx]
Default: docs/input_homogeneo_IMA.xlsx  docs/Pontos Padrao ADMS_v6.xlsx
"""
import sys
from collections import Counter

# openpyxl py3.14 shim: alguns .xlsx trazem PatternFill extLst nao suportado
import openpyxl.styles.fills as _fills
_orig_init = _fills.PatternFill.__init__
def _patched(self, *a, extLst=None, **kw):
    _orig_init(self, *a, **kw)
_fills.PatternFill.__init__ = _patched

import openpyxl
from rapidfuzz import fuzz

from tdt.dados.lista_padrao import ListaPadraoADMS
from tdt.normalizacao.estruturador_homogeneo import detectar_header, _col, _normaliza_celula

FASE3 = {"TAP", "COMTAP"}


def main():
    inp = sys.argv[1] if len(sys.argv) > 1 else "docs/input_homogeneo_IMA.xlsx"
    lp_path = sys.argv[2] if len(sys.argv) > 2 else "docs/Pontos Padrao ADMS_v6.xlsx"
    lp = ListaPadraoADMS.carregar(lp_path)
    siglas_padrao = sorted(lp.siglas)

    wb = openpyxl.load_workbook(inp, read_only=True, data_only=True)
    pend = Counter()
    exemplo: dict[str, str] = {}
    for sn in wb.sheetnames:
        rows = [tuple(r) for r in wb[sn].iter_rows(values_only=True)]
        h = detectar_header(rows)
        if h is None:
            continue
        i_sig = _col(rows[h], "SIGLA SINAL")
        i_uso = _col(rows[h], "UTILIZADO?")
        i_desc = _col(rows[h], "DESCRICAO DO PONTO")
        for row in rows[h + 1:]:
            if i_uso is None or i_uso >= len(row) or _normaliza_celula(row[i_uso]) != "SIM":
                continue
            s = str(row[i_sig] or "").strip() if i_sig is not None else ""
            if s and lp.por_sigla(s) is None:
                pend[s] += 1
                if s not in exemplo:
                    exemplo[s] = str(row[i_desc] or "")[:44] if i_desc is not None else ""
    wb.close()

    print(f"input: {inp}")
    print(f"lista padrao: {lp_path}")
    print(f"siglas distintas pendentes: {len(pend)} | ocorrencias: {sum(pend.values())}\n")
    print(f"{'sigla':<10}{'n':>4}  {'classe':<28} exemplo")
    for s, n in pend.most_common():
        # (b) checagem: variante de grafia proxima de alguma sigla padrao?
        melhor = max(((fuzz.token_sort_ratio(s.upper(), p), p) for p in siglas_padrao),
                     default=(0, ""))
        if s.upper() in FASE3:
            classe = "(a) resolvido Fase 3"
        elif melhor[0] >= 85:
            classe = f"(b?) ~{melhor[1]} r={melhor[0]:.0f}"
        else:
            classe = "(a) faltando na lista v6"
        print(f"{s:<10}{n:>4}  {classe:<28} {exemplo[s]}")


if __name__ == "__main__":
    main()
