"""Compara o TDT gerado com o TDT real por endereço DNP3 (INCOORDS).

A descrição bruta não existe no TDT real (col Description vazia); a única
chave estável entre os dois é o endereço (Input Coordinates, col 31 0-based).
Sigla = último token após '_' do Signal Name (col 0). Ver spec
docs/superpowers/specs/2026-07-01-sp-discriminador-generico-design.md.
"""
from __future__ import annotations

from dataclasses import dataclass

import openpyxl

# Input Coordinates 0-based por sheet — índices REAIS, medidos no template e no
# export real: DiscreteSignals=31, AnalogSignals=47, DiscreteAnalog=34. A chave
# de comparação é (sheet, endereço), não só o endereço: discreto/analógico/
# discreteanalog compartilham o keyspace de endereços DNP3 (todos começam em 0),
# então chavear só por int faria um discreto colidir com um analógico de mesmo
# índice. Chavear por (sheet, addr) compara cada categoria isoladamente e inclui
# os analógicos (antes silenciosamente excluídos por lerem a coluna 31 errada).
_SHEETS_INCOORDS = {
    "DNP3_DiscreteSignals": 31,
    "DNP3_AnalogSignals": 47,
    "DNP3_DiscreteAnalog": 34,
}
_COL_NOME = 0
_PRIMEIRA_LINHA_DADOS = 5  # 1-based; rows 1-4 são cabeçalho


def carregar_siglas_por_endereco(caminho: str) -> dict[tuple[str, int], tuple[str, str]]:
    """Devolve ``{(sheet, endereço): (nome, sigla)}``.

    Chave composta por (sheet, endereço) — ver nota em ``_SHEETS_INCOORDS``:
    endereços DNP3 não são únicos entre categorias, só dentro de cada sheet.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    out: dict[tuple[str, int], tuple[str, str]] = {}
    for sn, col_ic in _SHEETS_INCOORDS.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in ws.iter_rows(min_row=_PRIMEIRA_LINHA_DADOS, values_only=True):
            nome = r[_COL_NOME] if len(r) > _COL_NOME else None
            addr = r[col_ic] if len(r) > col_ic else None
            if not nome or not isinstance(addr, int):
                continue
            sigla = str(nome).split("_")[-1]
            out[(sn, addr)] = (str(nome), sigla)
    wb.close()
    return out


@dataclass(frozen=True)
class Resultado:
    comum: int
    iguais: int
    pct: float
    divergencias: list[tuple[int, str, str, str]]  # addr, real, nosso, nome_real


def comparar(nosso: str, real: str) -> Resultado:
    d_nosso = carregar_siglas_por_endereco(nosso)
    d_real = carregar_siglas_por_endereco(real)
    comuns = sorted(set(d_nosso) & set(d_real))
    iguais = 0
    divergencias: list[tuple[int, str, str, str]] = []
    for chave in comuns:
        _sheet, addr = chave
        nome_real, sig_real = d_real[chave]
        _, sig_nosso = d_nosso[chave]
        if sig_real.upper() == sig_nosso.upper():
            iguais += 1
        else:
            divergencias.append((addr, sig_real, sig_nosso, nome_real))
    pct = 100.0 * iguais / len(comuns) if comuns else 0.0
    return Resultado(len(comuns), iguais, pct, divergencias)
