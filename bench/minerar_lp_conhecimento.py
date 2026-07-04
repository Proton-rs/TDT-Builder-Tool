"""Gera o esqueleto de docs/conhecimento_sinais.md a partir da LP.

Por família: siglas, descrição-padrão, tipo ADMS, direção (Read/Write/RW),
estados (message mapping ...@...___...@...), grupo/categoria, DE->PARA.

Fonte de verdade: docs/Pontos Padrao ADMS_v2.xlsx (mesma LP usada por
src/tdt/dados/lista_padrao.py e por todos os benchmarks). Agrupamento por
família reusa a MESMA lógica de src/tdt/motor_regras._numero_lider (prefixo
ANSI de 2 dígitos). Siglas sem prefixo ANSI numérico são agrupadas por
prefixo alfabético de 2 letras (fallback documentado); o que sobrar vai
para "Outras/Não classificadas" — nenhuma sigla é descartada em silêncio.

Uso: PYTHONPATH=src python bench/minerar_lp_conhecimento.py > docs/conhecimento_sinais.md
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path
from typing import NamedTuple

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from tdt.motor_regras import _numero_lider  # reuso da lógica oficial de agrupamento

_ROOT = Path(__file__).resolve().parent.parent
_LP_PATH = _ROOT / "docs" / "Pontos Padrao ADMS_v2.xlsx"

_INVALIDOS = {None, "", "#N/A", "NONE", "NULL"}


def _val(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return None if s.upper() in _INVALIDOS else s


class Sinal(NamedTuple):
    sigla: str
    descricao: str
    signal_type: str
    direction: str | None
    categoria: str  # "Discrete" | "Analog"
    mm: str | None  # chave de message mapping (Discrete)
    tipo_medicao: str | None  # Analog
    unidade: str | None  # Analog
    grupo_severidade: str | None  # TYPE SEVERIDADE (Discrete)


def _coluna(header: tuple, nome: str) -> int | None:
    alvo = nome.strip().upper()
    for i, h in enumerate(header):
        if h is not None and str(h).strip().upper() == alvo:
            return i
    return None


def _ler_discretos(ws) -> list[Sinal]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_sigla = _coluna(header, "SINAL")
    idx_desc = _coluna(header, "DESCRIÇÃO NOVA")
    idx_tipo = _coluna(header, "SIGNAL TYPE")
    idx_dir = _coluna(header, "DIRECTION")
    idx_mm = _coluna(header, "MM")
    idx_sev = _coluna(header, "TYPE SEVERIDADE")

    sinais = []
    for row in rows:
        sigla = _val(row[idx_sigla]) if idx_sigla is not None else None
        if sigla is None:
            continue
        sinais.append(
            Sinal(
                sigla=sigla,
                descricao=_val(row[idx_desc]) if idx_desc is not None else "" or "",
                signal_type=(_val(row[idx_tipo]) if idx_tipo is not None else None) or "",
                direction=_val(row[idx_dir]) if idx_dir is not None else None,
                categoria="Discrete",
                mm=_val(row[idx_mm]) if idx_mm is not None else None,
                tipo_medicao=None,
                unidade=None,
                grupo_severidade=_val(row[idx_sev]) if idx_sev is not None else None,
            )
        )
    return sinais


def _ler_analogicos(ws) -> list[Sinal]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_sigla = _coluna(header, "SINAL")
    idx_desc = _coluna(header, "DESCRIÇÃO NOVA")
    idx_tipo = _coluna(header, "SIGNAL TYPE")
    idx_dir = _coluna(header, "DIREÇÃO DO FLUXO")
    idx_medicao = _coluna(header, "TIPO DE MEDIÇÃO")
    idx_unid = _coluna(header, "UNIDADE DE EXIBIÇÃO")

    sinais = []
    for row in rows:
        sigla = _val(row[idx_sigla]) if idx_sigla is not None else None
        if sigla is None:
            continue
        sinais.append(
            Sinal(
                sigla=sigla,
                descricao=(_val(row[idx_desc]) if idx_desc is not None else None) or "",
                signal_type=(_val(row[idx_tipo]) if idx_tipo is not None else None) or "",
                direction=_val(row[idx_dir]) if idx_dir is not None else None,
                categoria="Analog",
                mm=None,
                tipo_medicao=_val(row[idx_medicao]) if idx_medicao is not None else None,
                unidade=_val(row[idx_unid]) if idx_unid is not None else None,
                grupo_severidade=None,
            )
        )
    return sinais


def _ler_de_para(ws) -> dict[str, str]:
    """Sheet 'DE->PARA': coluna SINAL = sigla antiga, DESCRIÇÃO NOVA = sigla nova
    (nomes de coluna reaproveitados de outras sheets; aqui é renomeação de sigla)."""
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    mapa = {}
    for row in rows:
        antigo = _val(row[0])
        novo = _val(row[1]) if len(row) > 1 else None
        if antigo is not None and novo is not None:
            mapa[str(antigo).strip().upper()] = str(novo).strip()
    return mapa


def _ler_message_mapping(ws) -> dict[str, str]:
    """Sheet 'Message Mapping': Name (com prefixo DI_) -> State (Message).
    Chave normalizada removendo o prefixo 'DI_' para casar com a coluna MM
    de DiscreteSignals (formato ...@...___...@...___...)."""
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_name = _coluna(header, "Name")
    idx_state = _coluna(header, "State (Message)")
    mapa = {}
    for row in rows:
        name = _val(row[idx_name]) if idx_name is not None else None
        state = _val(row[idx_state]) if idx_state is not None else None
        if name is None:
            continue
        chave = name.strip()
        if chave.upper().startswith("DI_"):
            chave = chave[3:]
        mapa[chave] = state or ""
    return mapa


def _alpha_prefix(sigla: str, n: int = 2) -> str | None:
    letras = ""
    for ch in sigla:
        if ch.isalpha():
            letras += ch
        else:
            break
    return letras[:n] if letras else None


def _familia(sigla: str) -> str:
    """Chave de agrupamento: prefixo ANSI (via _numero_lider oficial) senão
    prefixo alfabético de 2 letras senão 'OUTRAS'."""
    lider = _numero_lider(sigla.upper())
    if lider is not None:
        return lider
    alpha = _alpha_prefix(sigla.upper())
    if alpha is not None:
        return f"ALPHA_{alpha}"
    return "OUTRAS"


def _titulo_familia(chave: str) -> str:
    if chave == "OUTRAS":
        return "Outras / Não classificadas"
    if chave.startswith("ALPHA_"):
        return f"Sigla {chave[len('ALPHA_'):]}*"
    return f"Família ANSI {chave}"


def _fmt(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("|", "\\|").replace("\n", " ").strip()
    return s


def _fmt_estados(mm: str | None, message_mapping: dict[str, str]) -> str:
    if not mm:
        return ""
    estado = message_mapping.get(mm)
    if estado:
        return _fmt(estado)
    return _fmt(mm)


def main() -> None:
    wb = openpyxl.load_workbook(_LP_PATH, read_only=True, data_only=True)
    try:
        discretos = _ler_discretos(wb["DiscreteSignals"])
        analogicos = _ler_analogicos(wb["AnalogSignals"])
        de_para = _ler_de_para(wb["DE->PARA"]) if "DE->PARA" in wb.sheetnames else {}
        message_mapping = (
            _ler_message_mapping(wb["Message Mapping"]) if "Message Mapping" in wb.sheetnames else {}
        )
    finally:
        wb.close()

    todos = discretos + analogicos

    familias: dict[str, list[Sinal]] = defaultdict(list)
    for s in todos:
        familias[_familia(s.sigla)].append(s)

    def ordem_familia(chave: str):
        if chave == "OUTRAS":
            return (2, chave)
        if chave.startswith("ALPHA_"):
            return (1, chave)
        return (0, chave.zfill(2))

    print("# Base de conhecimento de sinais — Lista Padrão ADMS")
    print()
    print(
        "Documento gerado automaticamente por `bench/minerar_lp_conhecimento.py` a partir de"
        f" `{_LP_PATH.relative_to(_ROOT).as_posix()}` (sheets `DiscreteSignals`, `AnalogSignals`,"
        " `DE->PARA`, `Message Mapping`)."
    )
    print()
    print(
        "Agrupamento por família reusa `tdt.motor_regras._numero_lider` (prefixo ANSI de"
        " 2 dígitos, ex.: `67N` -> `67`). Siglas sem prefixo ANSI numérico são agrupadas por"
        " prefixo alfabético de 2 letras (`ALPHA_xx`); o restante cai em"
        " **Outras / Não classificadas** — nenhuma sigla é descartada."
    )
    print()
    print(
        "Este é o **esqueleto** (Task 1 do plano SP-L): dados minerados diretamente da LP,"
        " sem curadoria de conteúdo além de formatação. Tasks 2/3 adicionam semântica de"
        " domínio, exemplos e cross-references."
    )
    print()
    print("---")
    print()

    total_cobertas = 0
    for chave in sorted(familias.keys(), key=ordem_familia):
        sinais = sorted(familias[chave], key=lambda s: s.sigla.upper())
        total_cobertas += len(sinais)
        print(f"## {_titulo_familia(chave)}")
        print()
        print(f"_{len(sinais)} sigla(s)._")
        print()
        print(
            "| Sigla | Descrição padrão | Tipo ADMS | Categoria | Direção | Estados / MM |"
            " Tipo medição | Unidade | DE->PARA |"
        )
        print("|---|---|---|---|---|---|---|---|---|")
        for s in sinais:
            de_para_str = de_para.get(s.sigla.upper(), "")
            estados_str = _fmt_estados(s.mm, message_mapping) if s.categoria == "Discrete" else ""
            print(
                f"| {_fmt(s.sigla)} | {_fmt(s.descricao)} | {_fmt(s.signal_type)} |"
                f" {_fmt(s.categoria)} | {_fmt(s.direction)} | {estados_str} |"
                f" {_fmt(s.tipo_medicao)} | {_fmt(s.unidade)} | {_fmt(de_para_str)} |"
            )
        print()

    print("---")
    print()
    print("## Cobertura")
    print()
    print(f"- Total de siglas na LP (linhas válidas, Discrete + Analog): **{len(todos)}**")
    print(f"- Total de siglas cobertas neste documento: **{total_cobertas}**")
    print(f"- Discrete: {len(discretos)} | Analog: {len(analogicos)}")
    print(f"- Famílias geradas: {len(familias)}")
    print()
    if total_cobertas == len(todos):
        print("Cobertura: OK — todas as siglas da LP aparecem em alguma família acima.")
    else:
        print(
            "Cobertura: DIVERGENTE — verificar script "
            "(alguma sigla da LP não apareceu em nenhuma família)."
        )


if __name__ == "__main__":
    main()
