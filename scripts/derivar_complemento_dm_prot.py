"""Deriva COMPLEMENTO_DM_PROT (spec 2026-07-20 §B1): siglas da lista padrão
com SIGNAL TYPE != RelayTrip que a fullbase mapeia consistentemente em
dispositivo PROT. Critério: nomes estilo-SE, linhas limpas (prefixo SE do
sinal == prefixo SE do DM), >=90% PROT, n>=20. Rodar e colar a saída em
tdt/defaults.py quando a lista padrão ou a fullbase mudarem."""
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

DOCS = Path(__file__).resolve().parents[1] / "docs"
FULLBASE = DOCS / "Export_base_Full__27_fev_2026.xlsx"
LISTA = DOCS / "Pontos Padrao ADMS_v8.xlsx"
RE_SE = re.compile(r"^([A-Z]{2,5})_")

# siglas nao-RelayTrip da lista padrao
wb = load_workbook(LISTA, read_only=True, data_only=True)
ws = wb["DiscreteSignals"]
rows = ws.iter_rows(values_only=True)
hdr = [str(h).strip() if h else "" for h in next(rows)]
i_s, i_st = hdr.index("SINAL"), hdr.index("SIGNAL TYPE")
nao_relaytrip = {
    str(r[i_s]).strip().upper()
    for r in rows
    if r[i_s] is not None and str(r[i_st] or "").strip() != "RelayTrip"
}

# %PROT por sigla na fullbase (linhas limpas)
wb = load_workbook(FULLBASE, read_only=True, data_only=True)
ws = wb["DNP3_DiscreteSignals"]
hdr = [str(h).strip() if h else "" for h in
       next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
i_sn, i_dm = hdr.index("Signal Name"), hdr.index("Device Mapping")
stats: dict[str, list[int]] = defaultdict(lambda: [0, 0])  # sigla -> [tot, prot]
for row in ws.iter_rows(min_row=5, values_only=True):
    sn = row[i_sn]
    if sn is None:
        continue
    sn, dm = str(sn), str(row[i_dm] or "").strip()
    m_sn, m_dm = RE_SE.match(sn), RE_SE.match(dm)
    if not m_sn or not m_dm or m_sn.group(1) != m_dm.group(1):
        continue  # religador numerico ou linha suja (SE trocada)
    sigla = sn.rsplit("_", 1)[-1].upper()
    stats[sigla][0] += 1
    if "PROT" in dm.upper():
        stats[sigla][1] += 1

aprovadas = sorted(
    s for s, (tot, prot) in stats.items()
    if s in nao_relaytrip and tot >= 20 and prot / tot >= 0.90
)
for s in aprovadas:
    tot, prot = stats[s]
    print(f"# {s}: {prot}/{tot} PROT ({100 * prot / tot:.0f}%)")
print("COMPLEMENTO_DM_PROT = frozenset({")
print("    " + ", ".join(f'"{s}"' for s in aprovadas))
print("})")
