"""Gera a v5: copia a v2, reescreve DESCRIÇÃO NOVA append-only nas duas sheets,
e emite sidecar de conflitos + diff CSV. Não toca em nenhuma outra coluna/sheet."""
from __future__ import annotations

import csv
import shutil
from pathlib import Path

import openpyxl

from composer import enriquecer

DOCS = Path("docs")
V2 = DOCS / "Pontos Padrao ADMS_v2.xlsx"
V5 = DOCS / "Pontos Padrao ADMS_v5.xlsx"
SHEETS = ("DiscreteSignals", "AnalogSignals")
COL_SIGLA = 0          # 0-based
COL_DESC = 1           # DESCRIÇÃO NOVA é a coluna 1 (B)


def _achar_col_desc(ws) -> int:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for i, h in enumerate(header):
        if h is not None and "DESCRI" in str(h).upper():
            return i
    return COL_DESC


def aplicar(ws, sheet_nome: str) -> tuple[int, list[tuple]]:
    col = _achar_col_desc(ws)
    tocadas = 0
    conflitos: list[tuple] = []
    diffs: list[tuple] = []
    for r in range(2, ws.max_row + 1):
        sigla = ws.cell(r, COL_SIGLA + 1).value
        v1_raw = ws.cell(r, col + 1).value
        if sigla in (None, "") or v1_raw in (None, ""):
            continue
        v1_raw_s = str(v1_raw)
        v1s = v1_raw_s.strip()
        novo, conflito = enriquecer(v1s, sheet_nome, sigla=str(sigla).strip())
        if novo != v1s:
            # acréscimo calculado sobre o v1 normalizado (strip), mas escrito
            # sobre o valor RAW da célula para preservar literalmente espaços
            # iniciais/finais do v1 -> invariante startswith(v1_raw) garantida.
            acrescimo = novo[len(v1s):]
            novo_raw = v1_raw_s + acrescimo
            ws.cell(r, col + 1).value = novo_raw
            tocadas += 1
            diffs.append((str(sigla).strip(), sheet_nome, v1s, novo_raw, acrescimo.lstrip(" —")))
        if conflito is not None:
            conflitos.append((str(sigla).strip(), sheet_nome, conflito, v1s))
    aplicar.ultimo_diff = diffs  # type: ignore[attr-defined]
    return tocadas, conflitos


def main() -> None:
    shutil.copyfile(V2, V5)
    wb = openpyxl.load_workbook(V5)
    todos_conf: list[tuple] = []
    todos_diff: list[tuple] = []
    for sh in SHEETS:
        toc, conf = aplicar(wb[sh], sh)
        todos_conf += conf
        todos_diff += getattr(aplicar, "ultimo_diff", [])
        print(f"[{sh}] linhas enriquecidas: {toc}; conflitos: {len(conf)}")
    wb.save(V5)
    # sidecar de conflitos
    from ansi_ref import CONFLITO_V1
    with open(DOCS / "v5_conflitos_ansi.md", "w", encoding="utf-8") as f:
        f.write("# Conflitos v1 × ANSI (revisar; v1 preservado)\n\n")
        for sig, sheet, code, v1 in todos_conf:
            nota = CONFLITO_V1.get(code, "")
            f.write(f"- **{sig}** ({sheet}) ANSI {code}: {v1!r} — {nota}\n")
    # diff CSV
    with open(DOCS / "v5_diff_descricoes.csv", "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sigla", "sheet", "v1", "v5", "acrescentado"])
        w.writerows(todos_diff)
    print(f"v5 salvo em {V5}; conflitos={len(todos_conf)}; diffs={len(todos_diff)}")


if __name__ == "__main__":
    main()
