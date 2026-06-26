import openpyxl
from gerar_v5 import aplicar, COL_DESC, COL_SIGLA


def _ws():
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["SINAL", "DESCRIÇÃO NOVA"])  # header (col 0,1)
    ws.append(["50N", "50 - SOBRECORRENTE INSTANTANEA NEUTRO"])
    ws.append(["61T", "61 - DESEQUILIBRIO TEMPORIZADO"])
    ws.append(["CMDE", "PROPRIEDADE DO COMANDO"])
    return ws


def test_aplicar_append_only_em_todas_as_linhas():
    ws = _ws()
    originais = [ws.cell(r, COL_DESC + 1).value for r in range(2, ws.max_row + 1)]
    tocadas, conflitos = aplicar(ws, "DiscreteSignals")
    for i, r in enumerate(range(2, ws.max_row + 1)):
        novo = ws.cell(r, COL_DESC + 1).value
        assert novo.startswith(originais[i])   # INVARIANTE central: append-only
    # 50N foi enriquecido; 61T flagado como conflito
    assert any(sig == "61T" for sig, *_ in conflitos)


def test_50n_recebeu_funcao_ansi():
    ws = _ws()
    aplicar(ws, "DiscreteSignals")
    v = ws.cell(2, COL_DESC + 1).value
    assert "ANSI 50" in v and "INSTANT" in v.upper()
