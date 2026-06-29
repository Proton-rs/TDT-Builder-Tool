"""Gera a v6: enriquece DESCRIÇÃO NOVA com variantes reais mineradas dos
arquivos de auditoria (Status == "decidido"), append-only, por sigla.

Pipeline (spec docs/superpowers/specs/2026-06-26-lista-adms-v6-...md, seção
"Atualização 2026-06-28"):

1. Minerar pares (desc_bruta, sigla) dos `output/**/*Auditoria_Revisao*.xlsx`,
   só linhas com Status == "decidido" (pares hipotéticos de "revisao" são
   descartados nesta iteração — ruído de herdar sigla errada > ganho).
2. Carregar docs/Pontos Padrao ADMS_v2.xlsx (sigla -> desc_v2 canonizada).
3. Para cada sigla, tokenizar as descrições brutas com o MESMO pipeline de
   canonização do matcher (tdt.normalizacao.normalizador.canonizar) e reter só
   termos que: (a) não estão em desc_v2 da própria sigla; (b) não são
   infinitivo; (c) não são ID de equipamento; (d) não são endereço; (e) não
   são grafia/acento/maiúscula-only de um termo já em desc_v2; (f) não
   aparecem em desc_v2 de NENHUMA outra sigla (contra-checagem); (g) aparecem
   com frequência mínima entre as descrições da sigla (filtro extra — sem
   isso, outliers de classificação errada do pipeline antigo entram como
   ruído; ver relatório).
4. Gera docs/Pontos Padrao ADMS_v6.xlsx (cópia byte-a-byte da v2 via openpyxl,
   só a coluna DESCRIÇÃO NOVA de DiscreteSignals/AnalogSignals muda) e
   docs/v6_variantes_propostas.csv (curadoria).

Uso: PYTHONPATH=src python scripts/enriquecer_v6/gerar_v6.py
"""
from __future__ import annotations

import csv
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_ROOT / "src"))

from tdt.config import Config
from tdt.dados.lista_padrao import ListaPadraoADMS
from tdt.normalizacao.normalizador import canonizar

DOCS = _ROOT / "docs"
OUTPUT = _ROOT / "output"
V2 = DOCS / "Pontos Padrao ADMS_v2.xlsx"
V6 = DOCS / "Pontos Padrao ADMS_v6.xlsx"
CSV_OUT = DOCS / "v6_variantes_propostas.csv"
SHEETS = ("DiscreteSignals", "AnalogSignals")
CFG = Config()

# --- siglas espúrias conhecidas: pipeline antigo decidia errado e a curadoria
# (spec, "Atualização 2026-06-28") já identificou DJA1 como sigla que não
# existe na TDT real da GTD. Excluída por segurança — não enriquecer.
SIGLAS_EXCLUIDAS: frozenset[str] = frozenset({"DJA1"})

# Frequência mínima (absoluta) de um termo dentro das descrições de UMA sigla
# para ser aceito como variante. Sem isso, descrições isoladas mal-classificadas
# pelo pipeline antigo (ex.: "Mola Descarregada" caindo em VAB) entram como
# ruído. Calibrado: 2+ ocorrências já distingue padrão real de outlier único.
FREQ_MIN_ABS = 2
# Fração mínima das descrições distintas da sigla em que o termo aparece.
# Reforça o corte de outliers quando a sigla tem poucas descrições distintas.
FREQ_MIN_FRAC = 0.05

# --- filtros (Recipe de enriquecimento, spec) -------------------------------

# 3) infinitivos (verbos de ação/comando) — exclusão obrigatória. Particípios
# de estado (Desligado, Aberto, Excluído...) NÃO são infinitivo — entram pela
# exceção de "termos de estado".
_INFINITIVOS: frozenset[str] = frozenset({
    "DESLIGAR", "LIGAR", "ABRIR", "FECHAR", "BLOQUEAR", "DESBLOQUEAR",
    "INCLUIR", "EXCLUIR", "ATUAR", "DISPARAR", "RESETAR", "ALARMAR",
    "ATIVAR", "DESATIVAR", "ENERGIZAR", "DESENERGIZAR", "COMANDAR",
    "OPERAR", "ACIONAR", "DESACIONAR", "HABILITAR", "DESABILITAR",
    "RESTAURAR", "REARMAR", "TESTAR", "SUPERVISIONAR", "MONITORAR",
    "INDICAR", "SINALIZAR", "VERIFICAR", "AJUSTAR", "PARAMETRIZAR",
})

# 4) IDs de equipamento: [A-Z0-9]{2,4}\d* tipo 01Q0, 52-1 (após canonizar, o
# hífen já colapsou em espaço e os IDs letra-número tipo 01Q0 já foram
# removidos por separar_ids_equipamento/N2 do próprio canonizar). Mantém regex
# de reforço para IDs que sobrevivam à canonização (ex.: números puros de
# endereço de equipamento, tipo "52", "89", quando não acompanhados de função
# de proteção reconhecida).
_ID_EQUIPAMENTO = re.compile(r"^[A-Z]{1,2}\d{1,3}[A-Z]?\d*$")
_PURO_NUMERO = re.compile(r"^\d+$")

# 5) endereços: \d+[./]\d+
_ENDERECO = re.compile(r"^\d+[./]\d+$")

# Termos de estado — exceção: incluídos mesmo sendo "ação" porque indicam
# contexto semântico (grandeza em estado de alarme/bloqueio), não comando.
_TERMOS_ESTADO: frozenset[str] = frozenset({
    "ALARME", "BLOQUEIO", "ATIVO", "ABERTO", "ABERTA", "FECHADO", "FECHADA",
    "DESLIGADO", "DESLIGADA", "LIGADO", "LIGADA", "TRIP", "ENERGIZADO",
    "ENERGIZADA", "DESENERGIZADO", "DESENERGIZADA", "ATUADO", "ATUADA",
    "EXCLUIDO", "EXCLUIDA", "INCLUIDO", "INCLUIDA", "INDEFINIDA", "INDEFINIDO",
})


def _e_infinitivo(token: str) -> bool:
    if token in _TERMOS_ESTADO:
        return False
    return token in _INFINITIVOS


def _e_id_equipamento(token: str) -> bool:
    if token in _TERMOS_ESTADO:
        return False
    return bool(_ID_EQUIPAMENTO.match(token)) or bool(_PURO_NUMERO.match(token))


def _e_endereco(token: str) -> bool:
    return bool(_ENDERECO.match(token))


def _e_grafia_only(token: str, tokens_v2: frozenset[str]) -> bool:
    """Token é mera variação de grafia/acento/maiúscula de um termo já em
    desc_v2. canonizar() já remove acentos e normaliza para upper, então se o
    token sobreviveu ao canonizar igual a um termo de tokens_v2, já teria sido
    descartado no passo 1 (subtração direta). Esta função cobre o caso de
    pequenas variações morfológicas óbvias (plural simples) que o canonizar
    não unifica."""
    if token in tokens_v2:
        return True
    if token.endswith("S") and token[:-1] in tokens_v2:
        return True
    if (token + "S") in tokens_v2:
        return True
    return False


def _tokens_canonicos(desc: str) -> list[str]:
    return canonizar(desc, CFG).split()


# --- Fase 1: minerar pares dos arquivos de auditoria ------------------------

_PADRAO_AUDITORIA = "*Auditoria_Revisao*.xlsx"


def _achar_arquivos_auditoria() -> list[Path]:
    if not OUTPUT.exists():
        return []
    return sorted(OUTPUT.rglob(_PADRAO_AUDITORIA))


def minerar_pares(arquivos: list[Path]) -> dict[str, list[str]]:
    """sigla -> [desc_bruta, ...] (com repetição = frequência), só decididos."""
    por_sigla: dict[str, list[str]] = defaultdict(list)
    for path in arquivos:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb["Auditoria"] if "Auditoria" in wb.sheetnames else wb.worksheets[0]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
            i_status = idx.get("Status")
            i_sigla = idx.get("Sigla Decidida")
            i_desc = idx.get("Descrição Bruta")
            if i_status is None or i_sigla is None or i_desc is None:
                continue
            for row in rows:
                if row[i_status] != "decidido":
                    continue
                sigla = row[i_sigla]
                desc = row[i_desc]
                if not sigla or not desc:
                    continue
                sigla_s = str(sigla).strip()
                if not sigla_s:
                    continue
                por_sigla[sigla_s].append(str(desc).strip())
        finally:
            wb.close()
    return por_sigla


# --- Fase 2: extrair e filtrar variantes ------------------------------------


def construir_indices_v2(lp: ListaPadraoADMS) -> tuple[dict[str, frozenset[str]], dict[str, set[str]]]:
    """sigla -> tokens_v2 (canonizados); token -> {siglas que o têm} (cross-check)."""
    tokens_por_sigla: dict[str, frozenset[str]] = {}
    siglas_por_token: dict[str, set[str]] = defaultdict(set)
    for s in (*lp.discretos, *lp.analogicos):
        sigla = s.sigla.strip().upper()
        toks = frozenset(_tokens_canonicos(s.descricao))
        # sigla pode repetir (linhas duplicadas) — desc é idêntica nesses casos
        tokens_por_sigla[sigla] = tokens_por_sigla.get(sigla, frozenset()) | toks
        for t in toks:
            siglas_por_token[t].add(sigla)
    return tokens_por_sigla, siglas_por_token


def extrair_variantes(
    por_sigla: dict[str, list[str]],
    tokens_por_sigla: dict[str, frozenset[str]],
    siglas_por_token: dict[str, set[str]],
) -> dict[str, list[tuple[str, int]]]:
    """sigla -> [(termo, freq)] aprovados, ordenados por freq desc."""
    resultado: dict[str, list[tuple[str, int]]] = {}
    for sigla, descs in por_sigla.items():
        sigla_up = sigla.strip().upper()
        if sigla_up in SIGLAS_EXCLUIDAS:
            continue
        tokens_v2 = tokens_por_sigla.get(sigla_up)
        if tokens_v2 is None:
            continue  # sigla não existe na v2 (espúria/descontinuada) — pula

        n_descs_distintas = len(set(descs))
        contagem = Counter()
        for desc in descs:
            for tok in set(_tokens_canonicos(desc)):  # set: 1 ocorrência por desc
                contagem[tok] += 1

        aprovados: list[tuple[str, int]] = []
        for tok, freq in contagem.items():
            if tok in tokens_v2:
                continue  # já presente na desc_v2 dessa sigla
            if not tok or tok.isspace():
                continue
            if _e_infinitivo(tok):
                continue
            if _e_id_equipamento(tok):
                continue
            if _e_endereco(tok):
                continue
            if _e_grafia_only(tok, tokens_v2):
                continue
            # contra-checagem cross-sigla: termo não pode estar na desc_v2 de
            # NENHUMA outra sigla (evita canibalização)
            outras = siglas_por_token.get(tok, set()) - {sigla_up}
            if outras:
                continue
            # frequência mínima: filtra outliers de classificação errada
            frac = freq / n_descs_distintas if n_descs_distintas else 0
            if freq < FREQ_MIN_ABS and frac < FREQ_MIN_FRAC:
                continue
            aprovados.append((tok, freq))

        if aprovados:
            aprovados.sort(key=lambda x: (-x[1], x[0]))
            resultado[sigla_up] = aprovados
    return resultado


# --- Fase 3: gerar v6.xlsx ---------------------------------------------------


def _achar_col(ws, nome_alvo: str) -> int | None:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    alvo = nome_alvo.strip().upper()
    for i, h in enumerate(header):
        if h is not None and str(h).strip().upper() == alvo:
            return i
    return None


def aplicar_v6(wb, variantes: dict[str, list[tuple[str, int]]]) -> list[tuple]:
    """Escreve DESCRIÇÃO NOVA enriquecida nas sheets de matching. Devolve as
    linhas do diff CSV: (sigla, sheet, desc_v2, desc_v6, termos_novos, freq, origem)."""
    diffs: list[tuple] = []
    for sheet_nome in SHEETS:
        ws = wb[sheet_nome]
        col_sigla = _achar_col(ws, "SINAL")
        col_desc = _achar_col(ws, "DESCRIÇÃO NOVA")
        if col_sigla is None or col_desc is None:
            continue
        for r in range(2, ws.max_row + 1):
            sigla_cell = ws.cell(r, col_sigla + 1).value
            if sigla_cell is None or str(sigla_cell).strip() == "":
                continue
            sigla = str(sigla_cell).strip().upper()
            termos = variantes.get(sigla)
            if not termos:
                continue
            desc_raw = ws.cell(r, col_desc + 1).value
            if desc_raw is None or str(desc_raw).strip() == "":
                continue
            desc_v2 = str(desc_raw)
            termos_str = ", ".join(t for t, _ in termos)
            desc_v6 = f"{desc_v2} — {termos_str}"
            ws.cell(r, col_desc + 1).value = desc_v6
            freq_str = ";".join(str(f) for _, f in termos)
            diffs.append((sigla, sheet_nome, desc_v2, desc_v6, termos_str, freq_str, "decidido"))
    return diffs


def main() -> None:
    arquivos = _achar_arquivos_auditoria()
    print(f"arquivos de auditoria encontrados: {len(arquivos)}")
    for a in arquivos:
        print(f"  - {a.relative_to(_ROOT)}")

    por_sigla = minerar_pares(arquivos)
    total_pares = sum(len(v) for v in por_sigla.values())
    print(f"siglas com pares decididos: {len(por_sigla)} | total de pares minerados: {total_pares}")

    lp = ListaPadraoADMS.carregar(V2)
    tokens_por_sigla, siglas_por_token = construir_indices_v2(lp)

    variantes = extrair_variantes(por_sigla, tokens_por_sigla, siglas_por_token)
    total_termos = sum(len(v) for v in variantes.values())
    print(f"siglas enriquecidas: {len(variantes)} | total de termos aprovados: {total_termos}")

    shutil.copyfile(V2, V6)
    wb = openpyxl.load_workbook(V6)
    try:
        diffs = aplicar_v6(wb, variantes)
        wb.save(V6)
    finally:
        wb.close()
    print(f"v6 salvo em {V6.relative_to(_ROOT)}; linhas tocadas: {len(diffs)}")

    with open(CSV_OUT, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "sigla", "sheet", "desc_v2", "desc_v6", "termos_novos", "freq",
            "origem", "benchmark_impact",
        ])
        for sigla, sheet, desc_v2, desc_v6, termos, freq, origem in diffs:
            w.writerow([sigla, sheet, desc_v2, desc_v6, termos, freq, origem, ""])
    print(f"relatório de curadoria salvo em {CSV_OUT.relative_to(_ROOT)}")


if __name__ == "__main__":
    main()
