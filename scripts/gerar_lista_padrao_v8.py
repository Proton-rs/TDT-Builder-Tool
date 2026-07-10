"""Gera docs/Pontos Padrao ADMS_v8.xlsx a partir da v7 (spec 2026-07-10 §2.3)
e extrai o catálogo real de MMs para tests/fixtures/mm_catalogo_real.txt.

Só altera a coluna MM de 43LR e 81U1-5 na sheet DiscreteSignals; o resto do
arquivo (todas as sheets, formatação) é preservado pelo openpyxl.
"""
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
V7 = RAIZ / "docs" / "Pontos Padrao ADMS_v7.xlsx"
V8 = RAIZ / "docs" / "Pontos Padrao ADMS_v8.xlsx"
EXPORT = RAIZ / "docs" / "Export_base_Full__27_fev_2026.xlsx"
FIXTURE = RAIZ / "tests" / "fixtures" / "mm_catalogo_real.txt"

CORRECOES = {
    "43LR": "null@null___REMOTO@LOCAL___Custom_S_TS_SS",
    "81U1": "null@ATIVAR___DESATIVADO@ATIVADO___Custom_S_TC_SS",
    "81U2": "null@ATIVAR___DESATIVADO@ATIVADO___Custom_S_TC_SS",
    "81U3": "null@ATIVAR___DESATIVADO@ATIVADO___Custom_S_TC_SS",
    "81U4": "null@ATIVAR___DESATIVADO@ATIVADO___Custom_S_TC_SS",
    "81U5": "null@ATIVAR___DESATIVADO@ATIVADO___Custom_S_TC_SS",
}


def gerar_v8() -> None:
    wb = openpyxl.load_workbook(V7)
    ws = wb["DiscreteSignals"]
    hdr = {str(c.value).strip().upper(): c.column for c in ws[1] if c.value}
    col_sig, col_mm = hdr["SINAL"], hdr["MM"]
    alteradas = 0
    for row in range(2, ws.max_row + 1):
        sigla = str(ws.cell(row, col_sig).value or "").strip()
        if sigla in CORRECOES:
            ws.cell(row, col_mm).value = CORRECOES[sigla]
            alteradas += 1
    assert alteradas == len(CORRECOES), f"esperava {len(CORRECOES)} siglas, alterou {alteradas}"
    wb.save(V8)
    print(f"v8 salva: {V8} ({alteradas} MMs corrigidos)")


def extrair_catalogo() -> None:
    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb["MessageMappings"]
    refs = sorted({
        str(r[0]).strip() for r in ws.iter_rows(values_only=True)
        if r[0] and isinstance(r[0], str) and "@" in r[0]
    })
    wb.close()
    FIXTURE.parent.mkdir(exist_ok=True)
    FIXTURE.write_text("\n".join(refs) + "\n", encoding="utf-8")
    print(f"catálogo: {FIXTURE} ({len(refs)} refs)")


if __name__ == "__main__":
    gerar_v8()
    extrair_catalogo()
