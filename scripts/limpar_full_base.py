"""Limpa o Export Full Base (~98 MB) e extrai sinais DNP3 válidos para JSON.

Lê em streaming (openpyxl read_only=True) as sheets DNP3_DiscreteSignals,
DNP3_AnalogSignals e DNP3_DiscreteAnalog. Cada sheet tem 4 linhas de cabeçalho
(label do grupo, marcador de grupo, códigos internos, labels legíveis) — as
colunas "Signal Name", "Signal Alias" e "Measurement Type" são localizadas
pelo nome na 4a linha de cabeçalho (a última, com labels legíveis), nunca
por índice fixo.

A sigla de cada sinal é o último segmento (separado por "_") do Signal Name.

Filtros aplicados (ver spec docs/superpowers/specs/2026-06-28-sp-gt-ground-truth-automatico-design.md):
1. Mantém apenas as 3 sheets DNP3 listadas acima.
2. Remove linhas cujo Signal Name não segue o formato hierárquico (>= 2
   segmentos separados por "_").
3. Remove linhas com Signal Alias vazio, "#N/A" ou nulo.
4. Remove duplicatas de (signal_alias, sigla).
5. Remove signal_alias idênticos associados a siglas diferentes (ambiguidade).
6. Remove siglas que não existem na Lista Padrão v2 (Fonte 1).

Uso: python scripts/limpar_full_base.py (da raiz do projeto)
Saída: docs/Export_base_Full_limpo.json
"""
from __future__ import annotations

import json
import pathlib
import re
import sys

import openpyxl

_SCRIPT_DIR = pathlib.Path(__file__).resolve().parent
_ROOT = _SCRIPT_DIR.parent
sys.path.insert(0, str(_ROOT / "src"))

from tdt.dados.lista_padrao import ListaPadraoADMS  # noqa: E402

EXPORT_PATH = _ROOT / "docs" / "Export_base_Full__27_fev_2026.xlsx"
LISTA_V2_PATH = _ROOT / "docs" / "Pontos Padrao ADMS_v2.xlsx"
SAIDA_PATH = _ROOT / "docs" / "Export_base_Full_limpo.json"

SHEETS_DNP3 = ("DNP3_DiscreteSignals", "DNP3_AnalogSignals", "DNP3_DiscreteAnalog")

_ALIAS_INVALIDOS = {None, "", "#N/A", "NONE", "NULL", "N/A"}

# Signal Name hierárquico: pelo menos 2 segmentos separados por "_",
# último segmento (a sigla) não pode ser vazio.
_RE_NOME_HIERARQUICO = re.compile(r"^[^_]+(?:_[^_]+)+$")


def _coluna(header: tuple, nome: str) -> int | None:
    alvo = nome.strip().upper()
    for i, h in enumerate(header):
        if h is not None and str(h).strip().upper() == alvo:
            return i
    return None


def _sigla_de(signal_name: str) -> str | None:
    if not signal_name or not _RE_NOME_HIERARQUICO.match(signal_name):
        return None
    return signal_name.rsplit("_", 1)[-1]


def _alias_valido(alias) -> bool:
    if alias is None:
        return False
    s = str(alias).strip()
    return s.upper() not in _ALIAS_INVALIDOS and s != ""


def _ler_sheet(ws, sheet_name: str) -> list[dict]:
    linhas = ws.iter_rows(values_only=True)
    next(linhas)  # linha 1: label do grupo (ex.: "Signal Details")
    next(linhas)  # linha 2: marcador de grupo (ex.: "DSIGNAL")
    next(linhas)  # linha 3: códigos internos (IDOBJ_NAME, IDOBJ_ALIAS, ...)
    header = next(linhas)  # linha 4: labels legíveis (Signal Name, Signal Alias, ...)

    i_nome = _coluna(header, "Signal Name")
    i_alias = _coluna(header, "Signal Alias")
    i_tipo = _coluna(header, "Measurement Type")
    if i_nome is None or i_alias is None:
        raise ValueError(f"colunas Signal Name/Signal Alias ausentes em {sheet_name}")

    entradas: list[dict] = []
    for row in linhas:
        signal_name = row[i_nome] if i_nome < len(row) else None
        signal_alias = row[i_alias] if i_alias < len(row) else None
        measurement_type = row[i_tipo] if i_tipo is not None and i_tipo < len(row) else None

        signal_name = str(signal_name).strip() if signal_name is not None else None
        if not signal_name:
            continue

        sigla = _sigla_de(signal_name)
        if sigla is None:  # filtro 2: formato hierárquico
            continue

        if not _alias_valido(signal_alias):  # filtro 3
            continue
        signal_alias = str(signal_alias).strip()

        signal_type = "Analog" if sheet_name == "DNP3_AnalogSignals" else "Discrete"

        entradas.append(
            {
                "sheet": sheet_name,
                "signal_name": signal_name,
                "signal_alias": signal_alias,
                "signal_type": signal_type,
                "measurement_type": str(measurement_type).strip() if measurement_type else None,
                "_sigla": sigla,
            }
        )
    return entradas


def _siglas_validas_lista_padrao() -> set[str]:
    lp = ListaPadraoADMS.carregar(LISTA_V2_PATH)
    return {s.sigla.upper() for s in (*lp.discretos, *lp.analogicos)}


def limpar(export_path: pathlib.Path = EXPORT_PATH) -> list[dict]:
    wb = openpyxl.load_workbook(export_path, read_only=True, data_only=True)
    try:
        entradas: list[dict] = []
        for sheet_name in SHEETS_DNP3:  # filtro 1
            if sheet_name not in wb.sheetnames:
                continue
            entradas.extend(_ler_sheet(wb[sheet_name], sheet_name))
    finally:
        wb.close()

    # filtro 4: dedup de (signal_alias, sigla)
    vistos: set[tuple[str, str]] = set()
    sem_dup: list[dict] = []
    for e in entradas:
        chave = (e["signal_alias"], e["_sigla"])
        if chave in vistos:
            continue
        vistos.add(chave)
        sem_dup.append(e)

    # filtro 5: remove signal_alias idênticos associados a siglas diferentes
    siglas_por_alias: dict[str, set[str]] = {}
    for e in sem_dup:
        siglas_por_alias.setdefault(e["signal_alias"], set()).add(e["_sigla"])
    aliases_ambiguos = {alias for alias, siglas in siglas_por_alias.items() if len(siglas) > 1}
    sem_ambiguidade = [e for e in sem_dup if e["signal_alias"] not in aliases_ambiguos]

    # filtro 6: remove siglas que não existem na Lista Padrão v2
    siglas_validas = _siglas_validas_lista_padrao()
    finais = [e for e in sem_ambiguidade if e["_sigla"].upper() in siglas_validas]

    for e in finais:
        del e["_sigla"]

    return finais


def main() -> None:
    entradas = limpar()
    SAIDA_PATH.write_text(json.dumps(entradas, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Export Full Base limpo: {len(entradas)} entradas -> {SAIDA_PATH}")


if __name__ == "__main__":
    main()
