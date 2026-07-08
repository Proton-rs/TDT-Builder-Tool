"""Restaura as data validations e a sheet oculta DMSMatchingTemplateInfo no template DNP3.

O template ``docs/dnp3_template.xlsx`` perdeu as data validations de lista
(dropdowns que referenciam a DMSMatchingTemplateInfo) e tem a sheet oculta
truncada em 5 linhas — a TDT real exportada do ADMS tem 279. Sem isso o
Excel não oferece opções de input nas células e não há domínio de valores
por coluna.

Copia da TDT real exportada:
1. a DMSMatchingTemplateInfo completa (valores);
2. todas as DVs das sheets em comum, colapsadas para a linha-modelo
   (ex.: ``E5:E1645`` -> ``E5``) — o engine_tdt expande para as linhas de
   dados na geração.

Uso (na raiz do repo): python scripts/restaurar_dvs_template_dnp3.py
"""

import copy
import re

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidationList

SRC = "docs/TDT/exportTDT_UTR_GTD_1_20260626.xlsx"
DST = "docs/dnp3_template.xlsx"

SHEET_INFO = "DMSMatchingTemplateInfo"


def _colapsar_token(tok: str) -> str:
    m = re.fullmatch(r"([A-Z]+)(\d+)(?::([A-Z]+)\d+)?", tok)
    if not m:
        return tok
    c1, r1, c2 = m.group(1), m.group(2), m.group(3)
    return f"{c1}{r1}:{c2}{r1}" if c2 and c2 != c1 else f"{c1}{r1}"


def colapsar(sqref) -> str:
    """'E5:E1645 G5:G1645' -> 'E5 G5' (só a linha-modelo de cada range)."""
    return " ".join(_colapsar_token(t) for t in str(sqref).split())


def main() -> None:
    src = openpyxl.load_workbook(SRC)
    dst = openpyxl.load_workbook(DST)

    si, di = src[SHEET_INFO], dst[SHEET_INFO]
    di.delete_rows(1, di.max_row)
    for row in si.iter_rows():
        for cell in row:
            if cell.value is not None:
                di.cell(cell.row, cell.column, cell.value)
    print(f"{SHEET_INFO}: {si.max_row} linhas x {si.max_column} colunas copiadas")

    for nome in dst.sheetnames:
        if nome not in src.sheetnames:
            continue
        dvs_src = src[nome].data_validations.dataValidation
        if not dvs_src:
            continue
        dst[nome].data_validations = DataValidationList()
        for dv in dvs_src:
            novo = copy.copy(dv)
            novo.sqref = colapsar(dv.sqref)
            dst[nome].add_data_validation(novo)
        print(f"{nome}: {len(dvs_src)} DVs restauradas")

    dst.save(DST)

    # verificação: contagem de DVs por sheet tem que bater com a TDT real
    check = openpyxl.load_workbook(DST)
    assert check[SHEET_INFO].max_row == si.max_row, check[SHEET_INFO].max_row
    for nome in check.sheetnames:
        if nome in src.sheetnames:
            n_src = len(src[nome].data_validations.dataValidation)
            n_dst = len(check[nome].data_validations.dataValidation)
            assert n_dst == n_src or n_src == 0, f"{nome}: {n_dst} != {n_src}"
    print("OK: template restaurado e verificado")


if __name__ == "__main__":
    main()
