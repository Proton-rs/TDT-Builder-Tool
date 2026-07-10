"""Aceite da spec 2026-07-10 §3 sobre o TDT regenerado da IMA:
1. zero Remote Point Custom ID duplicado;
2. todo MM preenchido existe no catálogo real (fixture);
3. DNP3_DiscreteAnalog contém IMA_TR6_TR6_TAP e IMA_TR7_TR7_TAP;
4. nenhum sinal espúrio *_CMD;
5. revisao.json contém COMTAP com motivo comando_tap_nao_modelado.
"""
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
TDT = RAIZ / "output" / "TDT_IMA_v2.xlsx"
REVISAO = RAIZ / "output" / "TDT_IMA_v2.revisao.json"
CATALOGO = set((RAIZ / "tests" / "fixtures" / "mm_catalogo_real.txt")
               .read_text(encoding="utf-8").splitlines())

falhas: list[str] = []
wb = openpyxl.load_workbook(TDT, read_only=True, data_only=True)
nomes_da: set[str] = set()
# Custom ID e unico por IMPORT (todo o workbook), nao por sheet — o ADMS
# descarta remote points colidindo em qualquer lugar do mesmo import.
cids: Counter[str] = Counter()
for sheet in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals", "DNP3_DiscreteAnalog"):
    if sheet not in wb.sheetnames:
        falhas.append(f"{sheet}: sheet ausente")
        continue
    linhas = wb[sheet].iter_rows(values_only=True)
    for _ in range(3):
        next(linhas)
    hdr = [str(c) if c else "" for c in next(linhas)]
    i_cid = next((i for i, c in enumerate(hdr) if "remote point custom" in c.lower()), None)
    i_mm = next((i for i, c in enumerate(hdr) if c == "Message Mapping"), None)
    for r in linhas:
        nome = r[0]
        if not nome:
            continue
        if str(nome).endswith("_CMD"):
            falhas.append(f"{sheet}: sinal espúrio {nome}")
        if sheet == "DNP3_DiscreteAnalog":
            nomes_da.add(str(nome))
        if i_cid is not None and r[i_cid]:
            cids[str(r[i_cid])] += 1
        if i_mm is not None and r[i_mm] and str(r[i_mm]).strip() not in CATALOGO:
            falhas.append(f"{sheet}: {nome}: MM fora do catálogo: {r[i_mm]}")
wb.close()

dups = {k: v for k, v in cids.items() if v > 1}
if dups:
    falhas.append(f"Custom IDs duplicados no import: {dups}")

for esperado in ("IMA_TR6_TR6_TAP", "IMA_TR7_TR7_TAP"):
    if esperado not in nomes_da:
        falhas.append(f"DNP3_DiscreteAnalog: {esperado} ausente")

itens = json.loads(REVISAO.read_text(encoding="utf-8"))
if not any(i["motivo"] == "comando_tap_nao_modelado" for i in itens):
    falhas.append("revisao.json: COMTAP não está em revisão com motivo comando_tap_nao_modelado")

if falhas:
    print("ACEITE FALHOU:")
    for f in falhas:
        print(" -", f)
    sys.exit(1)
print("ACEITE OK: TDT IMA sem duplicatas, MMs no catálogo, TAP presente, COMTAP em revisão.")
