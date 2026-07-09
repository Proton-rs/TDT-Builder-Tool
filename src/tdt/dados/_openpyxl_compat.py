"""Shim de compatibilidade openpyxl.

A versão de openpyxl do projeto (py3.14) quebra ao ler um ``PatternFill`` que
traz o atributo ``extLst`` (extensão OOXML) — presente em alguns .xlsx reais de
concessionária (ex. ``docs/RGE GAU 2026 - Lista de Pontos v09.xlsx``). O
``__init__`` do ``PatternFill`` não aceita o kwarg e levanta ``TypeError`` antes
de qualquer leitura de dados, derrubando a UI ao abrir o arquivo.

Este shim embrulha ``PatternFill.__init__`` para descartar ``extLst`` (é só
metadado de extensão, não afeta os valores das células). Idempotente — importar
cedo (feito por ``tdt/__init__.py``).

ponytail: monkeypatch mínimo em vez de dar upgrade no openpyxl (que arrasta
mudanças de API em todo o projeto). Remover quando o openpyxl do ambiente
passar a aceitar/ignorar ``extLst`` nativamente.
"""
from __future__ import annotations

import openpyxl.styles.fills as _fills

if not getattr(_fills.PatternFill, "_extlst_shim", False):
    _orig_init = _fills.PatternFill.__init__

    def _init(self, *args, extLst=None, **kwargs):  # noqa: N802 (assinatura openpyxl)
        _orig_init(self, *args, **kwargs)

    _fills.PatternFill.__init__ = _init
    _fills.PatternFill._extlst_shim = True
