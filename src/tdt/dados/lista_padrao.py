"""Lê a lista padrão ADMS (``Pontos Padrao ADMS_v1.xlsx``).

É o gabarito de respostas: o que os scorers comparam contra. Colunas
localizadas pelo nome (row 1), nunca por índice. Linhas inválidas (sigla
vazia / ``#N/A``) são ignoradas.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path

import openpyxl

_INVALIDOS = {None, "", "#N/A", "NONE", "NULL"}


@dataclass(frozen=True)
class SinalPadrao:
    sigla: str
    descricao: str
    signal_type: str
    direction: str | None
    mm: str | None
    categoria: str  # "Discrete" | "Analog"
    estados_brutos: str | None = None
    valores_scada: tuple[int, ...] = ()
    tipo_medicao: str | None = None  # "Corrente", "Tensão", ... (lista padrão, PT)
    unidade_exibicao: str | None = None  # "A", "kV", "Grau", "-", ...
    type_severidade: str | None = None  # "PROT", "FALHAS FCOM/VCA/VCC", ... (só discretos)
    normal_value: int | None = None
    remote_point_type: str | None = None
    output_data_type: str | None = None
    device_mapping_ref: str | None = None
    aplicabilidade: str | None = None
    severidade: str | None = None  # display-only (LP), nunca em corpus/score


def _val(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return None if s.upper() in _INVALIDOS else s


def _coluna(header: tuple, nome: str) -> int | None:
    alvo = nome.strip().upper()
    for i, h in enumerate(header):
        if h is not None and str(h).strip().upper() == alvo:
            return i
    return None


def _ler_sheet(ws, categoria: str, mapa: dict[str, str]) -> list[SinalPadrao]:
    linhas = ws.iter_rows(values_only=True)
    header = next(linhas)
    idx = {
        chave: (_coluna(header, col) if col is not None else None)
        for chave, col in mapa.items()
    }
    if idx["sigla"] is None:
        raise ValueError(f"coluna SINAL ausente em {ws.title}")

    sinais: list[SinalPadrao] = []
    for row in linhas:
        sigla = _val(row[idx["sigla"]])
        if sigla is None:
            continue

        def get(chave):
            i = idx[chave]
            return _val(row[i]) if i is not None and i < len(row) else None

        valores_raw = get("valores")
        try:
            valores = tuple(int(p) for p in valores_raw.split(";")) if valores_raw else ()
        except ValueError:
            valores = ()  # "#N/A" ou formato inesperado

        sinais.append(
            SinalPadrao(
                sigla=sigla,
                descricao=get("descricao") or "",
                signal_type=get("signal_type") or "",
                direction=get("direction"),
                mm=get("mm"),
                categoria=categoria,
                estados_brutos=get("estados"),
                valores_scada=valores,
                tipo_medicao=get("tipo_medicao"),
                unidade_exibicao=get("unidade_exibicao"),
                type_severidade=get("type_severidade"),
                severidade=get("severidade"),
            )
        )
    return sinais


def _ler_sheet_discrete_analog(ws) -> list[SinalPadrao]:
    linhas = ws.iter_rows(values_only=True)
    header = next(linhas)
    idx = {n: _coluna(header, n) for n in (
        "SINAL", "DESCRIÇÃO NOVA", "SIGNAL TYPE", "FASES", "DIRECTION",
        "NORMAL VALUE", "REMOTE POINT TYPE", "OUTPUT DATA TYPE",
        "DEVICE MAPPING REF", "APLICABILIDADE")}
    out: list[SinalPadrao] = []
    for row in linhas:
        def get(nome):
            i = idx[nome]
            return _val(row[i]) if i is not None and i < len(row) else None

        sigla = get("SINAL")
        if sigla is None:
            continue
        nv = get("NORMAL VALUE")
        out.append(SinalPadrao(
            sigla=sigla, descricao=get("DESCRIÇÃO NOVA") or "",
            signal_type=get("SIGNAL TYPE") or "", direction=get("DIRECTION"),
            mm=None, categoria="DiscreteAnalog",
            normal_value=int(nv) if nv is not None else None,
            remote_point_type=get("REMOTE POINT TYPE"),
            output_data_type=get("OUTPUT DATA TYPE"),
            device_mapping_ref=get("DEVICE MAPPING REF"),
            aplicabilidade=get("APLICABILIDADE"),
        ))
    return out


@dataclass(frozen=True)
class ListaPadraoADMS:
    discretos: tuple[SinalPadrao, ...]
    analogicos: tuple[SinalPadrao, ...]
    discrete_analog: tuple[SinalPadrao, ...] = ()
    de_para: dict[str, str] = field(default_factory=dict)

    @classmethod
    def carregar(cls, path: str | Path) -> "ListaPadraoADMS":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            disc = _ler_sheet(
                wb["DiscreteSignals"],
                "Discrete",
                {
                    "sigla": "SINAL",
                    "descricao": "DESCRIÇÃO NOVA",
                    "signal_type": "SIGNAL TYPE",
                    "direction": "DIRECTION",
                    "mm": "MM",
                    "estados": "FUNÇÃO",
                    "valores": "VALOR",
                    "tipo_medicao": None,
                    "unidade_exibicao": None,
                    "type_severidade": "TYPE SEVERIDADE",
                    "severidade": "SEVERIDADE",
                },
            )
            ana = _ler_sheet(
                wb["AnalogSignals"],
                "Analog",
                {
                    "sigla": "SINAL",
                    "descricao": "DESCRIÇÃO NOVA",
                    "signal_type": "SIGNAL TYPE",
                    "direction": "DIREÇÃO DO FLUXO",
                    "mm": None,
                    "estados": None,
                    "valores": None,
                    "tipo_medicao": "TIPO DE MEDIÇÃO",
                    "unidade_exibicao": "UNIDADE DE EXIBIÇÃO",
                    "type_severidade": None,
                    "severidade": None,
                },
            )
            da: list[SinalPadrao] = []
            if "DiscreteAnalog" in wb.sheetnames:
                da = _ler_sheet_discrete_analog(wb["DiscreteAnalog"])
            de_para: dict[str, str] = {}
            if "DE->PARA" in wb.sheetnames:
                for r in wb["DE->PARA"].iter_rows(min_row=2, values_only=True):
                    de = str(r[0]).strip().upper() if r and r[0] is not None else ""
                    para = (str(r[1]).strip().upper()
                            if r and len(r) > 1 and r[1] is not None else "")
                    if de and para:
                        de_para[de] = para
        finally:
            wb.close()
        return cls(tuple(disc), tuple(ana), tuple(da), de_para)

    def _todos(self) -> tuple[SinalPadrao, ...]:
        return (*self.discretos, *self.analogicos, *self.discrete_analog)

    def por_sigla(self, sigla: str) -> SinalPadrao | None:
        alvo = sigla.strip().upper()
        for s in self._todos():
            if s.sigla.upper() == alvo:
                return s
        return None

    @property
    def siglas(self) -> frozenset[str]:
        """Todas as siglas (discretos + analógicos + discrete_analog), normalizadas
        maiúsculas — usado pela detecção de coluna de sigla em listas não-homogêneas."""
        return frozenset(s.sigla.upper() for s in self._todos())


@lru_cache(maxsize=4)
def descricoes_por_sigla(path: str) -> dict[str, str]:
    """Mapa sigla UPPER -> descrição de uma lista padrão (Signal Alias da v1).

    Arquivo ausente/ilegível -> {} : a geração de TDT nunca quebra por causa
    do alias (fallback = descrição bruta do cliente, comportamento antigo).
    """
    try:
        lp = ListaPadraoADMS.carregar(path)
    except Exception:
        return {}
    return {
        s.sigla.upper(): s.descricao
        for s in lp._todos()
        if s.descricao
    }
