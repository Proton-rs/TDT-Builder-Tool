"""Gera o TDT DNP3 preenchendo o template (preserva fórmulas, estilos e tabelas).

Carrega ``dnp3_template.xlsx`` e escreve os dados a partir da row 5, localizando
colunas pelo **display name (row 4)** — os field names da row 3 se repetem
(seção do sinal vs. seção do remote point), então não servem de chave única.

Escopo atual: sinais discretos (DNP3_DiscreteSignals) e analógicos
(DNP3_AnalogSignals, só leitura). Pareamento D+C de comando fica para o
próximo corte.
ponytail: analógico sem unidade/escala/Measurement Type ainda; entra quando o
input fornecer a grandeza física.
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import replace
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from tdt.contracts import ItemRevisao, ListaHomogenea, SignalRecord
from tdt.dados.lista_padrao import ListaPadraoADMS
from tdt.defaults import COMPLEMENTO_DM_PROT
from tdt.normalizacao.normalizador import familia_do_id

SHEET_DISCRETOS = "DNP3_DiscreteSignals"
COLUNAS_ESPERADAS = 43
SHEET_ANALOGICOS = "DNP3_AnalogSignals"
COLUNAS_ESPERADAS_ANALOG = 61
SHEET_DISCRETE_ANALOG = "DNP3_DiscreteAnalog"
COLUNAS_ESPERADAS_DISCRETE_ANALOG = 48
PRIMEIRA_LINHA_DADOS = 5

_DIRECAO = {"Input": "Read", "Output": "Write", "InputOutput": "ReadWrite"}


def _mapa_colunas(ws) -> dict[str, int]:
    """display name (row 4) -> índice de coluna (1-based)."""
    return {
        ws.cell(4, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(4, c).value
    }


_BARRA_SUFIXO = {"Principal": "P", "Auxiliar": "A"}


def nome_hierarquico(
    subestacao: str | None,
    modulo_nome: str | None,
    equipamento: str | None,
    barra: str | None,
    sigla: str,
) -> str:
    partes = []
    if subestacao:
        partes.append(subestacao)
    # "LT 1" → "LT1", "AL 11" → "AL11", "TR 1" → "TR1"
    modulo_fmt = modulo_nome.replace(" ", "") if modulo_nome else None
    if modulo_fmt:
        partes.append(modulo_fmt)
    if equipamento:
        partes.append(equipamento)
    elif modulo_fmt:
        partes.append(modulo_fmt)  # sem equipamento: repete o módulo
    sufixo_barra = _BARRA_SUFIXO.get(barra or "")
    if sufixo_barra:
        partes.append(sufixo_barra)
    partes.append(sigla)
    return "_".join(partes)


_nome_hierarquico = nome_hierarquico  # alias: bench/diag_cva.py e testes ainda importam o nome privado


def _eh_alimentador(modulo_nome: str | None) -> bool:
    if not modulo_nome:
        return False
    norm = modulo_nome.replace(" ", "").upper()
    return bool(re.match(r"^AL\d", norm))


def _aor_group(subestacao: str | None, alimentador: bool) -> str | None:
    if not subestacao:
        return None
    return f"{subestacao} {'Distr' if alimentador else 'Trans'}"


def _remote_unit(subestacao: str | None) -> str | None:
    return f"UTR_{subestacao}_1" if subestacao else None


# Sufixo de família do equipamento no Device Mapping (spec 2026-07-20 §A1;
# fullbase: DJ 16.866, TR 4.610, SEC 2.157 como último segmento). TC/TP da
# spec saem no ramo analógico (`<MOD>_TC`/`<MOD>_TP`) — familia_do_id não
# classifica TC/TP por ID; estender aqui se a whitelist ganhar esses IDs.
_SUFIXO_FAMILIA: dict[str, str] = {
    "Disjuntor": "DJ",
    "Seccionadora": "SEC",
    "Transformador": "TR",
}


def _dm_prot(sigla: str | None, sp) -> bool:
    """Flag do ramo PROT do device mapping (spec 2026-07-20 §B1): RelayTrip
    da lista padrão manda; o complemento cobre siglas não-RelayTrip que a
    fullbase mapeia consistentemente em PROT (ex. 2649). NÃO é o conceito
    ANSI de função de proteção (79 é função e mesmo assim cai no DJ)."""
    if sp is not None and sp.signal_type == "RelayTrip":
        return True
    return (sigla or "").strip().upper() in COMPLEMENTO_DM_PROT


def _device_mapping(
    nome: str,
    sigla: str,
    dm_prot: bool,
    subestacao: str | None = None,
    modulo_nome: str | None = None,
    barra: str | None = None,
    equipamento: str | None = None,
    disjuntor: str | None = None,
) -> str:
    """Padrão RGE (spec 2026-07-20): proteção cai no módulo duplicado;
    não-proteção cai no equipamento com sufixo de família (_DJ/_SEC/_TR).
    Sem equipamento, o fallback módulo-duplicado emerge sozinho (sem sufixo).
    Base terminando em sufixo de barra fica sem sufixo de família
    (conservador — fullbase não tem exemplo com barra + sufixo).
    Exceção §A2: proteção de alimentador com disjuntor único conhecido usa o
    disjuntor (não o módulo repetido) como 2º nível — supersede a correção
    16/07 só para esse caso; demais módulos mantêm o módulo duplicado."""
    if dm_prot:
        if _eh_alimentador(modulo_nome) and disjuntor:
            # decisão 20/07 (supersede 16/07): alimentador usa o DISJUNTOR do
            # módulo (não o equipamento da linha) como 2º nível do PROT
            return nome_hierarquico(subestacao, modulo_nome, disjuntor, barra, f"PROT_{sigla}")
        return nome_hierarquico(subestacao, modulo_nome, None, barra, f"PROT_{sigla}")
    base = nome[: len(nome) - len(sigla) - 1] if nome.endswith(f"_{sigla}") else nome
    suf = _SUFIXO_FAMILIA.get(familia_do_id(equipamento) or "")
    if suf and equipamento and base.endswith(equipamento):
        return f"{base}_{suf}"
    return base


def _normal_value(sp: "SinalPadrao | None") -> int | None:
    if sp is None or not sp.estados_brutos or not sp.valores_scada:
        return None
    estados = sp.estados_brutos.split(";")
    try:
        i = estados.index("NORMAL")
    except ValueError:
        return None
    return sp.valores_scada[i] if i < len(sp.valores_scada) else None


def _alias_hoje() -> str:
    return date.today().strftime("%Y%m%d")


def _signal_alias(rec: SignalRecord, alias_v1: "dict[str, str] | None") -> str:
    """Descrição da lista padrão v1 quando a sigla está no mapa; senão a
    descrição bruta do cliente (Custom/sem sigla/mapa ausente)."""
    if alias_v1 and rec.sigla_sinal:
        desc = alias_v1.get(rec.sigla_sinal.upper())
        if desc:
            return desc
    return rec.descricoes.bruta


def _coords_comando(indices: tuple[int, ...], duplo: bool = True) -> str:
    if len(indices) == 1:
        return f"{indices[0]};{indices[0]}" if duplo else str(indices[0])
    return ";".join(str(i) for i in indices)


def _output_data_type(coords_saida: str | None) -> str | None:
    """Domínio DNP3OutputType (DMSMatchingTemplateInfo): ``SingleCoord`` quando
    o comando escreve numa coordenada só (``i;i`` ou índice único),
    ``MultiCoord`` quando escreve em duas distintas (trip;close)."""
    if not coords_saida:
        return None
    partes = coords_saida.split(";")
    return "SingleCoord" if len(set(partes)) == 1 else "MultiCoord"


_FASE_PHASECODE: dict[str, str] = {"CA": "AC"}  # interno -> PhaseCode ADMS
_PHASECODE = frozenset({"ABC", "AB", "BC", "AC", "A", "B", "C", "N"})  # DMSMatchingTemplateInfo


def _fase_saida(fase: str | None) -> str:
    """Fase para a coluna TDT ``Phases``: default ``ABC`` quando vazia, e
    fallback ``ABC`` para qualquer valor fora do domínio ``PhaseCode`` (guard
    de domínio — o ADMS rejeita fase inválida). Traduz a convenção interna de
    campo (``CA``) para o par alfabético do domínio ADMS (``AC``)."""
    fase = _FASE_PHASECODE.get(fase, fase)
    return fase if fase in _PHASECODE else "ABC"


def dm_registro(rec, subestacao, sp, disjuntor: str | None = None) -> tuple[str, str]:
    """(Signal Name, Device Mapping) do registro — derivação ÚNICA, usada por
    _valores, particionar_tipo_duplicado, particionar_custom_id_duplicado
    (quando lista_padrao é informado) e pelas colunas derivadas da UI.
    Signal Name (decisão do usuário 21/07 — reverte a exceção da Task 11):
    sempre módulo-duplicado quando não há equipamento explícito, mesmo no
    PROT de alimentador com disjuntor conhecido; só o Device Mapping usa o
    disjuntor nesse caso (mantido, Task 2)."""
    sigla = rec.sigla_sinal or "?"
    dm_prot = _dm_prot(rec.sigla_sinal, sp)
    equipamento = rec.eletrico.nome_equipamento
    nome = nome_hierarquico(
        subestacao, rec.modulo.nome, equipamento,
        rec.eletrico.barra, sigla,
    )
    dm = _device_mapping(
        nome, sigla, dm_prot, subestacao,
        rec.modulo.nome, rec.eletrico.barra,
        equipamento=rec.eletrico.nome_equipamento, disjuntor=disjuntor,
    )
    return nome, dm


def _valores(rec: SignalRecord, subestacao: str | None, padrao: ListaPadraoADMS,
             alias_v1: "dict[str, str] | None" = None,
             disjuntor: "str | None" = None) -> dict:
    sp = padrao.por_sigla(rec.sigla_sinal) if rec.sigla_sinal else None
    nome, dm = dm_registro(rec, subestacao, sp, disjuntor)
    alimentador = _eh_alimentador(rec.modulo.nome)
    remote_unit = _remote_unit(subestacao)
    rp_custom = f"{nome}_{remote_unit}" if remote_unit else None
    indices = rec.enderecamento.indices
    coords = indices[0] if len(indices) == 1 else ";".join(str(i) for i in indices)
    direcao = rec.tipo_sinal.direcao
    tem_comando = direcao in ("Output", "InputOutput")
    if direcao == "Output":
        # Comando órfão (sem par de status): o próprio endereço é o de escrita,
        # não há leitura — não preencher Input Coordinates.
        coords_entrada = None
        coords_saida = _coords_comando(rec.enderecamento.indices, rec.tipo_sinal.comando_duplo)
    else:
        coords_entrada = coords
        coords_saida = (
            _coords_comando(rec.enderecamento.indices_saida, rec.tipo_sinal.comando_duplo)
            if rec.enderecamento.indices_saida else ""
        )
    return {
        "Signal Name": nome,
        "Signal Alias": _signal_alias(rec, alias_v1),
        "Measurement Type": "Status",  # ponytail: default; refinar por signal_type
        "Signal Type": sp.signal_type if sp else "Custom",
        "Side": "None",
        "Output Register": False,
        "Remote Point Type": "Status",
        "Remote Point Name": nome,
        "Phases": _fase_saida(rec.eletrico.fase),
        "Signal AOR Group": _aor_group(subestacao, alimentador),
        "Device Mapping": dm,
        "Direction": _DIRECAO.get(direcao, "Read"),
        "Message Mapping": sp.mm if sp else None,
        "Input Data Type": rec.tipo_sinal.datatype,
        "Input Coordinates": coords_entrada,
        "Output Data Type": _output_data_type(coords_saida) if tem_comando else None,
        "Output Coordinates": coords_saida if tem_comando and coords_saida else None,
        "Remote Unit": remote_unit,
        "Remote Point Custom ID": rp_custom,
        "Remote Point Alias": _alias_hoje(),
        "Normal Value": _normal_value(sp),
    }


_MEASUREMENT_TYPE_PT_EN: dict[str, str] = {
    "CORRENTE": "Current",
    "TENSÃO": "Voltage",
    "POTÊNCIA ATIVA": "ActivePower",
    "POTÊNCIA REATIVA": "ReactivePower",
    "TEMPERATURA": "Temperature",
    # auditoria 09jul (lista padrao v6) — todos confirmados no dominio
    # MeasurementType do DMSMatchingTemplateInfo:
    "COMPRIMENTO": "Unitless",      # KMDF: distancia de defeito e unitless no ADMS (decisao do usuario; o dominio tem "Length" mas usamos Unitless)
    "FREQUÊNCIA": "Frequency",
    "FATOR DE POTÊNCIA": "CosPhi",  # fator de potencia = cos(phi); o dominio usa CosPhi, NAO PowerFactor
    "POTÊNCIA APARENTE": "ApparentPower",
    "ÂNGULO DE TENSÃO": "VoltageAngle",
    "UMIDADE": "Humidity",
    "DISCRETO": "Discrete",
}


def _measurement_type(sp) -> str | None:
    if sp is None or not sp.tipo_medicao:
        return None
    return _MEASUREMENT_TYPE_PT_EN.get(sp.tipo_medicao.strip().upper())
# ponytail: tabela cobre os 12 tipos da lista padrao v6; expandir se novos tipos forem adicionados ao dominio MeasurementType do DMSMatchingTemplateInfo.


_SIGNAL_TYPE_ANALOG_PT_EN: dict[str, str] = {
    "VALOR MEDIDO": "MeasuredValue",
    "GRAVADOR DE FALHA": "FaultRecorder",
    "CONTAGEM DE OPERAÇÃO": "Custom",  # AnalogSignalType não tem OperationCount
}


def _signal_type_analog(sp) -> str:
    """Domínio AnalogSignalType (DMSMatchingTemplateInfo) a partir do tipo PT
    da lista padrão — o ADMS rejeita valor fora do domínio no import."""
    if sp is None or not sp.signal_type:
        return "Custom"
    st = sp.signal_type.strip()
    return _SIGNAL_TYPE_ANALOG_PT_EN.get(st.upper(), st)
# ponytail: cobre os 4 valores da lista padrão v2; valor novo passa reto (teste de domínio pega).


# Grandeza (Measurement Type PT da lista padrão) -> entidade do device
# mapping analógico, padrão RGE (spec 2026-07-15).
_MEDIDAS_TC = frozenset({"CORRENTE", "POTÊNCIA ATIVA", "POTÊNCIA REATIVA", "POTÊNCIA APARENTE"})
_MEDIDAS_TP = frozenset({"TENSÃO"})


def disjuntor_por_modulo(registros) -> "dict[str | None, str | None]":
    """Disjuntor único de cada módulo, usado pelo DM analógico e pelo ramo
    PROT de alimentador (spec 2026-07-20 §A2/§A3). 0 ou 2+ disjuntores
    -> None (fallback módulo-duplicado; o aviso de ambiguidade já foi emitido
    por atribuir_id_por_registro no pipeline)."""
    por_mod: dict[str | None, set[str]] = defaultdict(set)
    for rec in registros:
        ne = rec.eletrico.nome_equipamento
        if ne and familia_do_id(ne) == "Disjuntor":
            por_mod[rec.modulo.nome].add(ne)
    return {m: next(iter(ids)) if len(ids) == 1 else None for m, ids in por_mod.items()}


def _medida_usa_disjuntor(tipo_medicao_pt: str | None) -> bool:
    """True quando a grandeza analógica NÃO é corrente/potência (_MEDIDAS_TC)
    nem tensão (_MEDIDAS_TP) — ramo que cai no disjuntor do módulo, tanto em
    `_device_mapping_analog` quanto em `_valores_analog` (spec 2026-07-20
    §A3/§C12; single source of truth desde Task 12, evita os dois lugares
    divergirem)."""
    t = (tipo_medicao_pt or "").strip().upper()
    return t not in _MEDIDAS_TC and t not in _MEDIDAS_TP


def _device_mapping_analog(
    subestacao: str | None,
    modulo_nome: str | None,
    tipo_medicao_pt: str | None,
    disjuntor: str | None,
) -> str:
    """Padrão RGE: corrente/potências -> <MOD>_TC; tensão -> <MOD>_TP;
    demais grandezas (KMDF, frequência, FP, temperatura...) -> disjuntor do
    módulo; sem disjuntor único -> módulo duplicado (<SUB>_<MOD>_<MOD>)."""
    modulo_fmt = modulo_nome.replace(" ", "") if modulo_nome else None
    partes = [p for p in (subestacao, modulo_fmt) if p]
    t = (tipo_medicao_pt or "").strip().upper()
    if _medida_usa_disjuntor(t):
        alvo = f"{disjuntor}_DJ" if disjuntor else modulo_fmt
    elif t in _MEDIDAS_TC:
        alvo = f"{modulo_fmt}_TC" if modulo_fmt else "TC"
    else:
        alvo = f"{modulo_fmt}_TP" if modulo_fmt else "TP"
    if alvo:
        partes.append(alvo)
    return "_".join(partes)


def _valores_analog(rec: SignalRecord, subestacao: str | None, padrao: ListaPadraoADMS,
                     alias_v1: "dict[str, str] | None" = None,
                     disjuntor: "str | None" = None) -> dict:
    sp = padrao.por_sigla(rec.sigla_sinal) if rec.sigla_sinal else None
    equipamento = rec.eletrico.nome_equipamento
    nome = nome_hierarquico(
        subestacao, rec.modulo.nome, equipamento,
        rec.eletrico.barra, rec.sigla_sinal or "?",
    )
    indices = rec.enderecamento.indices
    coords = indices[0] if len(indices) == 1 else ";".join(str(i) for i in indices)
    alimentador = _eh_alimentador(rec.modulo.nome)
    remote_unit = _remote_unit(subestacao)
    rp_custom = f"{nome}_{remote_unit}" if remote_unit else None
    return {
        "Signal Name": nome,
        "Signal Alias": _signal_alias(rec, alias_v1),
        "Signal Type": _signal_type_analog(sp),
        "Phases": _fase_saida(rec.eletrico.fase),
        "Direction": "Read",
        "Input Coordinates": coords,
        "Side": "None",
        "Output Register": False,
        "Remote Point Type": "Analog",
        "Remote Point Name": nome,
        "Signal AOR Group": _aor_group(subestacao, alimentador),
        "Device Mapping": _device_mapping_analog(
            subestacao, rec.modulo.nome, sp.tipo_medicao if sp else None, disjuntor),
        "Remote Unit": remote_unit,
        "Remote Point Custom ID": rp_custom,
        "Remote Point Alias": _alias_hoje(),
        "Measurement Type": _measurement_type(sp),
        "Display Unit": sp.unidade_exibicao if sp and sp.unidade_exibicao not in (None, "-") else None,
    }


def _eh_discrete_analog(rec: SignalRecord, padrao: ListaPadraoADMS) -> bool:
    sp = padrao.por_sigla(rec.sigla_sinal) if rec.sigla_sinal else None
    return bool(sp and sp.categoria == "DiscreteAnalog")


def _valores_discrete_analog(rec: SignalRecord, subestacao: str | None,
                              padrao: ListaPadraoADMS,
                              alias_v1: "dict[str, str] | None" = None) -> dict:
    """TAP real (GTD DNP3_DiscreteAnalog): Measurement Type=Discrete,
    Signal Type=TapPosition, Remote Point Type=Analog, Normal Value=9,
    Device Mapping -> comando COMTAP no mesmo módulo."""
    sp = padrao.por_sigla(rec.sigla_sinal) if rec.sigla_sinal else None
    nome = nome_hierarquico(
        subestacao, rec.modulo.nome, rec.eletrico.nome_equipamento,
        rec.eletrico.barra, rec.sigla_sinal or "?",
    )
    remote_unit = _remote_unit(subestacao)
    indices = rec.enderecamento.indices
    coords = indices[0] if len(indices) == 1 else ";".join(str(i) for i in indices)
    device = nome
    ref = sp.device_mapping_ref if sp else None
    if ref and rec.sigla_sinal and nome.endswith(rec.sigla_sinal):
        device = nome[: len(nome) - len(rec.sigla_sinal)] + ref
    return {
        "Signal Name": nome,
        "Signal Alias": _signal_alias(rec, alias_v1),
        "Measurement Type": "Discrete",
        "Signal Type": (sp.signal_type if sp and sp.signal_type else "Custom"),
        "Phases": _fase_saida(rec.eletrico.fase),
        "Side": "None",
        "Direction": "Read",
        "Normal Value": sp.normal_value if sp else None,
        "Device Mapping": device,
        "Signal AOR Group": _aor_group(subestacao, _eh_alimentador(rec.modulo.nome)),
        "Input Coordinates": coords,
        "Remote Point Type": (sp.remote_point_type if sp and sp.remote_point_type else "Analog"),
        "Remote Point Name": nome,
        "Remote Unit": remote_unit,
        "Remote Point Custom ID": f"{nome}_{remote_unit}" if remote_unit else None,
        "Remote Point Alias": _alias_hoje(),
    }


def particionar_custom_id_duplicado(
    lista: ListaHomogenea,
    lista_padrao=None,
) -> tuple[ListaHomogenea, tuple[ItemRevisao, ...]]:
    """Gate de unicidade (spec 2026-07-10): o ADMS descarta remote points com
    Custom ID repetido no mesmo import. Grupos que colidem saem TODOS do TDT
    e vão para revisão — nunca saem calados no xlsx.
    `lista_padrao` (spec 2026-07-20, fix pós-Task 2/11): quando informado, usa
    a MESMA derivação de Signal Name de `dm_registro` — evita o gate validar
    unicidade contra um nome que não é o que realmente sai no xlsx. Quando
    None (compat retroativa dos testes existentes), mantém o comportamento
    antigo (nome_hierarquico puro)."""
    remote_unit = _remote_unit(lista.subestacao)
    disj = disjuntor_por_modulo(lista.registros) if lista_padrao is not None else {}
    por_cid: dict[str, list[SignalRecord]] = defaultdict(list)
    for rec in lista.registros:
        if lista_padrao is not None:
            sp = lista_padrao.por_sigla(rec.sigla_sinal) if rec.sigla_sinal else None
            nome, _ = dm_registro(rec, lista.subestacao, sp, disj.get(rec.modulo.nome))
        else:
            nome = nome_hierarquico(
                lista.subestacao, rec.modulo.nome, rec.eletrico.nome_equipamento,
                rec.eletrico.barra, rec.sigla_sinal or "?",
            )
        cid = f"{nome}_{remote_unit}" if remote_unit else nome
        por_cid[cid].append(rec)
    duplicados = {id(r) for grupo in por_cid.values() if len(grupo) > 1 for r in grupo}
    if not duplicados:
        return lista, ()
    motivo_por_id: dict[int, str] = {}
    for grupo in por_cid.values():
        if len(grupo) <= 1:
            continue
        # mesma derivação de modelo_tabela.sheet_origem
        sheets = {r.id.rsplit(":", 1)[0] if ":" in r.id else "" for r in grupo}
        motivo = "modulo_duplicado_entre_sheets" if len(sheets) > 1 else "custom_id_duplicado"
        for r in grupo:
            motivo_por_id[id(r)] = motivo
    revisao = tuple(
        ItemRevisao(replace(r, status="revisao"), motivo=motivo_por_id[id(r)])
        for r in lista.registros if id(r) in duplicados
    )
    restantes = tuple(r for r in lista.registros if id(r) not in duplicados)
    return replace(lista, registros=restantes), revisao


def particionar_endereco_duplicado(
    lista: ListaHomogenea,
) -> tuple[ListaHomogenea, tuple[ItemRevisao, ...]]:
    """Gate por módulo (SP-CVA2 E6.2): dois pontos do MESMO módulo, no MESMO
    espaço de endereçamento (categoria, in/out), com o mesmo índice, saem
    TODOS da lista e vão pra revisão. Sintoma típico: direção errada na
    origem (comando lido como Input colide com o Input real de mesmo índice
    — hipótese CVA11). Módulo entra na chave: índice local reusado entre
    módulos DISTINTOS (IEDs/linhas diferentes) é endereçamento normal, não
    colisão (achado do decision gate original, ver nota no topo da task) —
    mesmo racional de `_chave` em dc_pairer/normalizador_estrutural e do
    Custom ID em `particionar_custom_id_duplicado` (`_remote_unit` é
    constante por SUBESTAÇÃO, não desambigua módulo). MultiCoord/DoubleBit
    contribuem cada índice individualmente."""
    grupos: dict[tuple, dict[str, SignalRecord]] = defaultdict(dict)
    for rec in lista.registros:
        mod = rec.modulo.nome
        cat = rec.tipo_sinal.categoria
        espaco_in = "out" if rec.tipo_sinal.direcao == "Output" else "in"
        for idx in rec.enderecamento.indices:
            grupos[(mod, cat, espaco_in, idx)][rec.id] = rec
        for idx in rec.enderecamento.indices_saida:
            grupos[(mod, cat, "out", idx)][rec.id] = rec

    colididos: dict[str, SignalRecord] = {}
    for regs in grupos.values():
        if len(regs) > 1:
            colididos.update(regs)
    if not colididos:
        return lista, ()
    restantes = tuple(r for r in lista.registros if r.id not in colididos)
    revisao = tuple(
        ItemRevisao(r, motivo="endereco_duplicado") for r in colididos.values()
    )
    return replace(lista, registros=restantes), revisao


#  Sinais cuja classificação real (fullbase 2026-07-20: 424/424 dispositivos
#  com 43LR+43TC no mesmo device tem 43LR=Custom) vale mesmo que o catálogo em
#  uso esteja desatualizado (ex. Pontos Padrao ADMS_v2.xlsx classifica
#  43LR=Local, incorreto vs. v8/produção real). Evita falso-positivo do gate
#  B3 quando o catálogo carregado é stale.
_FORCA_CUSTOM_TIPO_DUPLICADO = frozenset({"43LR"})


def particionar_tipo_duplicado(
    lista: ListaHomogenea,
    lista_padrao,
) -> tuple[ListaHomogenea, tuple[ItemRevisao, ...]]:
    """Gate (spec 2026-07-20 §B3): dois sinais não-Custom com o MESMO Signal
    Type caindo DIRETO no mesmo dispositivo (DM final) conflitam no ADMS —
    o grupo inteiro sai do TDT e vai pra revisão (padrão custom_id: nunca
    sai calado no xlsx). Sinais do ramo PROT ficam de fora — dentro de
    proteção o repetido é válido (decisão do usuário 20/07).

    Restrito a sinais Discretos (categoria == "Discrete"; exclui também
    "DiscreteAnalog"): no catálogo real (`Pontos Padrao ADMS_v8.xlsx`, sheet
    AnalogSignals), 55 das 62 siglas Analógicas compartilham
    `SIGNAL TYPE == "Valor Medido"` — genérico demais pra servir de chave de
    colisão; qualquer dispositivo com 2+ medições analógicas (ex. corrente +
    potência) colidiria por acaso (achado da revisão 2026-07-20 contra dado
    real SAN2: 100% falso-positivo). A categoria vem preferencialmente do
    CATÁLOGO (`sp.categoria`), não de `rec.tipo_sinal.categoria`: a
    classificação estrutural por linha (`estruturador.estruturar`) tem
    default "Discrete" e só reclassifica pra "Analog" com marcador de seção
    OU grandeza contínua na descrição — em planilhas não-homogêneas com
    coluna SIGLA dedicada (ex. SAN2) o sheet "Analogicos" nunca aciona
    nenhum dos dois gatilhos, então TODO sinal analógico chega aqui com
    `rec.tipo_sinal.categoria == "Discrete"` (confirmado instrumentando o
    pipeline com dado real). `sp.categoria` é atribuído pela PRÓPRIA aba do
    catálogo (AnalogSignals -> "Analog", DiscreteSignals -> "Discrete") e
    não depende dessa heurística — só cai no fallback de `rec.tipo_sinal`
    quando a sigla está fora do catálogo (`sp is None`, já isento via
    "Custom" antes de chegar no agrupamento na prática, mas mantido por
    completude). `43LR` é forçado para "Custom" via
    `_FORCA_CUSTOM_TIPO_DUPLICADO` independente do que o catálogo carregado
    diga (fullbase real: 43LR=Custom em 424/424 dispositivos; a lista padrão
    v2, usada por alguns testes/fixtures antigos, classifica errado como
    "Local", o que colidiria com 43TC)."""
    disj = disjuntor_por_modulo(lista.registros)
    grupos: dict[tuple[str, str], list[SignalRecord]] = defaultdict(list)
    for rec in lista.registros:
        sp = lista_padrao.por_sigla(rec.sigla_sinal) if rec.sigla_sinal else None
        categoria = sp.categoria if sp is not None else rec.tipo_sinal.categoria
        if categoria != "Discrete":
            continue
        sigla_norm = (rec.sigla_sinal or "").strip().upper()
        if sigla_norm in _FORCA_CUSTOM_TIPO_DUPLICADO:
            st = "Custom"
        else:
            st = (sp.signal_type if sp else "Custom") or "Custom"
        if st == "Custom" or _dm_prot(rec.sigla_sinal, sp):
            continue
        _, dm = dm_registro(rec, lista.subestacao, sp, disj.get(rec.modulo.nome))
        grupos[(dm, st)].append(rec)
    colididos = {r.id for regs in grupos.values() if len(regs) > 1 for r in regs}
    if not colididos:
        return lista, ()
    revisao = tuple(
        ItemRevisao(replace(r, status="revisao"), motivo="tipo_duplicado_dispositivo")
        for r in lista.registros if r.id in colididos
    )
    restantes = tuple(r for r in lista.registros if r.id not in colididos)
    return replace(lista, registros=restantes), revisao


def dispositivos_43lr_sem_43tc(registros) -> tuple[str, ...]:
    """Dispositivos (módulo, equipamento) com 43LR e sem 43TC — precisa haver
    um sinal Local (43TC) no dispositivo (spec 2026-07-20 §B2; catálogo v8:
    43LR=Custom, 43TC=Local)."""
    lr: set[tuple[str, str]] = set()
    tc: set[tuple[str, str]] = set()
    for rec in registros:
        sigla = (rec.sigla_sinal or "").strip().upper()
        if sigla not in ("43LR", "43TC"):
            continue
        chave = (rec.modulo.nome or "?", rec.eletrico.nome_equipamento or "?")
        (lr if sigla == "43LR" else tc).add(chave)
    return tuple(f"{m}/{e}" for m, e in sorted(lr - tc))


def gerar(
    lista: ListaHomogenea,
    template_path: str | Path,
    lista_padrao: ListaPadraoADMS,
    alias_v1: "dict[str, str] | None" = None,
) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(template_path)  # mantém fórmulas/estilos
    regs_da = [r for r in lista.registros if _eh_discrete_analog(r, lista_padrao)]
    ids_da = {id(r) for r in regs_da}
    regs_disc = [r for r in lista.registros
                 if r.tipo_sinal.categoria == "Discrete" and id(r) not in ids_da]
    regs_ana = [r for r in lista.registros
                if r.tipo_sinal.categoria == "Analog" and id(r) not in ids_da]
    disj = disjuntor_por_modulo(lista.registros)
    _escrever_sheet(
        wb[SHEET_DISCRETOS], SHEET_DISCRETOS, COLUNAS_ESPERADAS,
        regs_disc,
        lambda rec, sub, padrao: _valores(
            rec, sub, padrao, alias_v1, disj.get(rec.modulo.nome)),
        lista.subestacao, lista_padrao,
    )
    _escrever_sheet(
        wb[SHEET_ANALOGICOS], SHEET_ANALOGICOS, COLUNAS_ESPERADAS_ANALOG,
        regs_ana,
        lambda rec, sub, padrao: _valores_analog(
            rec, sub, padrao, alias_v1, disj.get(rec.modulo.nome)),
        lista.subestacao, lista_padrao,
    )
    if regs_da:
        _escrever_sheet(
            wb[SHEET_DISCRETE_ANALOG], SHEET_DISCRETE_ANALOG,
            COLUNAS_ESPERADAS_DISCRETE_ANALOG, regs_da,
            lambda rec, sub, padrao: _valores_discrete_analog(rec, sub, padrao, alias_v1),
            lista.subestacao, lista_padrao,
        )
    return wb


def _escrever_sheet(ws, sheet_nome, colunas_esperadas, registros, valores_fn, subestacao, padrao):
    if ws.max_column != colunas_esperadas:
        raise ValueError(
            f"{sheet_nome} tem {ws.max_column} colunas, esperado "
            f"{colunas_esperadas} — template desatualizado?"
        )
    colunas = _mapa_colunas(ws)
    linha = PRIMEIRA_LINHA_DADOS
    for rec in registros:
        for display, valor in valores_fn(rec, subestacao, padrao).items():
            col = colunas.get(display)
            if col and valor is not None:
                ws.cell(linha, col, valor)
        linha += 1
    ultima = linha - 1
    _expandir_tabela(ws, sheet_nome, ultima)
    if ultima >= PRIMEIRA_LINHA_DADOS:
        _expandir_cf(ws, ultima_linha=ultima)
        _expandir_dv(ws, ultima_linha=ultima)


def _expandir_tabela(ws, sheet_nome: str, ultima_linha: int) -> None:
    """Ajusta o ref do ListObject para cobrir as linhas de dados escritas."""
    if sheet_nome not in ws.tables:
        return
    ultima_col = get_column_letter(ws.max_column)
    fim = max(ultima_linha, PRIMEIRA_LINHA_DADOS)
    ws.tables[sheet_nome].ref = f"A4:{ultima_col}{fim}"


def _expandir_range_row5(sqref: str, até_linha: int) -> str:
    """Expande range(s) que começam na row 5 para até a linha dada.

    Ex: 'A5' -> 'A5:A{n}', 'AP5:AQ5' -> 'AP5:AQ{n}',
    'B5 Y5' (multi-range da TDT real) -> 'B5:B{n} Y5:Y{n}'.
    """
    return " ".join(_expandir_token_row5(t, até_linha) for t in sqref.split())


def _expandir_token_row5(token: str, até_linha: int) -> str:
    m = re.match(r'^([A-Z]+)5(?::([A-Z]+)5)?$', token)
    if not m:
        return token
    col1 = m.group(1)
    col2 = m.group(2)
    if col2:
        return f"{col1}5:{col2}{até_linha}"
    return f"{col1}5:{col1}{até_linha}"


def _expandir_cf(ws, ultima_linha: int) -> None:
    """Expande conditional formatting rules da row 5 para todo o range de dados."""
    cfl = ws.conditional_formatting
    items = []
    for cf in cfl:
        sqref = _expandir_range_row5(str(cf.sqref), ultima_linha)
        for rule in cf.rules:
            items.append((sqref, rule))
    if not items:
        return
    cfl._cf_rules.clear()
    for sqref, rule in items:
        cfl.add(sqref, rule)


def _expandir_dv(ws, ultima_linha: int) -> None:
    """Expande data validations da row 5 para todo o range de dados."""
    for dv in ws.data_validations.dataValidation:
        novo = _expandir_range_row5(str(dv.sqref), ultima_linha)
        if novo != str(dv.sqref):
            dv.sqref = novo


def salvar(wb: openpyxl.Workbook, destino: str | Path) -> None:
    wb.save(destino)
