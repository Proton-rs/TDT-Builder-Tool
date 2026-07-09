"""Gera Auditoria_<SUB>_<YYYYMMDD>.xlsx (nome via ``nomes_saida.nome_saida``,
sufixo ``_vN`` em colisão): uma linha por sinal com status, sigla
decidida, scores por método e os candidatos descartados — para cruzar com
a TDT na auditoria pós-classificação."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tdt.contracts import ItemRevisao, SignalRecord
from tdt.nomes_saida import nome_saida
from tdt.semantica_estados import detectar_estado

_CODIGO_RE = re.compile(r"^[A-Z0-9_/-]+$")


def descricao_para_exibicao(desc_bruta: str, sigla: str | None) -> str:
    """Sigla como descrição exibida quando ``desc_bruta`` é um código interno
    (ex: "SND_LT67SAN_LT67SAN_79", sem espaços) em vez de descrição real --
    listas padronizadas com NOME no lugar da descrição (sigla não-homogênea)."""
    if sigla and _CODIGO_RE.match(desc_bruta.strip().upper()):
        return sigla
    return desc_bruta

_LARGURA_MAX_COLUNA = 40

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_HEADER_ALIGN = Alignment(horizontal="center")

CABECALHO = [
    "ID Sinal", "Descrição Bruta", "Tipo", "Endereço", "Status",
    "Sigla Decidida", "Score Final", "Motivo Revisão",
    "Candidato 1", "Score tfidf 1", "Score vetorial 1", "Score fuzzy 1",
    "Candidato 2", "Score tfidf 2", "Score vetorial 2", "Score fuzzy 2",
    "Candidato 3", "Score tfidf 3", "Score vetorial 3", "Score fuzzy 3",
]

_COLUNAS_ESTENDIDAS = [
    "Sheet Origem", "Desc Normalizada", "Desc Canônica", "Equip Alvo (N0)",
    "Nome Equip", "Barra", "Fase", "Estado Semântico", "Regras Aplicadas",
    "Gap", "Gap Exigido", "Etapa Decisora", "Endereço Bruto",
]
_N_COLUNAS_BASE = len(CABECALHO)  # candidatos/scores -- coluna fixa antes das colunas estendidas
CABECALHO = CABECALHO + _COLUNAS_ESTENDIDAS


def _scores_metodo(rec: SignalRecord, sigla: str | None) -> tuple[str, str, str]:
    if rec.diagnostico is None or sigla is None:
        return ("", "", "")
    por = rec.diagnostico.scores_por_metodo.get(sigla, {})
    return tuple(
        f"{por[m]:.3f}" if m in por else "" for m in ("tfidf", "vetorial", "fuzzy")
    )


def _sheet_origem(rec: SignalRecord) -> str:
    """Sheet de origem, extraída do ``id`` estável (``f"{sheet}:{linha}"``)."""
    return rec.id.rsplit(":", 1)[0] if ":" in rec.id else ""


def _estado_semantico(rec: SignalRecord) -> str:
    est = detectar_estado(rec.descricoes.normalizada)
    if est is None:
        return ""
    return f"{est.classe}/{est.polaridade}" if est.polaridade else est.classe


def _colunas_estendidas(rec: SignalRecord, diag: dict | None) -> list:
    diag = diag or {}
    e = rec.eletrico
    gap = diag.get("gap")
    gap_exigido = diag.get("gap_exigido")
    return [
        _sheet_origem(rec),
        diag.get("desc_normalizada", ""),
        rec.descricoes.normalizada,  # "Desc Canônica": forma pós canonizar() (N1..N6)
        e.equipamento_alvo or "",
        e.nome_equipamento or "",
        e.barra or "",
        e.fase or "",
        _estado_semantico(rec),
        diag.get("regras_aplicadas", ""),
        f"{gap:.3f}" if gap is not None else "",
        f"{gap_exigido:.3f}" if gap_exigido is not None else "",
        diag.get("etapa_decisora", ""),
        diag.get("endereco_bruto", ""),
    ]


def gerar_relatorio_revisao(
    registros: list[SignalRecord],
    revisao: tuple[ItemRevisao, ...],
    destino: str | Path,
    diagnostico: "dict[str, dict] | None" = None,
    subestacao: str | None = None,
) -> Path:
    """``diagnostico``: dict opcional ``id -> {chaves}`` com contexto de decisão
    que não mora no ``SignalRecord`` (regras aplicadas, gap/gap exigido, etapa
    decisora, endereço bruto pré-pareamento, descrição pós-N1). Montado pelo
    pipeline (ver ``tdt.pipeline.executar``); ausente ou sem entrada para um
    id -> colunas correspondentes ficam vazias (nunca fabricadas)."""
    motivo_por_id = {it.registro.id: it.motivo for it in revisao}
    diagnostico = diagnostico or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"
    ws.append(CABECALHO)
    for rec in registros:
        linha = [
            rec.id,
            descricao_para_exibicao(rec.descricoes.bruta, rec.sigla_sinal),
            f"{rec.tipo_sinal.categoria}/{rec.tipo_sinal.direcao}",
            ";".join(str(i) for i in rec.enderecamento.indices),
            rec.status,
            rec.sigla_sinal or "",
            f"{rec.candidatos[0].score:.3f}" if rec.candidatos else "",
            motivo_por_id.get(rec.id, ""),
        ]
        for c in rec.candidatos[:3]:
            linha.append(c.sigla)
            linha.extend(_scores_metodo(rec, c.sigla))
        linha += [""] * (_N_COLUNAS_BASE - len(linha))
        linha.extend(_colunas_estendidas(rec, diagnostico.get(rec.id)))
        ws.append(linha)
    _formatar_cabecalho(ws)
    _ajustar_largura_colunas(ws)
    saida = nome_saida("Auditoria", subestacao, destino)
    wb.save(str(saida))
    return saida


def _formatar_cabecalho(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER
        cell.alignment = _HEADER_ALIGN


def _ajustar_largura_colunas(ws) -> None:
    for col in ws.columns:
        maior = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(maior + 2, _LARGURA_MAX_COLUNA)
