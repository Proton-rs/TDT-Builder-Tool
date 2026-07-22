import re
from dataclasses import replace
from datetime import date

import openpyxl

from tdt import criador_lista_homogenea, engine_tdt
from tdt.contracts import (
    Descricoes,
    Eletrico,
    Enderecamento,
    ListaHomogenea,
    Modulo,
    SignalRecord,
    TipoSinal,
)
from tdt.dados.lista_padrao import ListaPadraoADMS, SinalPadrao
from tdt.engine_tdt import (
    gerar,
    salvar,
    _nome_hierarquico,
    _expandir_range_row5,
    _eh_alimentador,
    _aor_group,
    _remote_unit,
    _device_mapping,
    _device_mapping_analog,
    disjuntor_por_modulo,
    _normal_value,
    _alias_hoje,
    _coords_comando,
    _measurement_type,
    _fase_saida,
    _output_data_type,
)


def test_coords_comando_duplica_indice_unico():
    assert _coords_comando((1500,)) == "1500;1500"


def test_coords_comando_preserva_multiplos_indices():
    assert _coords_comando((1500, 1501)) == "1500;1501"


def _rec(rid, sigla, indices, direcao="Input", double=False, fase="ABC"):
    return SignalRecord(
        id=rid,
        modulo=Modulo("3", "sheet_name"),
        tipo_sinal=TipoSinal("Discrete", "DoubleBit" if double else "SingleBit", direcao),
        enderecamento=Enderecamento("DNP3", tuple(indices)),
        descricoes=Descricoes(f"{sigla} BRUTO", sigla),
        sigla_sinal=sigla,
        status="decidido",
    )


def _rec_equip(rid, sigla, equipamento, modulo="AL11", indices=(10,), direcao="Input",
               categoria="Discrete"):
    return SignalRecord(
        id=rid,
        modulo=Modulo(modulo, "sheet_name"),
        tipo_sinal=TipoSinal(categoria, "SingleBit", direcao),
        enderecamento=Enderecamento("DNP3", tuple(indices)),
        descricoes=Descricoes(f"{sigla} BRUTO", sigla),
        sigla_sinal=sigla,
        status="decidido",
        eletrico=Eletrico(fase=None, equipamento_alvo=None,
                          nome_equipamento=equipamento, barra=None),
    )


def _lista():
    return ListaHomogenea(
        subestacao="IMA",
        protocolo="DNP3",
        registros=(
            _rec("LT3:1", "DJ", [17], direcao="Input"),
            _rec("LT3:2", "SECC", [100, 101], double=True),
        ),
    )


def test_signal_alias_usa_descricao_do_mapa_v1(template_dnp3_path, lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp,
               alias_v1={"DJ": "DISJUNTOR DE LINHA"})
    ws = wb["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(5, col["Signal Alias"]).value == "DISJUNTOR DE LINHA"
    # SECC fora do mapa -> mantém descrição bruta do cliente
    assert ws.cell(6, col["Signal Alias"]).value == "SECC BRUTO"


def test_signal_alias_sem_mapa_mantem_bruta(template_dnp3_path, lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)  # alias_v1 default None
    ws = wb["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(5, col["Signal Alias"]).value == "DJ BRUTO"


def test_preserva_43_colunas_e_tabela(template_dnp3_path, lista_padrao_path, tmp_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    out = tmp_path / "tdt.xlsx"
    salvar(wb, out)
    rb = openpyxl.load_workbook(out)
    ws = rb["DNP3_DiscreteSignals"]
    assert ws.max_column == 43
    # field names (row 3) e display (row 4) preservados
    assert ws.cell(3, 1).value == "IDOBJ_NAME"
    assert ws.cell(4, 1).value == "Signal Name"
    # ListObject preservado
    assert "DNP3_DiscreteSignals" in ws.tables


def test_escreve_registro_single_bit(template_dnp3_path, lista_padrao_path, tmp_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    # primeira linha de dados = row 5
    # hierarchical: subestacao + modulo.nome("3") + módulo repetido (sem equip) + sigla
    assert ws.cell(5, col["Signal Name"]).value == "IMA_3_3_DJ"
    assert ws.cell(5, col["Input Data Type"]).value == "SingleBit"
    assert ws.cell(5, col["Input Coordinates"]).value == 17
    assert ws.cell(5, col["Direction"]).value == "Read"


def test_readwrite_escreve_output_coords(template_dnp3_path, lista_padrao_path, tmp_path):
    from dataclasses import replace

    base = _rec("LT3:1", "DJ", [5], direcao="InputOutput")
    rw = replace(base, enderecamento=replace(base.enderecamento, indices_saida=(0,)))
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(rw,))
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    ws = gerar(lista, template_dnp3_path, lp)["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(5, col["Direction"]).value == "ReadWrite"
    assert ws.cell(5, col["Output Coordinates"]).value == "0;0"  # comando simples duplica o indice
    # dominio DNP3OutputType (DMSMatchingTemplateInfo): coords iguais -> SingleCoord
    assert ws.cell(5, col["Output Data Type"]).value == "SingleCoord"


def test_output_orfao_escreve_output_coords_nao_input(template_dnp3_path, lista_padrao_path, tmp_path):
    """Comando Output sem par (órfão): endereço próprio vai para Output Coordinates,
    não para Input Coordinates (sinal não tem leitura, só escrita)."""
    cmd = _rec("LT3:3", "ABRIR", [42], direcao="Output")
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(cmd,))
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    ws = gerar(lista, template_dnp3_path, lp)["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(5, col["Direction"]).value == "Write"
    assert ws.cell(5, col["Output Coordinates"]).value == "42;42"  # comando simples duplica o indice
    assert ws.cell(5, col["Input Coordinates"]).value is None


def test_outcoords_comando_s_sem_duplicar(template_dnp3_path, lista_padrao_path, tmp_path):
    from dataclasses import replace

    # comando_duplo=False (comando S) -> Output Coordinates "1504" (sem duplicar)
    base_s = _rec("LT3:10", "81U1", [1539], direcao="InputOutput")
    rec_s = replace(
        base_s,
        tipo_sinal=replace(base_s.tipo_sinal, comando_duplo=False),
        enderecamento=replace(base_s.enderecamento, indices_saida=(1504,)),
    )

    # comando_duplo=True (comando D, default) -> Output Coordinates "1502;1502"
    base_d = _rec("LT3:11", "81U2", [1540], direcao="InputOutput")
    rec_d = replace(
        base_d,
        enderecamento=replace(base_d.enderecamento, indices_saida=(1502,)),
    )

    # MultiCoord -> Input Data Type "MultiCoord"
    base_mc = _rec("LT3:12", "81U3", [1600, 1601], direcao="Input")
    rec_mc = replace(
        base_mc,
        tipo_sinal=replace(base_mc.tipo_sinal, datatype="MultiCoord"),
    )

    lista = ListaHomogenea(
        subestacao="IMA", protocolo="DNP3", registros=(rec_s, rec_d, rec_mc),
    )
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    ws = gerar(lista, template_dnp3_path, lp)["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(5, col["Output Coordinates"]).value == "1504"
    assert ws.cell(6, col["Output Coordinates"]).value == "1502;1502"
    assert ws.cell(7, col["Input Data Type"]).value == "MultiCoord"


def test_output_orfao_multiplos_indices_nao_duplica(template_dnp3_path, lista_padrao_path, tmp_path):
    cmd = _rec("LT3:4", "ABRIR2", [42, 43], direcao="Output")
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(cmd,))
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    ws = gerar(lista, template_dnp3_path, lp)["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(5, col["Output Coordinates"]).value == "42;43"
    # coords distintas (trip;close) -> MultiCoord
    assert ws.cell(5, col["Output Data Type"]).value == "MultiCoord"


def test_double_bit_preserva_dois_indices(template_dnp3_path, lista_padrao_path, tmp_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(6, col["Input Data Type"]).value == "DoubleBit"
    assert ws.cell(6, col["Input Coordinates"]).value == "100;101"


# ── Tarefa 1: table ref ──────────────────────────────────────────────────────

def test_table_ref_comeca_em_a4(template_dnp3_path, lista_padrao_path, tmp_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteSignals"]
    ref = ws.tables["DNP3_DiscreteSignals"].ref
    assert ref.startswith("A4:"), f"Esperado A4:..., obteve {ref}"
    assert ref.endswith("6"), f"Esperado A4:AQ6, obteve {ref}"
    assert ref == "A4:AQ6", f"ref inesperado: {ref}"


# ── Tarefa 2: conditional formatting ─────────────────────────────────────────

def test_cf_expandido_para_todas_linhas(template_dnp3_path, lista_padrao_path, tmp_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteSignals"]
    ranges = [str(cf.sqref) for cf in ws.conditional_formatting]
    assert len(ranges) == 13
    for r in ranges:
        assert "6" in r, f"CF range {r} não cobre row 6"
        assert r.endswith("6"), f"CF range {r} não termina em 6"


# ── Tarefa 3: data validations ──────────────────────────────────────────────

def test_dv_expandido_para_todas_linhas(template_dnp3_path, lista_padrao_path, tmp_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteSignals"]
    sqrefs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert len(sqrefs) == 15  # DVs do template restaurado (4 numericas + 11 de lista)
    for sq in sqrefs:
        for token in sq.split():
            assert token.endswith("6"), f"DV sqref {sq} não cobre row 6"


def test_dv_lista_referencia_sheet_oculta(template_dnp3_path, lista_padrao_path):
    """Dropdowns vêm da DMSMatchingTemplateInfo (não de listas inline)."""
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    listas = [dv for dv in ws.data_validations.dataValidation if dv.type == "list"]
    assert listas, "nenhuma DV de lista na sheet gerada"
    assert all("DMSMatchingTemplateInfo!" in dv.formula1 for dv in listas)
    # colunas-chave têm dropdown cobrindo as linhas de dados
    from openpyxl.utils import get_column_letter
    cobertas = set()
    for dv in listas:
        for token in str(dv.sqref).split():
            cobertas.add(token.split(":")[0].rstrip("0123456789"))
    for nome in ("Phases", "Direction", "Signal Type", "Measurement Type", "Output Data Type"):
        assert get_column_letter(col[nome]) in cobertas, f"{nome} sem dropdown"


def _dominios_dv(wb, sheet):
    """{índice de coluna: valores permitidos} das DVs de lista da sheet."""
    from openpyxl.utils import column_index_from_string
    info = wb["DMSMatchingTemplateInfo"]
    ws = wb[sheet]
    dominios = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        m = re.fullmatch(
            r"DMSMatchingTemplateInfo!\$([A-Z]+)\$(\d+):\$[A-Z]+\$(\d+)", dv.formula1
        )
        assert m, f"formula de DV inesperada: {dv.formula1}"
        ci = column_index_from_string(m.group(1))
        permitidos = {
            info.cell(r, ci).value for r in range(int(m.group(2)), int(m.group(3)) + 1)
        }
        for token in str(dv.sqref).split():
            col = column_index_from_string(token.split(":")[0].rstrip("0123456789"))
            dominios[col] = permitidos
    return dominios


def test_valores_escritos_dentro_do_dominio_das_dvs(template_dnp3_path, lista_padrao_path):
    """Todo valor escrito numa coluna com dropdown pertence ao domínio da
    DMSMatchingTemplateInfo — pega bugs tipo Output Data Type = 'SingleBit'
    (valor de Input) que o ADMS rejeita no import."""
    from dataclasses import replace

    from tdt.contracts import Eletrico

    base = _rec("LT3:1", "DJ", [5], direcao="InputOutput")
    rw = replace(base, enderecamento=replace(base.enderecamento, indices_saida=(0,)))
    cmd = _rec("LT3:3", "ABRIR2", [42, 43], direcao="Output")
    analog_ca = replace(_rec_analog("A:2", "IN62", [30]), eletrico=Eletrico(fase="CA"))
    lista = ListaHomogenea(
        subestacao="IMA", protocolo="DNP3",
        registros=_lista().registros + (rw, cmd) + _lista_analog().registros + (analog_ca,),
    )
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(lista, template_dnp3_path, lp)
    for sheet in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals"):
        ws = wb[sheet]
        dominios = _dominios_dv(wb, sheet)
        for row in range(5, ws.max_row + 1):
            for col, permitidos in dominios.items():
                v = ws.cell(row, col).value
                if v is None or v == "":
                    continue
                if isinstance(v, bool):  # TRUE/FALSE como boolean nativo (igual à TDT real)
                    continue
                assert v in permitidos, (
                    f"{sheet} row {row} col {col}: {v!r} fora do domínio {sorted(map(str, permitidos))}"
                )


def test_output_data_type_dominio():
    assert _output_data_type(None) is None
    assert _output_data_type("") is None
    assert _output_data_type("1504") == "SingleCoord"
    assert _output_data_type("2501;2501") == "SingleCoord"
    assert _output_data_type("650;651") == "MultiCoord"


# ── Tarefa 4: nome hierárquico ──────────────────────────────────────────────

def test_nome_hierarquico_4_partes():
    nome = _nome_hierarquico("GTA", "AL 11", "52-22", None, "43TC")
    assert nome == "GTA_AL11_52-22_43TC"


def test_nome_hierarquico_2_partes_fallback():
    nome = _nome_hierarquico("GTA", None, None, None, "BATA")
    assert nome == "GTA_BATA"


def test_nome_hierarquico_sem_subestacao():
    nome = _nome_hierarquico(None, "LT 1", "52-10", None, "DJ")
    assert nome == "LT1_52-10_DJ"


def test_nome_hierarquico_sigla_only():
    nome = _nome_hierarquico(None, None, None, None, "BATA")
    assert nome == "BATA"


def test_nome_hierarquico_modulo_sem_equip():
    # sem equipamento: repete o módulo
    nome = _nome_hierarquico("SE", "TR 1", None, None, "FA")
    assert nome == "SE_TR1_TR1_FA"


def test_nome_hierarquico_barra_principal_auxiliar():
    nome_p = _nome_hierarquico("GTA", "LT 1", "52-10", "Principal", "DJ")
    assert nome_p == "GTA_LT1_52-10_P_DJ"
    nome_a = _nome_hierarquico("GTA", "LT 1", "52-10", "Auxiliar", "DJ")
    assert nome_a == "GTA_LT1_52-10_A_DJ"


def test_signal_name_hierarquico_no_output(template_dnp3_path, lista_padrao_path, tmp_path):
    """Verifica que o Signal Name no output segue o formato hierárquico."""
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    # _lista() tem subestacao="IMA", modulo.nome="3", sem equipamento -> repete módulo
    assert ws.cell(5, col["Signal Name"]).value == "IMA_3_3_DJ"
    assert ws.cell(6, col["Signal Name"]).value == "IMA_3_3_SECC"


# ── _expandir_range_row5 ────────────────────────────────────────────────────

def test_expandir_range_single():
    assert _expandir_range_row5("A5", 10) == "A5:A10"


def test_expandir_range_range():
    assert _expandir_range_row5("AP5:AQ5", 10) == "AP5:AQ10"


def test_expandir_range_nao_row5():
    assert _expandir_range_row5("A1", 10) == "A1"  # sem mudança


def test_expandir_range_multiplos_tokens():
    # sqref multi-range (ex.: 'B5 Y5' da TDT real) expande cada token
    assert _expandir_range_row5("B5 Y5", 10) == "B5:B10 Y5:Y10"
    assert _expandir_range_row5("AG5 AH5 AN5", 8) == "AG5:AG8 AH5:AH8 AN5:AN8"


# ── Tarefa 3: campos derivados ──────────────────────────────────────────────

def test_eh_alimentador():
    assert _eh_alimentador("AL11") is True
    assert _eh_alimentador("AL 12") is True
    assert _eh_alimentador("3") is False
    assert _eh_alimentador("TR1") is False
    assert _eh_alimentador(None) is False


def test_aor_group():
    assert _aor_group("IMA", True) == "IMA Distr"
    assert _aor_group("IMA", False) == "IMA Trans"
    assert _aor_group(None, True) is None


def test_remote_unit():
    assert _remote_unit("IMA") == "UTR_IMA_1"
    assert _remote_unit(None) is None


def test_device_mapping():
    # correção 16/07: proteção ignora o `nome` recebido e reconstrói do
    # módulo duplicado — nunca cai no equipamento específico da linha.
    assert _device_mapping(
        "IMA_3_20T", "20T", True, subestacao="IMA", modulo_nome="3",
    ) == "IMA_3_3_PROT_20T"
    # spec 2026-07-15: non-protection signals fall directly to equipment, sans sigla
    assert _device_mapping("IMA_3_DJ", "DJ", False) == "IMA_3"
    # sem subestacao/modulo (contexto não disponível) -> só PROT_<SIGLA>
    assert _device_mapping("BATA", "BATA", True) == "PROT_BATA"


def test_device_mapping_protecao_cai_no_modulo_duplicado():
    # correção 16/07: LVA_AL11_52-11_PROT_CAFL estava errado (proteção não
    # deve carregar o equipamento específico, mesmo quando inequívoco); o
    # correto é o módulo duplicado, igual ao fallback de não-proteção sem
    # equipamento.
    assert _device_mapping(
        "LVA_AL11_52-11_CAFL", "CAFL", True, subestacao="LVA", modulo_nome="AL11",
    ) == "LVA_AL11_AL11_PROT_CAFL"


def test_device_mapping_nao_protecao_cai_no_equipamento():
    # spec 2026-07-15: não-proteção cai direto no equipamento, sem sigla.
    assert _device_mapping("LVA_AL11_52-11_CAFL", "CAFL", False) == "LVA_AL11_52-11"


def test_device_mapping_nao_protecao_seccionadora():
    assert _device_mapping("LVA_AL11_89-1_SECC", "SECC", False) == "LVA_AL11_89-1"


def test_device_mapping_nao_protecao_sem_equipamento_cai_no_modulo():
    # nome_hierarquico repete o módulo quando não há equipamento
    assert _device_mapping("LVA_AL11_AL11_MOLA", "MOLA", False) == "LVA_AL11_AL11"


def test_device_mapping_nome_igual_sigla_nao_quebra():
    assert _device_mapping("CAFL", "CAFL", False) == "CAFL"


def test_dm_equipamento_disjuntor_ganha_sufixo_dj():
    # nao-protecao caindo em equipamento 52-1 -> sufixo _DJ (spec 20/07 §A1)
    dm = engine_tdt._device_mapping(
        "IMA_AL11_52-1_DJF1", "DJF1", False,
        subestacao="IMA", modulo_nome="AL11", equipamento="52-1",
    )
    assert dm == "IMA_AL11_52-1_DJ"


def test_dm_equipamento_seccionadora_ganha_sufixo_sec():
    dm = engine_tdt._device_mapping(
        "IMA_AL11_89-4_SECF", "SECF", False,
        subestacao="IMA", modulo_nome="AL11", equipamento="89-4",
    )
    assert dm == "IMA_AL11_89-4_SEC"


def test_dm_equipamento_transformador_ganha_sufixo_tr():
    dm = engine_tdt._device_mapping(
        "IMA_TR1_TR1_86", "86", False,
        subestacao="IMA", modulo_nome="TR1", equipamento="TR1",
    )
    assert dm == "IMA_TR1_TR1_TR"


def test_dm_equipamento_fora_da_whitelist_sem_sufixo():
    # familia_do_id devolve None p/ "RT1" -> comportamento atual preservado
    dm = engine_tdt._device_mapping(
        "IMA_SE_RT1_CAFL", "CAFL", False,
        subestacao="IMA", modulo_nome="SE", equipamento="RT1",
    )
    assert dm == "IMA_SE_RT1"


def test_dm_sem_equipamento_fallback_modulo_duplicado_sem_sufixo():
    dm = engine_tdt._device_mapping(
        "IMA_AL11_AL11_DJF1", "DJF1", False,
        subestacao="IMA", modulo_nome="AL11", equipamento=None,
    )
    assert dm == "IMA_AL11_AL11"


def test_dm_prot_alimentador_usa_disjuntor():
    # decisao 20/07 (supersede correcao 16/07): AL com disjuntor unico ->
    # 2o modulo vira o disjuntor, SEM sufixo (fullbase CNC_AL11_52-22_PROT_51F)
    dm = engine_tdt._device_mapping(
        "CVA_AL11_52-1_CAFL", "CAFL", True,
        subestacao="CVA", modulo_nome="AL11", disjuntor="52-1",
    )
    assert dm == "CVA_AL11_52-1_PROT_CAFL"


def test_dm_prot_alimentador_sem_disjuntor_fallback_modulo():
    dm = engine_tdt._device_mapping(
        "CVA_AL11_AL11_CAFL", "CAFL", True,
        subestacao="CVA", modulo_nome="AL11", disjuntor=None,
    )
    assert dm == "CVA_AL11_AL11_PROT_CAFL"


def test_dm_prot_nao_alimentador_ignora_disjuntor():
    dm = engine_tdt._device_mapping(
        "CVA_TR1BT_TR1BT_CAFL", "CAFL", True,
        subestacao="CVA", modulo_nome="TR1BT", disjuntor="52-7",
    )
    assert dm == "CVA_TR1BT_TR1BT_PROT_CAFL"


def test_dm_prot_complemento_2649():
    # 2649 e Enabled na lista padrao mas mapeia PROT (decisao 20/07)
    sp = SinalPadrao(sigla="2649", descricao="X", signal_type="Enabled",
                     direction=None, mm=None, categoria="Discrete")
    assert engine_tdt._dm_prot("2649", sp) is True


def test_dm_prot_79_fica_fora():
    sp = SinalPadrao(sigla="79", descricao="X", signal_type="ReclosingEnabled",
                     direction=None, mm=None, categoria="Discrete")
    assert engine_tdt._dm_prot("79", sp) is False


def test_dm_prot_relaytrip_preservado():
    sp = SinalPadrao(sigla="TRIP", descricao="X", signal_type="RelayTrip",
                     direction=None, mm=None, categoria="Discrete")
    assert engine_tdt._dm_prot("TRIP", sp) is True


# ── Tarefa 11: revertida 21/07 (decisao do usuario) — Signal Name sempre ──
# modulo-duplicado, mesmo no PROT de alimentador com disjuntor; so o Device
# Mapping usa o disjuntor (Task 2, mantido).

def test_signal_name_prot_alimentador_ignora_disjuntor():
    # decisao 21/07: Signal Name sempre modulo-duplicado; Device Mapping
    # continua usando o disjuntor (Task 2)
    sp = SinalPadrao(sigla="51F", descricao="X", signal_type="RelayTrip",
                      direction=None, mm=None, categoria="Discrete")
    rec = _rec_equip("AL11:1", "51F", None, modulo="AL11")  # sem equipamento explicito
    nome, dm = engine_tdt.dm_registro(rec, "CNC", sp, disjuntor="52-22")
    assert nome == "CNC_AL11_AL11_51F"
    assert dm == "CNC_AL11_52-22_PROT_51F"


def test_signal_name_prot_alimentador_sem_disjuntor_mantem_fallback_modulo():
    sp = SinalPadrao(sigla="51F", descricao="X", signal_type="RelayTrip",
                      direction=None, mm=None, categoria="Discrete")
    rec = _rec_equip("AL11:1", "51F", None, modulo="AL11")
    nome, dm = engine_tdt.dm_registro(rec, "CNC", sp, disjuntor=None)
    assert nome == "CNC_AL11_AL11_51F"


def test_signal_name_nao_prot_ignora_disjuntor():
    # sinal nao-PROT com equipamento explicito: fallback nao deve mexer
    sp = SinalPadrao(sigla="DJF1", descricao="X", signal_type="Enabled",
                      direction=None, mm=None, categoria="Discrete")
    rec = _rec_equip("AL11:1", "DJF1", "52-1", modulo="AL11")
    nome, dm = engine_tdt.dm_registro(rec, "CNC", sp, disjuntor="52-1")
    assert nome == "CNC_AL11_52-1_DJF1"  # ja usava equipamento antes, sem mudanca


# ── Tarefa 12 (contexto historico, ver reversao 21/07 acima) ──

def test_signal_name_prot_alimentador_equipamento_igual_modulo_ignora_disjuntor():
    # decisao 21/07: equipamento-placeholder (== modulo) nao muda o Signal
    # Name; Device Mapping continua usando o disjuntor (Task 2)
    sp = SinalPadrao(sigla="51N", descricao="X", signal_type="Enabled",
                      direction=None, mm=None, categoria="Discrete")
    rec = _rec_equip("AL21:1", "51N", "AL21", modulo="AL21")  # equipamento == modulo
    nome, dm = engine_tdt.dm_registro(rec, "LVA", sp, disjuntor="52-21")
    assert nome == "LVA_AL21_AL21_51N"
    assert dm == "LVA_AL21_52-21_PROT_51N"


def test_signal_name_analog_alimentador_ignora_disjuntor():
    # decisao 21/07: Signal Name analogico tambem sempre modulo-duplicado;
    # Device Mapping analogico continua usando o disjuntor
    lp = _lp_fake({"FREQ": "MeasuredValue"}, categorias={"FREQ": "Analog"},
                   tipos_medicao={"FREQ": "Frequência"})
    rec = _rec_equip("AL21:1", "FREQ", "AL21", modulo="AL21", categoria="Analog")
    valores = engine_tdt._valores_analog(rec, "LVA", lp, disjuntor="52-21")
    assert valores["Signal Name"] == "LVA_AL21_AL21_FREQ"
    assert valores["Device Mapping"] == "LVA_AL21_52-21_DJ"


def test_signal_name_nao_alimentador_mantem_modulo_duplicado():
    # nao-regressao: modulo nao-alimentador (ex. TR1) continua modulo-duplicado
    # mesmo com equipamento-placeholder e disjuntor conhecido
    sp = SinalPadrao(sigla="51N", descricao="X", signal_type="Enabled",
                      direction=None, mm=None, categoria="Discrete")
    rec = _rec_equip("TR1:1", "51N", "TR1", modulo="TR1")
    nome, dm = engine_tdt.dm_registro(rec, "LVA", sp, disjuntor="52-7")
    assert nome == "LVA_TR1_TR1_51N"  # inalterado -- _eh_alimentador(TR1) e False


def test_signal_name_analog_tc_tp_mantem_modulo_mesmo_alimentador():
    # nao-regressao: medida de corrente/tensao (_MEDIDAS_TC/_MEDIDAS_TP) NUNCA
    # usa disjuntor, mesmo em alimentador com equipamento-placeholder
    lp = _lp_fake({"VAB": "MeasuredValue"}, categorias={"VAB": "Analog"},
                   tipos_medicao={"VAB": "Tensão"})
    rec = _rec_equip("AL21:1", "VAB", "AL21", modulo="AL21", categoria="Analog")
    valores = engine_tdt._valores_analog(rec, "LVA", lp, disjuntor="52-21")
    assert valores["Signal Name"] == "LVA_AL21_AL21_VAB"  # modulo-duplicado, TP e excecao


def test_dm_registro_equipamento_especifico_diferente_do_modulo_intocado():
    # nao-regressao: equipamento JA especifico (ex. "52-1", diferente do
    # modulo) nao deve ser sobrescrito pelo disjuntor
    sp = SinalPadrao(sigla="DJF1", descricao="X", signal_type="Enabled",
                      direction=None, mm=None, categoria="Discrete")
    rec = _rec_equip("AL11:1", "DJF1", "52-1", modulo="AL11")
    nome, dm = engine_tdt.dm_registro(rec, "CNC", sp, disjuntor="52-1")
    assert nome == "CNC_AL11_52-1_DJF1"  # inalterado (ja testado na Task 11, reconfirma)


# -- edge cases extras (escrutinio adicional: 2o bug real da mesma forma) --

def test_signal_name_analog_nao_alimentador_mantem_modulo_duplicado():
    # equivalente analogico do teste nao-alimentador acima
    lp = _lp_fake({"FREQ": "MeasuredValue"}, categorias={"FREQ": "Analog"},
                   tipos_medicao={"FREQ": "Frequência"})
    rec = _rec_equip("TR1:1", "FREQ", "TR1", modulo="TR1", categoria="Analog")
    valores = engine_tdt._valores_analog(rec, "LVA", lp, disjuntor="52-7")
    assert valores["Signal Name"] == "LVA_TR1_TR1_FREQ"


def test_signal_name_analog_alimentador_sem_disjuntor_mantem_fallback_modulo():
    # sem disjuntor conhecido (0 ou 2+ no modulo): fallback modulo-duplicado,
    # sem crash
    lp = _lp_fake({"FREQ": "MeasuredValue"}, categorias={"FREQ": "Analog"},
                   tipos_medicao={"FREQ": "Frequência"})
    rec = _rec_equip("AL21:1", "FREQ", "AL21", modulo="AL21", categoria="Analog")
    valores = engine_tdt._valores_analog(rec, "LVA", lp, disjuntor=None)
    assert valores["Signal Name"] == "LVA_AL21_AL21_FREQ"


def test_signal_name_analog_equipamento_especifico_intocado():
    # equipamento ja especifico (diferente do modulo) nao e sobrescrito
    lp = _lp_fake({"FREQ": "MeasuredValue"}, categorias={"FREQ": "Analog"},
                   tipos_medicao={"FREQ": "Frequência"})
    rec = _rec_equip("AL21:1", "FREQ", "89-1", modulo="AL21", categoria="Analog")
    valores = engine_tdt._valores_analog(rec, "LVA", lp, disjuntor="52-21")
    assert valores["Signal Name"] == "LVA_AL21_89-1_FREQ"


def test_medida_usa_disjuntor_tc_tp_false():
    assert engine_tdt._medida_usa_disjuntor("Corrente") is False
    assert engine_tdt._medida_usa_disjuntor("CORRENTE") is False
    assert engine_tdt._medida_usa_disjuntor("Tensão") is False


def test_medida_usa_disjuntor_resto_true():
    assert engine_tdt._medida_usa_disjuntor("Frequência") is True
    assert engine_tdt._medida_usa_disjuntor(None) is True
    assert engine_tdt._medida_usa_disjuntor("") is True


def test_custom_id_duplicado_sem_lista_padrao_comportamento_antigo():
    # compat retroativa: lista_padrao=None -> comportamento antigo, sem disjuntor
    lista = ListaHomogenea(subestacao="CNC", protocolo="DNP3", registros=(
        _rec_equip("AL11:1", "51F", None, modulo="AL11"),
    ))
    restante, revisao = engine_tdt.particionar_custom_id_duplicado(lista)
    assert len(restante.registros) == 1 and revisao == ()


def test_custom_id_duplicado_com_lista_padrao_usa_disjuntor():
    lp = _lp_fake({"51F": "RelayTrip"})
    lista = ListaHomogenea(subestacao="CNC", protocolo="DNP3", registros=(
        _rec_equip("AL11:1", "51F", None, modulo="AL11"),
    ))
    restante, revisao = engine_tdt.particionar_custom_id_duplicado(lista, lp)
    assert len(restante.registros) == 1 and revisao == ()


def test_normal_value():
    sp = SinalPadrao("20T", "", "RelayTrip", None, None, "Discrete",
                     estados_brutos="Transit;NORMAL;ATUADO;Error",
                     valores_scada=(0, 1, 2, 3))
    assert _normal_value(sp) == 1
    assert _normal_value(None) is None
    sp_sem = SinalPadrao("X", "", "Custom", None, None, "Discrete")
    assert _normal_value(sp_sem) is None


def test_alias_hoje_formato_yyyymmdd(monkeypatch):
    import tdt.engine_tdt as eng
    from datetime import date
    class _FakeDate(date):
        @classmethod
        def today(cls):
            return cls(2026, 7, 9)
    monkeypatch.setattr(eng, "date", _FakeDate)
    # TDT real (GTD DNP3_DiscreteAnalog) usa 20260204 — YYYYMMDD, nao MMDDYYYY
    assert eng._alias_hoje() == "20260709"
    assert eng._alias_hoje() != "07092026"


def _rec_analog(rid, sigla, indices):
    return SignalRecord(
        id=rid,
        modulo=Modulo("AL11", "sheet_name"),
        tipo_sinal=TipoSinal("Analog", "SingleBit", "Input"),
        enderecamento=Enderecamento("DNP3", tuple(indices)),
        descricoes=Descricoes(f"{sigla} BRUTO", sigla),
        sigla_sinal=sigla,
        status="decidido",
    )


def _lista_analog():
    return ListaHomogenea(
        subestacao="IMA", protocolo="DNP3",
        registros=(_rec_analog("A:1", "IN61", [20]),),
    )


def test_measurement_type_traduz_pt_para_en():
    sp = SinalPadrao(sigla="IA", descricao="", signal_type="Current", direction=None,
                      mm=None, categoria="Analog", tipo_medicao="Corrente", unidade_exibicao="A")
    assert _measurement_type(sp) == "Current"


def test_measurement_type_none_sem_tipo_medicao():
    sp = SinalPadrao(sigla="IA", descricao="", signal_type="Current", direction=None,
                      mm=None, categoria="Analog")
    assert _measurement_type(sp) is None


def test_kmdf_comprimento_vira_unitless():
    sp = SinalPadrao(sigla="KMDF", descricao="d", signal_type="MeasuredValue",
                     direction=None, mm=None, categoria="Analog", tipo_medicao="Comprimento")
    assert _measurement_type(sp) == "Unitless"


def test_todos_tipos_da_lista_padrao_v6_tem_traducao():
    def sp(t):
        return SinalPadrao(sigla="X", descricao="d", signal_type="MeasuredValue",
                           direction=None, mm=None, categoria="Analog", tipo_medicao=t)
    # os 12 tipos reais da aba AnalogSignals da lista padrao v6
    tipos = ["Corrente", "Tensão", "Potência Ativa", "Potência Reativa", "Temperatura",
             "Comprimento", "Frequência", "Fator de Potência", "Potência Aparente",
             "Ângulo de Tensão", "Umidade", "Discreto"]
    sem = [t for t in tipos if _measurement_type(sp(t)) is None]
    assert sem == [], f"tipos sem traducao: {sem}"


def test_fase_saida_default_abc_para_none():
    assert _fase_saida(None) == "ABC"


def test_fase_saida_fallback_abc_para_valor_invalido():
    assert _fase_saida("F") == "ABC"
    assert _fase_saida("XYZ") == "ABC"


def test_fase_saida_preserva_fase_valida():
    assert _fase_saida("A") == "A"
    assert _fase_saida("ABC") == "ABC"
    assert _fase_saida("N") == "N"


def test_fase_saida_traduz_para_dominio_phasecode():
    assert _fase_saida("CA") == "AC"      # ADMS PhaseCode usa par alfabetico
    assert _fase_saida("AB") == "AB"
    assert _fase_saida(None) == "ABC"     # default preservado
    assert _fase_saida("XY") == "ABC"     # guard de dominio preservado


def test_escreve_sheet_analogica(template_dnp3_path, lista_padrao_path, tmp_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista_analog(), template_dnp3_path, lp)
    ws = wb["DNP3_AnalogSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(5, col["Signal Name"]).value == "IMA_AL11_AL11_IN61"
    assert ws.cell(5, col["Input Coordinates"]).value == 20
    assert ws.cell(5, col["Direction"]).value == "Read"
    assert ws.cell(5, col["Remote Point Type"]).value == "Analog"
    assert ws.cell(5, col["Side"]).value == "None"
    assert ws.cell(5, col["Signal AOR Group"]).value == "IMA Distr"  # AL11 = alimentador
    assert ws.cell(5, col["Remote Point Name"]).value == "IMA_AL11_AL11_IN61"
    # IN61 = "Corrente"/"A" na Lista Padrao -> Measurement Type/Display Unit
    assert ws.cell(5, col["Measurement Type"]).value == "Current"
    assert ws.cell(5, col["Display Unit"]).value == "A"
    # signal_type PT da lista padrao ("Valor Medido") -> dominio EN AnalogSignalType
    assert ws.cell(5, col["Signal Type"]).value == "MeasuredValue"
    # table ref começa em A4
    assert ws.tables["DNP3_AnalogSignals"].ref.startswith("A4:")


def test_discreto_intacto_com_analog(template_dnp3_path, lista_padrao_path, tmp_path):
    """Gerar com registros das duas categorias preenche as duas sheets."""
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    lista = ListaHomogenea(
        subestacao="IMA", protocolo="DNP3",
        registros=_lista().registros + _lista_analog().registros,
    )
    wb = gerar(lista, template_dnp3_path, lp)
    wsd = wb["DNP3_DiscreteSignals"]
    wsa = wb["DNP3_AnalogSignals"]
    cold = {wsd.cell(4, c).value: c for c in range(1, wsd.max_column + 1)}
    cola = {wsa.cell(4, c).value: c for c in range(1, wsa.max_column + 1)}
    assert wsd.cell(5, cold["Signal Name"]).value == "IMA_3_3_DJ"
    assert wsa.cell(5, cola["Signal Name"]).value == "IMA_AL11_AL11_IN61"


def test_campos_novos_no_output(template_dnp3_path, lista_padrao_path, tmp_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    wb = gerar(_lista(), template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteSignals"]
    col = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    # row 5 = "IMA_3_3_DJ" (modulo "3" não é alimentador -> Trans, sem equip repete módulo)
    assert ws.cell(5, col["Side"]).value == "None"
    assert ws.cell(5, col["Output Register"]).value is False
    assert ws.cell(5, col["Remote Point Type"]).value == "Status"
    assert ws.cell(5, col["Remote Point Name"]).value == "IMA_3_3_DJ"
    assert ws.cell(5, col["Signal AOR Group"]).value == "IMA Trans"
    # spec 2026-07-15: non-protection signals (DJ is not RelayTrip) fall directly to equipment
    assert ws.cell(5, col["Device Mapping"]).value == "IMA_3_3"
    assert ws.cell(5, col["Remote Unit"]).value == "UTR_IMA_1"
    assert ws.cell(5, col["Remote Point Custom ID"]).value == "IMA_3_3_DJ_UTR_IMA_1"
    import re as _re
    assert _re.fullmatch(r"\d{8}", ws.cell(5, col["Remote Point Alias"]).value)


def test_tap_sai_na_sheet_discrete_analog(template_dnp3_path):
    """TAP (categoria DiscreteAnalog na lista padrão v7) roteia para a sheet
    DNP3_DiscreteAnalog e não para Discrete/Analog, mesmo com categoria de
    sinal legada 'Discrete' (o roteamento é por sigla, não por tipo_sinal)."""
    from tdt.dados.lista_padrao import ListaPadraoADMS

    lp = ListaPadraoADMS.carregar("docs/Pontos Padrao ADMS_v7.xlsx")
    rec = SignalRecord(
        id="TR1:1",
        modulo=Modulo("TR 1", "sheet_name"),
        tipo_sinal=TipoSinal("Discrete", "SingleBit", "Input"),
        enderecamento=Enderecamento("DNP3", (9,)),
        descricoes=Descricoes("TAP BRUTO", "TAP"),
        sigla_sinal="TAP",
        status="decidido",
    )
    lista = ListaHomogenea(subestacao="GTD", protocolo="DNP3", registros=(rec,))
    wb = gerar(lista, template_dnp3_path, lp)
    ws = wb["DNP3_DiscreteAnalog"]
    nomes = [ws.cell(r, 1).value for r in range(5, 8)]
    assert any(n and str(n).endswith("_TAP") for n in nomes)
    # e NÃO saiu nas sheets Discrete/Analog
    for sn in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals"):
        col1 = [wb[sn].cell(r, 1).value for r in range(5, 8)]
        assert not any(v and str(v).endswith("_TAP") for v in col1)
    # colunas-chave preenchidas conforme o dado real (lista padrão v7)
    hdr = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    linha_tap = next(
        r for r in range(5, 8)
        if ws.cell(r, 1).value and str(ws.cell(r, 1).value).endswith("_TAP")
    )
    assert ws.cell(linha_tap, hdr["Measurement Type"]).value == "Discrete"
    assert ws.cell(linha_tap, hdr["Signal Type"]).value == "TapPosition"
    assert ws.cell(linha_tap, hdr["Remote Point Type"]).value == "Analog"
    assert ws.cell(linha_tap, hdr["Normal Value"]).value == 9
    assert ws.cell(linha_tap, hdr["Input Coordinates"]).value == 9
    assert ws.cell(linha_tap, hdr["Device Mapping"]).value == "GTD_TR1_TR1_COMTAP"


# ── Tarefa 3: gate de unicidade de Remote Point Custom ID ───────────────────

def _rec_gate(id_, sigla, endereco, nome_equipamento=None):
    from tdt.contracts import (Descricoes, Eletrico, Enderecamento, Modulo,
                               SignalRecord, TipoSinal)
    return SignalRecord(
        id=id_,
        modulo=Modulo("LT3", "coluna:MODULO"),
        tipo_sinal=TipoSinal("Discrete"),
        enderecamento=Enderecamento("DNP3", (endereco,)),
        descricoes=Descricoes("43 - CHAVE LOCAL REMOTO", "chave local remoto"),
        eletrico=Eletrico(nome_equipamento=nome_equipamento),
        sigla_sinal=sigla, status="decidido",
    )


def test_gate_custom_id_duplicado_manda_grupo_inteiro_pra_revisao():
    from tdt.contracts import ListaHomogenea
    from tdt.engine_tdt import particionar_custom_id_duplicado
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_gate("LT 3:1", "43LR", 1),        # IMA_LT3_LT3_43LR
        _rec_gate("LT 3:2", "43LR", 22),       # IMA_LT3_LT3_43LR (colide)
        _rec_gate("LT 3:3", "CCFL", 2),        # único -> fica
        _rec_gate("LT 3:4", "43LR", 6, "89-14"),  # IMA_LT3_89-14_43LR -> fica
    ))
    lista_ok, rev = particionar_custom_id_duplicado(lista)
    assert [r.id for r in lista_ok.registros] == ["LT 3:3", "LT 3:4"]
    assert {it.registro.id for it in rev} == {"LT 3:1", "LT 3:2"}
    assert all(it.motivo == "custom_id_duplicado" for it in rev)


def test_gate_sem_duplicatas_nao_mexe():
    from tdt.contracts import ListaHomogenea
    from tdt.engine_tdt import particionar_custom_id_duplicado
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_gate("LT 3:1", "43LR", 1, "89-16"),
        _rec_gate("LT 3:2", "CCFL", 2),
    ))
    lista_ok, rev = particionar_custom_id_duplicado(lista)
    assert lista_ok == lista and rev == ()


def test_colisao_entre_sheets_ganha_motivo_especifico():
    from tdt.contracts import ListaHomogenea
    from tdt.engine_tdt import particionar_custom_id_duplicado
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_gate("BC1:10", "43LR", 1),
        _rec_gate("BC2:12", "43LR", 2),
    ))
    lista_ok, rev = particionar_custom_id_duplicado(lista)
    assert {it.motivo for it in rev} == {"modulo_duplicado_entre_sheets"}


def test_colisao_mesma_sheet_mantem_motivo_atual():
    from tdt.contracts import ListaHomogenea
    from tdt.engine_tdt import particionar_custom_id_duplicado
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_gate("BC1:10", "43LR", 1),
        _rec_gate("BC1:11", "43LR", 2),
    ))
    lista_ok, rev = particionar_custom_id_duplicado(lista)
    assert {it.motivo for it in rev} == {"custom_id_duplicado"}


# ── SP-CVA2 E6.2: gate de endereço duplicado por módulo ─────────────────────

def _rec_end(rid, sigla, direcao, modulo, indices, desc):
    return SignalRecord(
        id=rid,
        modulo=Modulo(modulo, "sheet_name"),
        tipo_sinal=TipoSinal("Discrete", "SingleBit", direcao),
        enderecamento=Enderecamento("DNP3", tuple(indices)),
        descricoes=Descricoes(desc, desc),
        sigla_sinal=sigla,
        status="decidido",
    )


def test_particionar_endereco_duplicado_mesmo_modulo_colide():
    """SP-CVA2 E6.2: dois Inputs Discrete do MESMO módulo com o mesmo índice
    -> grupo inteiro pra revisão (sintoma de direção errada na origem)."""
    a = _rec_end("S1:1", "27", "Input", "M1", [10], "PROT 27 ATUADO")
    b = _rec_end("S1:2", "50BF", "Input", "M1", [10], "ATUADO 50 BF")
    lista = criador_lista_homogenea.montar([a, b], subestacao="SE1")
    lista2, rev = engine_tdt.particionar_endereco_duplicado(lista)
    assert len(lista2.registros) == 0
    assert sorted(it.registro.id for it in rev) == ["S1:1", "S1:2"]
    assert {it.motivo for it in rev} == {"endereco_duplicado"}


def test_particionar_endereco_duplicado_modulos_distintos_nao_colidem():
    """Achado do decision gate (14jul): índice local reusado entre módulos
    (IEDs/linhas) DISTINTOS é endereçamento normal, não colisão — mesmo
    padrão de `_chave` em dc_pairer/normalizador_estrutural."""
    a = _rec_end("S1:1", "FCOM", "Input", "M1", [10], "FALHA COMUNICACAO")
    b = _rec_end("S2:1", "FCOM", "Input", "M2", [10], "FALHA COMUNICACAO")
    lista = criador_lista_homogenea.montar([a, b], subestacao="SE1")
    lista2, rev = engine_tdt.particionar_endereco_duplicado(lista)
    assert not rev and len(lista2.registros) == 2


def test_particionar_endereco_duplicado_espacos_distintos_nao_colidem():
    """Analog@0 e Discrete@0 são espaços distintos; Input@5 e Output@5 idem
    (mesmo módulo em todos os casos)."""
    recs = [
        _rec_end("S1:1", "VAB", "Input", "M1", [0], "TENSAO BARRA AB"),
        _rec_end("S1:2", "27", "Input", "M1", [0], "PROT 27 ATUADO"),
        _rec_end("S1:3", "DJF1", "Output", "M1", [5], "DISJ ABRIR FECHAR"),
        _rec_end("S1:4", "MOLA", "Input", "M1", [5], "MOLA DESCARREGADA"),
    ]
    # VAB precisa ser Analog: ajustar helper/replace da categoria
    recs[0] = replace(recs[0], tipo_sinal=replace(recs[0].tipo_sinal, categoria="Analog"))
    lista = criador_lista_homogenea.montar(recs, subestacao="SE1")
    lista2, rev = engine_tdt.particionar_endereco_duplicado(lista)
    assert not rev and len(lista2.registros) == 4


def test_particionar_endereco_duplicado_indices_saida_no_espaco_out():
    fundido = _rec_end("S1:1", "DJF1", "InputOutput", "M1", [10], "DISJ ABERTO")
    fundido = replace(
        fundido, enderecamento=replace(fundido.enderecamento, indices_saida=(90,))
    )
    outro_cmd = _rec_end("S1:2", "SECC", "Output", "M1", [90], "SEC CARGA ABRIR FECHAR")
    lista = criador_lista_homogenea.montar([fundido, outro_cmd], subestacao="SE1")
    lista2, rev = engine_tdt.particionar_endereco_duplicado(lista)
    assert sorted(it.registro.id for it in rev) == ["S1:1", "S1:2"]


def test_dm_analog_corrente_e_potencias_caem_no_tc():
    assert _device_mapping_analog("LVA", "AL 11", "Corrente", "52-11") == "LVA_AL11_AL11_TC"
    assert _device_mapping_analog("LVA", "AL11", "Potência Ativa", None) == "LVA_AL11_AL11_TC"
    assert _device_mapping_analog("LVA", "AL11", "POTÊNCIA REATIVA", None) == "LVA_AL11_AL11_TC"
    assert _device_mapping_analog("LVA", "AL11", "Potência Aparente", None) == "LVA_AL11_AL11_TC"


def test_dm_analog_tensao_cai_no_tp():
    assert _device_mapping_analog("LVA", "AL11", "Tensão", "52-11") == "LVA_AL11_AL11_TP"


def test_dm_analog_resto_cai_no_disjuntor():
    # KMDF (Comprimento), frequência, FP, temperatura... -> disjuntor do módulo
    assert _device_mapping_analog("LVA", "AL11", "Comprimento", "52-11") == "LVA_AL11_52-11_DJ"  # sufixo _DJ (spec 20/07 §A3)
    assert _device_mapping_analog("LVA", "AL11", "Frequência", "52-11") == "LVA_AL11_52-11_DJ"  # sufixo _DJ (spec 20/07 §A3)
    assert _device_mapping_analog("LVA", "AL11", None, "52-11") == "LVA_AL11_52-11_DJ"  # sufixo _DJ (spec 20/07 §A3)


def test_dm_analog_sem_disjuntor_cai_no_modulo_duplicado():
    assert _device_mapping_analog("LVA", "AL11", "Comprimento", None) == "LVA_AL11_AL11"


def test_dm_analog_disjuntor_ganha_sufixo_dj():
    # fullbase analog: ultimo segmento DJ 4.570x (spec 20/07 §A3)
    dm = engine_tdt._device_mapping_analog("CVA", "AL11", "FREQUÊNCIA", "52-1")
    assert dm == "CVA_AL11_52-1_DJ"


def test_dm_analog_corrente_continua_mod_tc():
    dm = engine_tdt._device_mapping_analog("CVA", "AL11", "CORRENTE", "52-1")
    assert dm == "CVA_AL11_AL11_TC"


def test_dm_analog_sem_disjuntor_fallback_modulo():
    dm = engine_tdt._device_mapping_analog("CVA", "AL11", "FREQUÊNCIA", None)
    assert dm == "CVA_AL11_AL11"


def _rec_eq(rid, modulo, nome_eq):
    return replace(
        _rec(rid, "DJ", [1]),
        modulo=Modulo(modulo, "sheet_name"),
        eletrico=Eletrico(nome_equipamento=nome_eq),
    )


def test_disjuntor_por_modulo():
    regs = [
        _rec_eq("a:1", "AL11", "52-11"),
        _rec_eq("a:2", "AL11", "89-1"),   # seccionadora não conta
        _rec_eq("a:3", "AL12", "52-12"),
        _rec_eq("a:4", "AL12", "24-1"),   # 2 disjuntores -> ambíguo
        _rec_eq("a:5", "AL13", None),
    ]
    disj = disjuntor_por_modulo(regs)
    assert disj["AL11"] == "52-11"
    assert disj["AL12"] is None
    assert disj.get("AL13") is None


# ── Tarefa 5: aviso 43LR sem 43TC (spec §B2) ────────────────────────────────

def test_43lr_sem_43tc_avisa():
    regs = (
        _rec_equip("AL11:1", "43LR", "52-1"),
    )
    assert engine_tdt.dispositivos_43lr_sem_43tc(regs) == ("AL11/52-1",)


def test_43lr_com_43tc_nao_avisa():
    regs = (
        _rec_equip("AL11:1", "43LR", "52-1"),
        _rec_equip("AL11:2", "43TC", "52-1"),
    )
    assert engine_tdt.dispositivos_43lr_sem_43tc(regs) == ()


def test_43tc_sozinho_nao_avisa():
    regs = (_rec_equip("AL11:1", "43TC", "52-1"),)
    assert engine_tdt.dispositivos_43lr_sem_43tc(regs) == ()


# ── Tarefa 6: gate tipo duplicado por dispositivo (spec §B3) ────────────────

def _lp_fake(siglas_tipos: dict[str, str], categorias: dict[str, str] | None = None,
             tipos_medicao: dict[str, str] | None = None):
    categorias = categorias or {}
    tipos_medicao = tipos_medicao or {}
    class _LP:
        def por_sigla(self, sigla):
            st = siglas_tipos.get(sigla)
            if st is None:
                return None
            return SinalPadrao(sigla=sigla, descricao="X", signal_type=st,
                               direction=None, mm=None,
                               categoria=categorias.get(sigla, "Discrete"),
                               tipo_medicao=tipos_medicao.get(sigla))
    return _LP()


def test_tipo_duplicado_mesmo_dispositivo_vai_para_revisao():
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_equip("AL11:1", "43TC", "52-1"),   # Local
        _rec_equip("AL11:2", "43XY", "52-1"),   # Local tambem -> conflito
    ))
    lp = _lp_fake({"43TC": "Local", "43XY": "Local"})
    restante, revisao = engine_tdt.particionar_tipo_duplicado(lista, lp)
    assert len(restante.registros) == 0
    assert {it.motivo for it in revisao} == {"tipo_duplicado_dispositivo"}


def test_tipo_duplicado_custom_isento():
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_equip("AL11:1", "CAB1", "52-1"),
        _rec_equip("AL11:2", "CAB2", "52-1"),
    ))
    lp = _lp_fake({"CAB1": "Custom", "CAB2": "Custom"})
    restante, revisao = engine_tdt.particionar_tipo_duplicado(lista, lp)
    assert len(restante.registros) == 2 and revisao == ()


def test_tipo_duplicado_prot_isento():
    # dois RelayTrip no mesmo modulo: caem no PROT -> repetido e valido
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_equip("AL11:1", "51F", "52-1"),
        _rec_equip("AL11:2", "50F1", "52-1"),
    ))
    lp = _lp_fake({"51F": "RelayTrip", "50F1": "RelayTrip"})
    restante, revisao = engine_tdt.particionar_tipo_duplicado(lista, lp)
    assert len(restante.registros) == 2 and revisao == ()


def test_tipo_duplicado_dispositivos_distintos_isento():
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_equip("AL11:1", "43TC", "52-1"),
        _rec_equip("AL11:2", "43TC", "89-4"),   # outro equipamento
    ))
    lp = _lp_fake({"43TC": "Local"})
    restante, revisao = engine_tdt.particionar_tipo_duplicado(lista, lp)
    assert len(restante.registros) == 2 and revisao == ()


def test_tipo_duplicado_analogico_isento():
    # Fix da revisao 2026-07-20: "Valor Medido" e generico demais (55/62
    # siglas analogicas do catalogo real compartilham esse signal_type) --
    # sinal Analog nunca entra no agrupamento deste gate, mesmo colidindo
    # em (dm, signal_type) se fosse Discrete. `rec.tipo_sinal.categoria`
    # fica "Discrete" (default do estruturador -- replica o caso real SAN2,
    # sheet "Analogicos" sem marcador de secao/coluna TIPO explicita, ver
    # nota do docstring do gate); e o `sp.categoria` do CATALOGO (fonte
    # confiavel) que precisa vencer e isentar o par.
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_equip("AL11:1", "IA", "52-1", categoria="Discrete"),
        _rec_equip("AL11:2", "IB", "52-1", categoria="Discrete"),
    ))
    lp = _lp_fake({"IA": "Valor Medido", "IB": "Valor Medido"},
                  categorias={"IA": "Analog", "IB": "Analog"})
    restante, revisao = engine_tdt.particionar_tipo_duplicado(lista, lp)
    assert len(restante.registros) == 2 and revisao == ()


def test_tipo_duplicado_43lr_sempre_custom_mesmo_com_catalogo_desatualizado():
    # Fix da revisao 2026-07-20: fullbase real (424/424 dispositivos com
    # 43LR+43TC) mostra 43LR=Custom sempre; a lista padrao v2 (fixture/
    # catalogo desatualizado usado por test_integracao_san2.py) classifica
    # errado como "Local", o que colidiria com 43TC. O gate precisa tratar
    # 43LR como Custom (isento) independente do que o catalogo diga.
    lista = ListaHomogenea(subestacao="IMA", protocolo="DNP3", registros=(
        _rec_equip("AL11:1", "43LR", "52-1"),
        _rec_equip("AL11:2", "43TC", "52-1"),
    ))
    lp = _lp_fake({"43LR": "Local", "43TC": "Local"})  # catalogo stale/incorreto
    restante, revisao = engine_tdt.particionar_tipo_duplicado(lista, lp)
    assert len(restante.registros) == 2 and revisao == ()
