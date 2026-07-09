import openpyxl
import pytest

from bench.gate_tdt_real import carregar_siglas_por_endereco, comparar


def _tdt_fake(path, linhas, sheet="DNP3_DiscreteSignals", col_addr=31):
    """linhas = list[(signal_name, incoords)] na sheet dada.

    col_addr 0-based do Input Coordinates: DiscreteSignals=31, DiscreteAnalog=34,
    AnalogSignals=47 (índices reais do template)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for _ in range(4):  # rows 1-4 cabeçalho
        ws.append([])
    for nome, addr in linhas:
        row = [None] * (col_addr + 1)
        row[0] = nome
        row[col_addr] = addr
        ws.append(row)
    wb.save(path)


def test_carregar_chaveia_por_sheet_e_endereco(tmp_path):
    f = tmp_path / "t.xlsx"
    _tdt_fake(f, [("GTD_AL13_52-13_CCMO", 1706), ("GTD_LTGTA_89-2_DSEC", 16)])
    d = carregar_siglas_por_endereco(str(f))
    assert d == {
        ("DNP3_DiscreteSignals", 1706): ("GTD_AL13_52-13_CCMO", "CCMO"),
        ("DNP3_DiscreteSignals", 16): ("GTD_LTGTA_89-2_DSEC", "DSEC"),
    }


def test_comparar_conta_iguais_e_lista_divergencias(tmp_path):
    real = tmp_path / "real.xlsx"; nosso = tmp_path / "nosso.xlsx"
    _tdt_fake(real, [("GTD_AL13_52-13_CCMO", 100), ("GTD_LTGTA_52-1_BBFC", 7)])
    _tdt_fake(nosso, [("GTD_AL13_52-13_CCMO", 100), ("GTD_LTGTA_52-1_LIGAR", 7)])
    r = comparar(str(nosso), str(real))
    assert r.comum == 2
    assert r.iguais == 1
    assert r.pct == pytest.approx(50.0)
    assert (7, "BBFC", "LIGAR", "GTD_LTGTA_52-1_BBFC") in r.divergencias


def test_discreto_e_analogico_no_mesmo_endereco_nao_colidem(tmp_path):
    """Endereço 0 num discreto e num analógico são entradas distintas — a chave
    (sheet, addr) impede que um sobrescreva o outro (o bug do keying achatado)."""
    f = tmp_path / "misto.xlsx"
    wb = openpyxl.Workbook()
    ws_d = wb.active
    ws_d.title = "DNP3_DiscreteSignals"
    for _ in range(4):
        ws_d.append([])
    row = [None] * 32
    row[0] = "GTD_TR1_52-1_DJF1"; row[31] = 0
    ws_d.append(row)
    ws_a = wb.create_sheet("DNP3_AnalogSignals")
    for _ in range(4):
        ws_a.append([])
    row = [None] * 48
    row[0] = "GTD_TR1_TR1_IA"; row[47] = 0  # mesmo endereço 0, outra categoria
    ws_a.append(row)
    wb.save(f)

    d = carregar_siglas_por_endereco(str(f))
    assert d[("DNP3_DiscreteSignals", 0)] == ("GTD_TR1_52-1_DJF1", "DJF1")
    assert d[("DNP3_AnalogSignals", 0)] == ("GTD_TR1_TR1_IA", "IA")


def test_le_analogicos_da_coluna_47(tmp_path):
    """AnalogSignals: Input Coordinates na coluna 47 (antes lia 31 = zero)."""
    f = tmp_path / "analog.xlsx"
    _tdt_fake(f, [("GTD_TR1_TR1_IA", 5)], sheet="DNP3_AnalogSignals", col_addr=47)
    d = carregar_siglas_por_endereco(str(f))
    assert d == {("DNP3_AnalogSignals", 5): ("GTD_TR1_TR1_IA", "IA")}


def test_comparar_identico_da_100pct():
    real = "docs/TDT/exportTDT_UTR_GTD_1_20260626.xlsx"
    r = comparar(real, real)
    assert r.pct == 100.0
    assert r.iguais == r.comum > 0
