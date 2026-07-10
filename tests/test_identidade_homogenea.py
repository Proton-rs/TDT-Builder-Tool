"""Resolvedor de identidade módulo/equipamento do caminho homogêneo
(spec 2026-07-10-sp-import-ima-erros). Oracle: coluna NOME da lista IMA."""
from pathlib import Path

import pytest

from tdt.normalizacao.identidade_homogenea import Identidade, extrair_bloco, resolver

_DOCS = Path(__file__).resolve().parent.parent / "docs"


# --- extrair_bloco -----------------------------------------------------------

def test_extrair_bloco_aceita_numero_operativo_classe_tensao_e_numero():
    rows = [
        ("MÓDULO - TRANSFORMADOR",),
        ("EQUIPAMENTO", "NÚMERO OPERATIVO / MNEMNICO"),
        ("MÓDULO  ", "6"),
        ("DJ AT", "52-6"),
        (),
        ("EQUIPAMENTO", "CLASSE DE TENSÃO"),
        ("BP AT", "69"),
        ("EQUIPAMENTO", "NÚMERO"),
        ("RET", "1"),
        ("header",),
    ]
    bloco = extrair_bloco(rows, header_idx=9)
    assert bloco["MODULO"] == "6"
    assert bloco["DJ AT"] == "52-6"
    assert bloco["BP AT"] == "69"
    assert bloco["RET"] == "1"


def test_extrair_bloco_ausente_devolve_vazio():
    assert extrair_bloco([("qualquer",), ("coisa",)], header_idx=2) == {}


# --- resolver: módulo --------------------------------------------------------

def test_modulo_geral_concatena_numero_do_bloco():
    ident = resolver({"MODULO": "3", "DJ": "52-3"}, "LT 3", "LT", "TC")
    assert ident.modulo == "LT3"
    assert ident.equipamento is None  # TC sem mnemônico -> repete módulo


def test_modulo_ja_numerado_mantem_comportamento_atual():
    ident = resolver({"MODULO": "99"}, "LT1", "LT 1", "DJ")
    assert ident.modulo == "LT 1"  # guarda: "<letras> <número>" não concatena


def test_modulo_com_digito_interno_concatena():
    # TSA 1 real: coluna TSA_P1 + bloco MÓDULO 40 -> cliente usa TSA_P140
    ident = resolver({"MODULO": "40"}, "TSA 1", "TSA_P1", "TSA")
    assert ident.modulo == "TSA_P140"


def test_modulo_lado_at_bt_compoe_prefixo_da_sheet():
    bloco = {"MODULO": "6", "DJ AT": "52-6", "DJ BT": "52-19"}
    assert resolver(bloco, "TR 1", "AT", "TC").modulo == "TR6AT"
    assert resolver(bloco, "TR 1", "BT", "TC").modulo == "TR6BT"
    assert resolver(bloco, "TR 1", "TR", "TR").modulo == "TR6"


def test_modulo_bp_compoe_classe_de_tensao():
    bloco = {"BP AT": "69", "BP BT1": "13.8", "BP BT2": "13.8"}
    assert resolver(bloco, "BARRA", "BP", "AT").modulo == "BP69"
    assert resolver(bloco, "BARRA", "BP1", "BT").modulo == "BP113.8"
    assert resolver(bloco, "BARRA", "BP2", "BT").modulo == "BP213.8"


def test_modulo_com_entrada_propria_no_bloco():
    # RET 1 real: coluna MÓDULO = TSA, bloco TSA->1 -> módulo TSA1
    ident = resolver({"TSA": "1", "RET": "1"}, "RET 1", "TSA", "RET")
    assert ident.modulo == "TSA1"
    assert ident.equipamento == "RET1"  # valor numérico concatena ao rótulo


# --- resolver: equipamento (segmento do meio) --------------------------------

def test_equipamento_mnemonico_com_hifen_usado_direto():
    bloco = {"MODULO": "3", "DJ": "52-3", "SECC": "89-16"}
    assert resolver(bloco, "LT 3", "LT", "SECC").equipamento == "89-16"
    assert resolver(bloco, "LT 3", "LT", "DJ").equipamento == "52-3"


def test_equipamento_lookup_por_lado():
    bloco = {"MODULO": "6", "DJ AT": "52-6", "DJ BT": "52-19"}
    assert resolver(bloco, "TR 1", "AT", "DJ").equipamento == "52-6"
    assert resolver(bloco, "TR 1", "BT", "DJ").equipamento == "52-19"


def test_sufixo_rele_p_a():
    bloco = {"MODULO": "3", "DJ": "52-3"}
    assert resolver(bloco, "LT 3", "LT", "DJ_P").equipamento == "52-3_P"
    assert resolver(bloco, "LT 3", "LT", "DJ_A").equipamento == "52-3_A"
    # sem mnemônico do base: sufixo vai no módulo (LT_P -> LT3_P)
    assert resolver(bloco, "LT 3", "LT", "LT_P").equipamento == "LT3_P"


def test_sem_bloco_fallback_sem_equipamento():
    ident = resolver({}, "AL 11", "AL", "DJ")
    assert ident.modulo == "AL"
    assert ident.equipamento is None


# --- oracle: coluna NOME da lista IMA (1404 linhas) ---------------------------

_EXCECOES_CLIENTE = {
    "IMA_BP169_BP169_FGOO", "IMA_BP269_BP269_FGOO",
    "IMA_TSA_RET_NEGT", "IMA_TSA_RET_POST",
}


def test_oracle_nome_cliente_ima():
    caminho = _DOCS / "input_homogeneo_IMA.xlsx"
    if not caminho.exists():
        pytest.skip("lista IMA não disponível")
    import unicodedata

    import openpyxl

    def norm(v):
        if v is None:
            return ""
        s = "".join(c for c in unicodedata.normalize("NFKD", str(v))
                    if not unicodedata.combining(c))
        return " ".join(s.upper().split())

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    total, iguais, divergentes = 0, 0, set()
    for sh in wb.sheetnames:
        rows = [r for r in wb[sh].iter_rows(values_only=True)]
        ih = next((i for i, r in enumerate(rows[:30])
                   if {"NOME", "SIGLA SINAL"} <= {norm(c) for c in r if c}), None)
        if ih is None:
            continue
        hdr = [norm(c) for c in rows[ih]]
        ix = {n: hdr.index(n) for n in
              ("UTILIZADO?", "SUBESTACAO", "MODULO", "EQUIPAMENTO",
               "SIGLA SINAL", "NOME")}
        bloco = extrair_bloco(rows, ih)
        for r in rows[ih + 1:]:
            if norm(r[ix["UTILIZADO?"]]) != "SIM":
                continue
            sigla = str(r[ix["SIGLA SINAL"]] or "").strip()
            nome_cli = str(r[ix["NOME"]] or "").strip()
            if not sigla or not nome_cli:
                continue
            total += 1
            ident = resolver(bloco, sh, norm(r[ix["MODULO"]]),
                             norm(r[ix["EQUIPAMENTO"]]))
            se = str(r[ix["SUBESTACAO"]] or "").strip()
            mod = ident.modulo.replace(" ", "")
            calc = f"{se}_{mod}_{ident.equipamento or mod}_{sigla}"
            if calc == nome_cli:
                iguais += 1
            else:
                divergentes.add(nome_cli)
    wb.close()
    assert total == 1404
    assert divergentes == _EXCECOES_CLIENTE
    assert iguais == 1400
