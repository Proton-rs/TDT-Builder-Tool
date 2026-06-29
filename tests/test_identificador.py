import openpyxl

from tdt.config import Config
from tdt.analise.identificador import classificar, ler_rows


def test_detecta_sheets_de_dados_no_nao_homogeneo(docs):
    wb = openpyxl.load_workbook(docs / "input_nao_homogeneo_1.xlsx", read_only=True, data_only=True)
    rota = classificar(wb, override="nao-homogeneo")
    assert rota.homogeneo is False
    assert "01F1_GTA_P" in rota.sheets_dados
    # a capa (sem colunas de dados) não entra
    assert "Capa" not in rota.sheets_dados
    wb.close()


def test_sheets_excluidas_da_config_ficam_fora_mesmo_passando_pelo_filtro_de_conteudo(docs):
    # "Consistidos" passa o filtro de conteúdo (_eh_sheet_dados) mas é um
    # índice/consolidação cross-módulo (mistura LT_GTA/87B-AT/TR1_MT na mesma
    # sheet) -- medição real 28-jun. Exclusão por nome via Config.sheets_excluidas.
    wb = openpyxl.load_workbook(docs / "input_nao_homogeneo_1.xlsx", read_only=True, data_only=True)
    rota = classificar(wb, override="nao-homogeneo", config=Config())
    assert "Consistidos" not in rota.sheets_dados
    assert "01F1_GTA_P" in rota.sheets_dados  # sheet de dados real não é afetada
    wb.close()


def test_sem_config_classificar_mantem_comportamento_atual(docs):
    # Compat: config é opcional -- sem ela, nenhuma exclusão por nome (só o
    # filtro de conteúdo de sempre).
    wb = openpyxl.load_workbook(docs / "input_nao_homogeneo_1.xlsx", read_only=True, data_only=True)
    rota = classificar(wb, override="nao-homogeneo")
    assert "Consistidos" in rota.sheets_dados
    wb.close()


def test_override_respeitado(docs):
    wb = openpyxl.load_workbook(docs / "input_nao_homogeneo_1.xlsx", read_only=True, data_only=True)
    assert classificar(wb, override="homogeneo").homogeneo is True
    wb.close()


def test_ler_rows_retorna_tuplas(docs):
    wb = openpyxl.load_workbook(docs / "input_nao_homogeneo_1.xlsx", read_only=True, data_only=True)
    rows = ler_rows(wb["01F1_GTA_P"])
    assert len(rows) > 10
    assert isinstance(rows[0], tuple)
    wb.close()
