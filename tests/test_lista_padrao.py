from pathlib import Path

from tdt.defaults import DEFAULT_LISTA
from tdt.dados.lista_padrao import ListaPadraoADMS
from tdt.dados.lista_padrao import descricoes_por_sigla
from tdt.dados.lista_padrao import validar_mm


def test_default_lista_aponta_para_v8_com_djf1_enriquecido():
    assert Path(DEFAULT_LISTA).name == "Pontos Padrao ADMS_v8.xlsx"
    lp = ListaPadraoADMS.carregar(DEFAULT_LISTA)
    sp = lp.por_sigla("DJF1")
    assert sp is not None
    assert "LIGADO" in sp.descricao.upper()
    assert "DESLIGADO" in sp.descricao.upper()


def test_carrega_discretos_e_analogicos(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    assert len(lp.discretos) > 100
    assert len(lp.analogicos) > 10


def test_sinal_discreto_conhecido(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    cmde = lp.por_sigla("CMDE")
    assert cmde is not None
    assert cmde.categoria == "Discrete"
    assert cmde.signal_type == "Custom"
    assert cmde.direction == "ReadWrite"
    assert cmde.mm.startswith("CMD_CEEE@CMD_RGE")


def test_sinal_analogico_conhecido(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    in61 = lp.por_sigla("IN61")
    assert in61.categoria == "Analog"
    assert in61.signal_type == "Valor Medido"


def test_siglas_inclui_discretos_e_analogicos(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    siglas = lp.siglas
    assert isinstance(siglas, frozenset)
    assert "CMDE" in siglas  # discreto
    assert "IN61" in siglas  # analógico
    assert len(siglas) == len({s.sigla for s in (*lp.discretos, *lp.analogicos)})


def test_siglas_normaliza_maiusculas(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    assert all(s == s.upper() for s in lp.siglas)


def test_ignora_linhas_invalidas(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    # nenhuma sigla vazia ou "#N/A" deve entrar
    siglas = {s.sigla for s in lp.discretos}
    assert "" not in siglas
    assert "#N/A" not in siglas
    assert None not in siglas


def test_sinal_discreto_tem_estados_e_valores(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    # Na v2, a sheet DiscreteSignals não tem nenhuma linha com FUNÇÃO/VALOR
    # preenchidos (colunas existem no header, mas vazias em produção) — ao
    # contrário da v1, onde "20T" trazia
    # "Transit;NORMAL;ATUADO;Error" / "0;1;2;3". Mantém a leitura validada
    # (campo None / tupla vazia quando a fonte não traz dado), sem inventar
    # valor que a v2 não tem.
    sp = lp.por_sigla("20T")
    assert sp is not None
    assert sp.estados_brutos is None
    assert sp.valores_scada == ()


def test_analogico_sem_estados(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    in61 = lp.por_sigla("IN61")
    assert in61.estados_brutos is None
    assert in61.valores_scada == ()


def test_le_tipo_medicao_e_unidade_exibicao_de_analog_signals(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    in61 = lp.por_sigla("IN61")
    assert in61.tipo_medicao == "Corrente"
    assert in61.unidade_exibicao == "A"


def test_le_type_severidade_dos_discretos(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    assert lp.por_sigla("TEA").type_severidade == "PROT"
    assert lp.por_sigla("DJF1").type_severidade == "DJ"
    assert lp.por_sigla("CMDE").type_severidade == "ALARMES PREDIAIS/VF/GRUPO"


def test_analogico_sem_type_severidade(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    assert lp.por_sigla("IN61").type_severidade is None


def test_descricoes_por_sigla_le_v1(docs):
    m = descricoes_por_sigla(str(docs / "Pontos Padrao ADMS_v1.xlsx"))
    assert m["TEA"] == "49 - ALARME TEMPERATURA ENROLAMENTO"
    assert "IN61" in m  # analógicos também entram


def test_descricoes_por_sigla_arquivo_ausente_devolve_vazio(tmp_path):
    assert descricoes_por_sigla(str(tmp_path / "nao_existe.xlsx")) == {}


def test_carrega_aba_discrete_analog(docs):
    lp = ListaPadraoADMS.carregar(docs / "Pontos Padrao ADMS_v7.xlsx")
    tap = lp.por_sigla("TAP")
    assert tap is not None
    assert tap.categoria == "DiscreteAnalog"
    assert tap.signal_type == "TapPosition"
    assert tap.normal_value == 9
    assert tap.remote_point_type == "Analog"
    assert tap.device_mapping_ref == "COMTAP"
    assert tap.aplicabilidade == "TRANSFORMADOR"
    assert "TAP" in lp.siglas


def test_lista_v2_sem_aba_nova_carrega_vazio(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    assert lp.discrete_analog == ()
    assert lp.por_sigla("TAP") is None


def _wb_minimo():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DiscreteSignals"
    ws.append(["SINAL"])
    ws.append(["CMDE"])
    wb.create_sheet("AnalogSignals").append(["SINAL"])
    return wb


def test_carregar_le_sheet_de_para(tmp_path):
    wb = _wb_minimo()
    ws = wb.create_sheet("DE->PARA")
    ws.append(["SINAL", "DESCRIÇÃO NOVA"])
    ws.append(["90", "R90"])
    ws.append(["21_1", "21Z1"])
    p = tmp_path / "lp.xlsx"
    wb.save(p)
    lp = ListaPadraoADMS.carregar(p)
    assert lp.de_para == {"90": "R90", "21_1": "21Z1"}


def test_carregar_sem_sheet_de_para(tmp_path):
    wb = _wb_minimo()
    p = tmp_path / "lp.xlsx"
    wb.save(p)
    lp = ListaPadraoADMS.carregar(p)
    assert lp.de_para == {}


def test_carregar_le_sheet_message_mapping(tmp_path):
    wb = _wb_minimo()
    ws = wb.create_sheet("Message Mapping")
    ws.append(["Name"])
    ws.append(["MM_A"])
    ws.append(["MM_B"])
    p = tmp_path / "lp.xlsx"
    wb.save(p)
    lp = ListaPadraoADMS.carregar(p)
    assert lp.catalogo_mm == frozenset({"MM_A", "MM_B"})


def test_carregar_sem_sheet_message_mapping(tmp_path):
    wb = _wb_minimo()
    p = tmp_path / "lp.xlsx"
    wb.save(p)
    lp = ListaPadraoADMS.carregar(p)
    assert lp.catalogo_mm == frozenset()


def test_validar_mm_com_sigla_mm_ausente_do_catalogo(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DiscreteSignals"
    ws.append(["SINAL", "MM"])
    ws.append(["CMDE", "MM_X"])
    wb.create_sheet("AnalogSignals").append(["SINAL"])
    ws_mm = wb.create_sheet("Message Mapping")
    ws_mm.append(["Name"])
    ws_mm.append(["MM_A"])
    ws_mm.append(["MM_B"])
    p = tmp_path / "lp.xlsx"
    wb.save(p)
    lp = ListaPadraoADMS.carregar(p)
    avisos = validar_mm(lp)
    assert avisos == ["MM MM_X da sigla CMDE ausente do catálogo"]


def test_validar_mm_com_catalogo_vazio_nao_gera_avisos(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DiscreteSignals"
    ws.append(["SINAL", "MM"])
    ws.append(["CMDE", "MM_X"])
    wb.create_sheet("AnalogSignals").append(["SINAL"])
    p = tmp_path / "lp.xlsx"
    wb.save(p)
    lp = ListaPadraoADMS.carregar(p)
    assert lp.catalogo_mm == frozenset()
    assert validar_mm(lp) == []


def test_le_severidade_dos_discretos(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DiscreteSignals"
    ws.append(["SINAL", "SEVERIDADE"])
    ws.append(["CMDE", "Severidade 4"])
    ws.append(["DJF1", None])
    wb.create_sheet("AnalogSignals").append(["SINAL"])
    p = tmp_path / "lp.xlsx"
    wb.save(p)
    lp = ListaPadraoADMS.carregar(p)
    assert lp.por_sigla("CMDE").severidade == "Severidade 4"
    assert lp.por_sigla("DJF1").severidade is None


def test_le_fases_dos_analogicos(lista_padrao_path):
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    assert lp.por_sigla("IN61").fases == "N"


def test_analogico_sem_fases_e_none(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DiscreteSignals"
    ws.append(["SINAL"])
    wsa = wb.create_sheet("AnalogSignals")
    wsa.append(["SINAL", "FASES"])
    wsa.append(["IA", "L1"])
    wsa.append(["IB", None])
    p = tmp_path / "lp.xlsx"
    wb.save(p)
    lp = ListaPadraoADMS.carregar(p)
    assert lp.por_sigla("IA").fases == "L1"
    assert lp.por_sigla("IB").fases is None
