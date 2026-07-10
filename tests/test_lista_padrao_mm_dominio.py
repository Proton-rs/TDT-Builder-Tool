"""Domínio: todo MM da lista padrão deve existir no catálogo real do ADMS
(spec 2026-07-10 §2.3). Whitelist = refs quebradas conhecidas ainda não
corrigidas (spec §4 — AJ*/DSAB/VFAR, ZERO, ICC, CDCO, CCIC)."""
from pathlib import Path

import openpyxl
import pytest

from tdt.defaults import DEFAULT_LISTA

_FIXTURE = Path(__file__).resolve().parent / "fixtures" / "mm_catalogo_real.txt"

# Pendências documentadas na spec §4 — remover daqui conforme a v8+n corrigir.
_WHITELIST = {
    "null@null___DESABILITADO@HABILITADO___Custom_S_TS_SV",   # AJ*/DSAB/VFAR
    "ZERAR@null___null@null___Custom_S_TC_SS",                # ZERO
    "RESET@null___null@null___Custom_S_TC_SS",                # ICC
    "MESTRE@INDIVIDUAL@COMANDO___MESTRE@INDIVIDUAL@COMANDO___Parallel___admsINV_D_TC",  # CDCO
    "CMD_RGE@null___RGE@CPFLT___Custom_S_TC_SS_CPFLT",        # CCIC
}


def test_mm_da_lista_padrao_existe_no_catalogo_real():
    if not Path(DEFAULT_LISTA).exists():
        pytest.fail(f"lista padrão não encontrada: {DEFAULT_LISTA}")
    catalogo = set(_FIXTURE.read_text(encoding="utf-8").splitlines())
    assert len(catalogo) > 400  # sanidade: fixture não truncada

    wb = openpyxl.load_workbook(DEFAULT_LISTA, read_only=True, data_only=True)
    ws = wb["DiscreteSignals"]
    linhas = ws.iter_rows(values_only=True)
    hdr = [str(c).strip().upper() if c else "" for c in next(linhas)]
    i_sig, i_mm = hdr.index("SINAL"), hdr.index("MM")
    fora = {}
    for r in linhas:
        sigla = str(r[i_sig]).strip() if r[i_sig] else ""
        mm = str(r[i_mm]).strip() if r[i_mm] else ""
        if sigla and mm and mm not in catalogo and mm not in _WHITELIST:
            fora.setdefault(mm, []).append(sigla)
    wb.close()
    assert fora == {}, f"MMs fora do catálogo real: {fora}"


def test_refs_corrigidas_na_v8():
    wb = openpyxl.load_workbook(DEFAULT_LISTA, read_only=True, data_only=True)
    ws = wb["DiscreteSignals"]
    linhas = ws.iter_rows(values_only=True)
    hdr = [str(c).strip().upper() if c else "" for c in next(linhas)]
    i_sig, i_mm = hdr.index("SINAL"), hdr.index("MM")
    mm = {str(r[i_sig]).strip(): str(r[i_mm]).strip() if r[i_mm] else ""
          for r in linhas if r[i_sig]}
    wb.close()
    assert mm["43LR"] == "null@null___REMOTO@LOCAL___Custom_S_TS_SS"
    for s in ("81U1", "81U2", "81U3", "81U4", "81U5"):
        assert mm[s] == "null@ATIVAR___DESATIVADO@ATIVADO___Custom_S_TC_SS"
