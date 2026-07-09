"""Regressao do shim openpyxl PatternFill extLst (py3.14).

O .xlsx real docs/RGE GAU 2026 - Lista de Pontos v09.xlsx traz PatternFill com
extLst e derrubava a leitura (TypeError) antes do shim. Importar `tdt` aplica o
patch; carregar o arquivo tem de funcionar.
"""
import openpyxl
import pytest

import tdt  # noqa: F401 — importar aplica o shim de compat

_ARQUIVO = "docs/RGE GAU 2026 - Lista de Pontos v09.xlsx"


@pytest.mark.skipif(
    not __import__("pathlib").Path(_ARQUIVO).exists(),
    reason="arquivo real da concessionaria ausente no ambiente",
)
def test_le_xlsx_com_patternfill_extlst_sem_crash():
    wb = openpyxl.load_workbook(_ARQUIVO, read_only=True, data_only=True)
    assert wb.sheetnames  # abriu e listou as abas, sem TypeError
    wb.close()


def test_shim_e_idempotente():
    import openpyxl.styles.fills as fills
    assert getattr(fills.PatternFill, "_extlst_shim", False) is True
    # PatternFill normal continua construindo
    pf = fills.PatternFill(patternType="solid", fgColor="FF0000")
    assert pf.patternType == "solid"
    # e agora aceita/descarta extLst sem quebrar
    pf2 = fills.PatternFill(patternType="solid", extLst="qualquer")
    assert pf2.patternType == "solid"
