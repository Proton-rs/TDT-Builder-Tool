from dataclasses import replace as _replace

import numpy as np
import openpyxl
import pytest

from tdt.auditoria import Auditoria
from tdt.config import Config
from tdt.contracts import (
    Candidato, Descricoes, Enderecamento, Modulo, SignalRecord, TipoSinal,
)
from tdt.dados.lista_padrao import ListaPadraoADMS
from tdt.pipeline import executar, _com_fase, _construir_scorers, _classificar_roteado
from tdt.pipeline import _desempatar_ambiguo, _gap, _timer


# bag-of-words sobre vocab de sinais: faz a coluna de descrição casar a lista ADMS
_VOCAB = ["FALHA", "COMUNICACAO", "DISJUNTOR", "ABERTO", "FECHADO", "CORRENTE",
          "TENSAO", "FASE", "POTENCIA", "SECCIONADORA", "RELE", "PROTECAO",
          "NEUTRO", "ALARME", "BAIXA", "PRESSAO", "SF6"]


def _fake_encoder(textos):
    return np.array(
        [[float(str(t).upper().split().count(w)) for w in _VOCAB] for t in textos],
        dtype="float32",
    )


def _input_sintetico(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # nome de sheet resolve via mapa_prefixo_modulo (GTD->AL) p/ confiança "alta"
    # em identidade_modulo — não testamos identidade de módulo aqui.
    ws.title = "GTD_11"
    ws.append(["", "", "SUBESTAÇÃO X", "", ""])
    ws.append(["IED", "Módulo", "Descrição do Ponto", "Tipo", "Endereço DNP3"])
    ws.append(["Digitais", "", "", "", ""])
    ws.append(["01F1", "LT_GTA", "FALHA COMUNICACAO", "Digital", "10"])
    ws.append(["01F1", "LT_GTA", "DISJUNTOR", "Digital", "11"])
    p = tmp_path / "input.xlsx"
    wb.save(p)
    return p


def test_pipeline_ponta_a_ponta_gera_tdt(tmp_path, template_dnp3_path, lista_padrao_path):
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_sintetico(tmp_path)
    resultado, wb = executar(
        inp,
        template_dnp3_path,
        lista_padrao_path,
        config=cfg,
        encoder=_fake_encoder,
        subestacao="X",
        modo="nao-homogeneo",
    )
    # gerou lista homogênea com sinais decididos
    assert len(resultado.lista.registros) >= 1
    # o TDT tem as 43 colunas e dados a partir da row 5
    ws = wb["DNP3_DiscreteSignals"]
    assert ws.max_column == 43
    nomes = [ws.cell(r, 1).value for r in range(5, 5 + len(resultado.lista.registros))]
    assert any(n and n.startswith("X_") for n in nomes)


def test_pipeline_emite_evento_de_progresso_com_atual_e_total(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    """Task 1.4: evento de progresso traz dados={'atual','total'} p/ a UI
    desenhar a barra, sem precisar de um nível dedicado em Auditoria."""
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_sintetico(tmp_path)
    aud = Auditoria()
    executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X",
        modo="nao-homogeneo", auditoria=aud,
    )
    eventos_progresso = [e for e in aud.eventos if e.dados and "atual" in e.dados and "total" in e.dados]
    assert eventos_progresso, "esperava ao menos um evento com dados de progresso"
    ultimo = eventos_progresso[-1]
    assert ultimo.dados["atual"] == ultimo.dados["total"]  # sempre emite no último sinal da sheet


def test_timer_registra_evento_perf_com_nome_e_duracao():
    """Task 1.5: _timer mede uma etapa e registra um evento "perf" com o
    nome esperado na mensagem."""
    aud = Auditoria()
    with _timer("etapa teste", aud):
        pass
    eventos_perf = [e for e in aud.eventos if e.modulo == "perf"]
    assert len(eventos_perf) == 1
    assert eventos_perf[0].msg.startswith("etapa teste: ")
    assert eventos_perf[0].msg.endswith("s")


def test_timer_propaga_excecao_sem_mascarar_e_ainda_registra_tempo():
    """_timer não deve suprimir exceções da etapa cronometrada — apenas
    registra o tempo até a falha e deixa a exceção propagar."""
    aud = Auditoria()
    with pytest.raises(ValueError, match="boom"):
        with _timer("etapa que falha", aud):
            raise ValueError("boom")
    eventos_perf = [e for e in aud.eventos if e.modulo == "perf"]
    assert len(eventos_perf) == 1
    assert eventos_perf[0].msg.startswith("etapa que falha: ")


def test_pipeline_emite_eventos_perf_para_scorers_e_etapa_final(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    """Task 1.5: executar() deve registrar eventos "perf" para a construção
    dos scorers (disc/ana) e para a etapa final de pareamento+tdt."""
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_sintetico(tmp_path)
    aud = Auditoria()
    executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X",
        modo="nao-homogeneo", auditoria=aud,
    )
    msgs_perf = [e.msg for e in aud.eventos if e.modulo == "perf"]
    assert any(m.startswith("construir scorers disc:") for m in msgs_perf)
    assert any(m.startswith("construir scorers ana:") for m in msgs_perf)
    assert any(m.startswith("dc_pairer + corrigir + montar + tdt:") for m in msgs_perf)


def test_pipeline_com_cache_scorers_reusa_entre_execucoes(tmp_path, template_dnp3_path, lista_padrao_path):
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_sintetico(tmp_path)
    cache_dir = tmp_path / "cache_scorers"

    resultado1, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
        cache_scorers_dir=cache_dir,
    )
    assert cache_dir.exists()
    assert any(cache_dir.iterdir())  # populou ao menos um diretório de hash

    # segunda execução: cache já existe, deve reusar e produzir o mesmo resultado
    resultado2, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
        cache_scorers_dir=cache_dir,
    )
    siglas1 = [r.sigla_sinal for r in resultado1.lista.registros]
    siglas2 = [r.sigla_sinal for r in resultado2.lista.registros]
    assert siglas1 == siglas2


def test_pipeline_classifica_falha_comunicacao(tmp_path, template_dnp3_path, lista_padrao_path):
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_sintetico(tmp_path)
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
    )
    siglas = {r.sigla_sinal for r in resultado.lista.registros}
    assert "FCOM" in siglas  # "FALHA COMUNICACAO" -> FCOM


def _input_multi_sheet(tmp_path):
    """2 sheets de dados idênticas em estrutura (GTD_11, GTD_22) — usado para
    testar filtro de seleção e rename de sheet/módulo end-to-end."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GTD_11"
    ws.append(["", "", "SUBESTAÇÃO X", "", ""])
    ws.append(["IED", "Módulo", "Descrição do Ponto", "Tipo", "Endereço DNP3"])
    ws.append(["Digitais", "", "", "", ""])
    ws.append(["01F1", "LT_GTA", "FALHA COMUNICACAO", "Digital", "10"])
    ws.append(["01F1", "LT_GTA", "DISJUNTOR", "Digital", "11"])
    ws2 = wb.create_sheet("GTD_22")
    ws2.append(["", "", "SUBESTAÇÃO X", "", ""])
    ws2.append(["IED", "Módulo", "Descrição do Ponto", "Tipo", "Endereço DNP3"])
    ws2.append(["Digitais", "", "", "", ""])
    ws2.append(["01F1", "LT_GTB", "FALHA COMUNICACAO", "Digital", "10"])
    ws2.append(["01F1", "LT_GTB", "DISJUNTOR", "Digital", "11"])
    p = tmp_path / "input_multi_sheet.xlsx"
    wb.save(p)
    return p


def test_pipeline_processa_apenas_sheets_selecionadas(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    """Task 2 (spK): `sheets` restringe quais sheets de `rota.sheets_dados`
    são processadas — sheet fora da lista não aparece no resultado."""
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_multi_sheet(tmp_path)
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
        sheets=["GTD_11"],
    )
    origens = {r.id.split(":")[0] for r in resultado.lista.registros}
    assert origens == {"GTD_11"}
    assert "GTD_22" not in origens


def test_pipeline_aplica_aliases_ao_nome_do_modulo(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    """Task 2 (spK): `aliases` (sheet original -> apelido, igual
    `estado.aliases` da tela inicial) renomeia `modulo.nome` dos registros
    vindos dessa sheet. Chave é o NOME DA SHEET (não o nome de módulo
    resolvido) — mesma chave que a UI usa em `_sheet_alterada`."""
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_multi_sheet(tmp_path)
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
        aliases={"GTD_11": "MODULO_RENOMEADO"},
    )
    nomes_modulo = {r.modulo.nome for r in resultado.lista.registros}
    assert "MODULO_RENOMEADO" in nomes_modulo
    assert "AL11" not in nomes_modulo  # renomeado, não duplicado
    assert "AL22" in nomes_modulo  # sheet não referenciada no alias, intacta


def _input_sem_endereco(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # nome de sheet resolve via mapa_prefixo_modulo (GTD->AL) p/ confiança "alta"
    # em identidade_modulo — não testamos identidade de módulo aqui.
    ws.title = "GTD_11"
    ws.append(["", "", "SUBESTAÇÃO X", "", ""])
    ws.append(["IED", "Módulo", "Descrição do Ponto", "Tipo", "Endereço DNP3"])
    ws.append(["Digitais", "", "", "", ""])
    ws.append(["01F1", "LT_GTA", "FALHA COMUNICACAO", "Digital", ""])  # sem endereco (futuro)
    ws.append(["01F1", "LT_GTA", "DISJUNTOR ABERTO", "Digital", "11"])
    p = tmp_path / "input_sem_endereco.xlsx"
    wb.save(p)
    return p


def test_sinal_sem_endereco_recebe_candidato_sugerido(tmp_path, template_dnp3_path, lista_padrao_path):
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_sem_endereco(tmp_path)
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
    )
    item = next(it for it in resultado.revisao if it.motivo == "sem_endereco")
    assert len(item.candidatos_sugeridos) > 0
    assert item.registro not in resultado.lista.registros  # nunca auto-aprovado


def _input_homogeneo_sintetico(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ALX"
    ws.append(["EQUIPAMENTO", "NÚMERO OPERATIVO / MNEMÔNICO"])
    ws.append(["DJ", "52-11"])
    ws.append([])
    ws.append(["Utilizado?", "SUBESTAÇÃO", "MÓDULO", "EQUIPAMENTO", "TIPO",
               "DESCRIÇÃO DO PONTO", "SIGLA SINAL", "NOME", "Tipo",
               "Nível Lógico 0", "Nível Lógico 1", "Escala",
               "Control Code / Qualificador", "INDEX DNP3"])
    ws.append(["SIM", "IMA", "ALX", "", "D", "FALHA COMUNICACAO", "FCOM",
               "IMA_ALX_FCOM", "-", "-", "-", "-", "-", "10"])
    ws.append(["SIM", "IMA", "ALX", "", "D", "DISJUNTOR ABERTO", "DJX",
               "IMA_ALX_DJX", "-", "-", "-", "-", "-", "11"])
    p = tmp_path / "input_homog.xlsx"
    wb.save(p)
    return p


def test_pipeline_sheet_homogenea_le_colunas_direto_sem_scoring(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_homogeneo_sintetico(tmp_path)
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="homogeneo",
    )
    # FCOM: sigla existente na lista padrao -> decidido direto pelas colunas,
    # sem candidatos nem justificativa (nao passou pelos scorers).
    fcom = next(r for r in resultado.lista.registros if r.sigla_sinal == "FCOM")
    assert fcom.modulo.nome == "ALX"  # da coluna MODULO, nao do nome da sheet
    assert fcom.enderecamento.indices == (10,)
    assert fcom.candidatos == ()
    assert fcom.justificativa is None
    # DJX: sigla nao existe na lista padrao -> cai no caminho de scoring normal
    # (continua decidindo algo, mas via candidatos/justificativa do scorer).
    djx = next(r for r in resultado.lista.registros if r.descricoes.bruta == "DISJUNTOR ABERTO")
    assert djx.candidatos != ()
    assert djx.justificativa is not None


def _input_djf1_sintetico(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # nome de sheet resolve via mapa_prefixo_modulo (GTD->AL) p/ confiança "alta"
    # em identidade_modulo — não testamos identidade de módulo aqui.
    ws.title = "GTD_12"
    ws.append(["", "", "SUBESTAÇÃO X", "", ""])
    ws.append(["IED", "Módulo", "Descrição do Ponto", "Tipo", "Endereço DNP3"])
    ws.append(["Digitais", "", "", "", ""])
    ws.append(["06F2", "06F2", "DISJUNTOR 52-2 DESLIGADO", "Digital", "100"])
    ws.append(["06F2", "06F2", "DISJUNTOR 52-2 LIGADO", "Digital", "101"])
    p = tmp_path / "input_djf1.xlsx"
    wb.save(p)
    return p


def test_djf1_par_ligado_desligado_converge_para_multicoord(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    # D3: par de POSIÇÃO (DJF1 é SwitchStatus na lista padrão) com estados
    # opostos e endereços consecutivos vira MultiCoord — não DoubleBit
    # (reservado ao par N;M nativo do input).
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_djf1_sintetico(tmp_path)
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
    )
    djf1 = [r for r in resultado.lista.registros if r.sigla_sinal == "DJF1"]
    assert len(djf1) == 1
    assert djf1[0].tipo_sinal.datatype == "MultiCoord"
    assert sorted(djf1[0].enderecamento.indices) == [100, 101]


def test_djf1_pareamento_desligado_via_flag_cai_no_scoring_normal(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05,
        parear_polaridade_equipamento=False,
    )
    inp = _input_djf1_sintetico(tmp_path)
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
    )
    djf1 = [r for r in resultado.lista.registros if r.sigla_sinal == "DJF1" and r.tipo_sinal.datatype == "DoubleBit"]
    assert djf1 == []  # sem o pareamento forcado, "DISJUNTOR...LIGADO/DESLIGADO" nao bate em "DISJUNTOR NF"


def _rec_min(sigla):
    return SignalRecord(
        id="s:1",
        modulo=Modulo("3", "sheet_name"),
        tipo_sinal=TipoSinal("Discrete", "SingleBit", "Input"),
        enderecamento=Enderecamento("DNP3", (1,)),
        descricoes=Descricoes(sigla, sigla),
        sigla_sinal=sigla,
        status="decidido",
    )


def test_com_fase_deriva_de_sigla_neutro():
    out = _com_fase(_rec_min("51N"))
    assert out.eletrico.fase == "N"


def test_com_fase_sem_fase_mantem_none():
    out = _com_fase(_rec_min("FCOM"))
    assert out.eletrico.fase is None


def test_com_fase_nao_sobrescreve_existente():
    base = _rec_min("51N")
    base = _replace(base, eletrico=_replace(base.eletrico, fase="ABC"))
    out = _com_fase(base)
    assert out.eletrico.fase == "ABC"


def test_com_fase_mapeia_F_generico_para_ABC():
    # PRTF/50F1 -> fase_da_sigla devolve "F" (genérica trifásica) -> ABC
    assert _com_fase(_rec_min("PRTF")).eletrico.fase == "ABC"
    assert _com_fase(_rec_min("50F1")).eletrico.fase == "ABC"


def test_com_fase_preserva_fase_especifica():
    assert _com_fase(_rec_min("51N")).eletrico.fase == "N"
    assert _com_fase(_rec_min("51A")).eletrico.fase == "A"


def _input_com_analogico(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    # nome de sheet resolve via mapa_prefixo_modulo (GTD->AL) p/ confiança "alta"
    # em identidade_modulo — não testamos identidade de módulo aqui.
    ws.title = "GTD_11"
    ws.append(["", "", "SUBESTAÇÃO X", "", ""])
    ws.append(["IED", "Módulo", "Descrição do Ponto", "Tipo", "Endereço DNP3"])
    ws.append(["Digitais", "", "", "", ""])
    ws.append(["01F1", "LT_GTA", "FALHA COMUNICACAO", "Digital", "10"])
    ws.append(["Analógicas", "", "", "", ""])
    ws.append(["01F1", "LT_GTA", "CORRENTE FASE A", "Analógico", "20"])
    p = tmp_path / "input_ana.xlsx"
    wb.save(p)
    return p


def test_pipeline_classifica_analogico(tmp_path, template_dnp3_path, lista_padrao_path):
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0,
                 peso_tfidf_analog=1.0, peso_vetorial_analog=0.0,
                 threshold_pct=0.3, threshold_gap=0.01,
                 threshold_pct_analog=0.3, threshold_gap_analog=0.01)
    inp = _input_com_analogico(tmp_path)
    resultado, wb = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
    )
    analog = [r for r in resultado.lista.registros if r.tipo_sinal.categoria == "Analog"]
    assert len(analog) >= 1
    # saiu na sheet analógica do TDT
    ws = wb["DNP3_AnalogSignals"]
    nomes = [ws.cell(r, 1).value for r in range(5, 5 + len(analog))]
    assert any(n for n in nomes)


def _rec_incerto(descricao="CORRENTE FASE A"):
    """SignalRecord com categoria_confiavel=False — aciona o dual-pass."""
    return SignalRecord(
        id="s:1",
        modulo=Modulo("3", "sheet_name"),
        tipo_sinal=TipoSinal("Discrete", "SingleBit", "Input", categoria_confiavel=False),
        enderecamento=Enderecamento("DNP3", (1,)),
        descricoes=Descricoes(descricao, descricao),
    )


def _bundles(lista_padrao_path, cfg: Config):
    """Constrói os bundles disc/ana a partir da config, espelhando a montagem de `executar`."""
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    cfg_analog = _replace(
        cfg,
        peso_tfidf=cfg.peso_tfidf_analog,
        peso_vetorial=cfg.peso_vetorial_analog,
        peso_fuzzy=cfg.peso_fuzzy_analog,
        threshold_pct=cfg.threshold_pct_analog,
        threshold_gap=cfg.threshold_gap_analog,
    )
    disc = _construir_scorers(lp, cfg, _fake_encoder, "Discrete", cfg)
    ana = _construir_scorers(lp, cfg, _fake_encoder, "Analog", cfg_analog)
    return disc, ana


def test_classificar_roteado_categoria_incerta_apenas_um_decide(lista_padrao_path):
    # discreto com threshold frouxo decide; analógico com threshold travado não decide.
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.01, threshold_gap=0.0,
        peso_tfidf_analog=1.0, peso_vetorial_analog=0.0,
        threshold_pct_analog=5.0, threshold_gap_analog=5.0,
    )
    disc, ana = _bundles(lista_padrao_path, cfg)
    rec = _rec_incerto()
    decidido, item = _classificar_roteado(rec, disc, ana, diagnostico=False)
    assert item is None
    assert decidido is not None
    assert decidido.status == "decidido"
    assert decidido.sigla_sinal == "FA"


def test_classificar_roteado_categoria_incerta_ambos_decidem_categoria_ambigua(lista_padrao_path):
    # thresholds frouxos nos dois bundles -> ambos decidem. A fixture tem
    # categoria="Discrete" (placeholder do estruturador) -> a barreira de
    # domínio bloqueia o lado Analog. Como os DOIS decidiram (conflito real,
    # não só score baixo), não se auto-aceita o lado "permitido" (Discrete):
    # vai para revisão como categoria_ambigua, sem rodar o desempate por
    # gap/centroide (esse só roda quando os dois domínios decididos são
    # admitidos pela categoria — ver test_dual_pass_discreteanalog_*).
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.01, threshold_gap=0.0,
        peso_tfidf_analog=1.0, peso_vetorial_analog=0.0,
        threshold_pct_analog=0.01, threshold_gap_analog=0.0,
    )
    disc, ana = _bundles(lista_padrao_path, cfg)
    rec = _rec_incerto()
    decidido, item = _classificar_roteado(rec, disc, ana, diagnostico=False)
    assert decidido is None
    assert item is not None
    assert item.motivo == "categoria_ambigua"
    assert len(item.candidatos_sugeridos) > 0


def test_dual_pass_barra_dominio_so_bundle_errado_decide(lista_padrao_path):
    # categoria="Analog" (evidência real, ex: veio de coluna Tipo, módulo
    # indefinido forçou confiavel=False). Disc decide (threshold frouxo);
    # ana não decide (threshold travado). A barreira bloqueia o disc (fora
    # do domínio Analog) e, como ana nem tentou (não houve conflito), não
    # auto-aceita nem manda pro desempate -> revisão categoria_incompativel.
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.01, threshold_gap=0.0,
        peso_tfidf_analog=1.0, peso_vetorial_analog=0.0,
        threshold_pct_analog=5.0, threshold_gap_analog=5.0,
    )
    disc, ana = _bundles(lista_padrao_path, cfg)
    rec = _replace(
        _rec_incerto(),
        tipo_sinal=TipoSinal("Analog", "SingleBit", "Input", categoria_confiavel=False),
    )
    decidido, item = _classificar_roteado(rec, disc, ana, diagnostico=False)
    assert decidido is None
    assert item is not None
    assert item.motivo == "categoria_incompativel"
    assert len(item.candidatos_sugeridos) > 0


def test_dual_pass_discreteanalog_ambos_decidem_aceita_desempate(lista_padrao_path):
    # categoria="DiscreteAnalog" admite os dois domínios -> a barreira não
    # restringe nada; ambos decidem e o desempate normal (gap/centroide)
    # roda como no dual-pass livre de antes.
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.01, threshold_gap=0.0,
        peso_tfidf_analog=1.0, peso_vetorial_analog=0.0,
        threshold_pct_analog=0.01, threshold_gap_analog=0.0,
    )
    disc, ana = _bundles(lista_padrao_path, cfg)
    rec = _replace(
        _rec_incerto(),
        tipo_sinal=TipoSinal("DiscreteAnalog", "SingleBit", "Input", categoria_confiavel=False),
    )
    decidido, item = _classificar_roteado(rec, disc, ana, diagnostico=False)
    assert decidido is not None
    assert item is None
    # tipo_sinal.categoria não é alterado por _classificar_sinal — segue
    # "DiscreteAnalog" (o que importa é que decidiu, via desempate normal).
    assert decidido.status == "decidido"
    assert decidido.sigla_sinal


def test_dual_pass_discreteanalog_so_um_decide_aceita(lista_padrao_path):
    # categoria="DiscreteAnalog"; só disc decide (ana travado). Sem conflito
    # (ana não decidiu) -> aceita direto, sem ir pra revisão.
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.01, threshold_gap=0.0,
        peso_tfidf_analog=1.0, peso_vetorial_analog=0.0,
        threshold_pct_analog=5.0, threshold_gap_analog=5.0,
    )
    disc, ana = _bundles(lista_padrao_path, cfg)
    rec = _replace(
        _rec_incerto(),
        tipo_sinal=TipoSinal("DiscreteAnalog", "SingleBit", "Input", categoria_confiavel=False),
    )
    decidido, item = _classificar_roteado(rec, disc, ana, diagnostico=False)
    assert item is None
    assert decidido is not None
    assert decidido.status == "decidido"


def _fake_enc_zeros(textos):
    import numpy as np
    return np.zeros((len(textos), 5), dtype="float32")


def _rec_ancorado(texto: str) -> SignalRecord:
    return SignalRecord(
        id="t:1",
        modulo=Modulo("M", "sheet"),
        tipo_sinal=TipoSinal("Discrete", "SingleBit", "Input", categoria_confiavel=True),
        enderecamento=Enderecamento("DNP3", (1,)),
        descricoes=Descricoes(texto, texto),
    )


def test_classificar_roteado_ancora_exata_decide_variante_pai_c1(lista_padrao_path):
    """Caso real H1 (diag SP-OBS-17JUL P5-Fase0): âncora exata "51N" no texto,
    top-3 são só irmãos da família 51N (51N, 51N1, 51N2) e a própria sigla
    âncora sobrevive no top -- C1 decide "51N" em vez de mandar p/ revisão."""
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, peso_fuzzy=0.0,
        threshold_pct=0.01, threshold_gap=5.0,
        ancora_sigla_ativa=True, ancora_sigla_score=0.85,
    )
    disc = _construir_scorers(lp, cfg, _fake_enc_zeros, "Discrete", cfg)
    cfg_analog = _replace(cfg, threshold_pct=5.0, threshold_gap=5.0)
    ana = _construir_scorers(lp, cfg, _fake_enc_zeros, "Analog", cfg_analog)

    rec = _rec_ancorado("PROTECAO 51N SOBRECORRENTE TEMPORIZADA NEUTRO")
    decidido, item = _classificar_roteado(rec, disc, ana, diagnostico=False, lista_padrao=lp)

    assert item is None
    assert decidido is not None
    assert decidido.status == "decidido"
    assert decidido.sigla_sinal == "51N"
    assert decidido.justificativa == "variante-pai exata da âncora (C1)"


def test_classificar_roteado_ancora_exata_sem_variante_pai_no_top3_vira_variante_ambigua(lista_padrao_path):
    """Mesma família 51N ancorada, mas a própria sigla "51N" não sobrevive no
    top-3 (H3) -- C1 não tem o que decidir; motivo vira "variante_ambigua"."""
    lp = ListaPadraoADMS.carregar(lista_padrao_path)
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, peso_fuzzy=0.0,
        threshold_pct=0.01, threshold_gap=5.0,
        ancora_sigla_ativa=True, ancora_sigla_score=0.85,
    )
    disc = _construir_scorers(lp, cfg, _fake_enc_zeros, "Discrete", cfg)
    cfg_analog = _replace(cfg, threshold_pct=5.0, threshold_gap=5.0)
    ana = _construir_scorers(lp, cfg, _fake_enc_zeros, "Analog", cfg_analog)

    rec = _rec_ancorado("SOBRECORRENTE TEMPORIZADA NEUTRO 51N ATUADO")
    decidido, item = _classificar_roteado(rec, disc, ana, diagnostico=False, lista_padrao=lp)

    assert decidido is None
    assert item is not None
    assert item.motivo == "variante_ambigua"
    siglas_sugeridas = {c.sigla for c in item.candidatos_sugeridos}
    assert "51N" not in siglas_sugeridas


def test_classificar_roteado_categoria_incerta_nenhum_decide_score_baixo(lista_padrao_path):
    # thresholds travados nos dois bundles -> nenhum decide -> revisão por score baixo.
    cfg = Config(
        peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=5.0, threshold_gap=5.0,
        peso_tfidf_analog=1.0, peso_vetorial_analog=0.0,
        threshold_pct_analog=5.0, threshold_gap_analog=5.0,
    )
    disc, ana = _bundles(lista_padrao_path, cfg)
    rec = _rec_incerto()
    decidido, item = _classificar_roteado(rec, disc, ana, diagnostico=False)
    assert decidido is None
    assert item is not None
    assert item.motivo == "score_baixo"
    assert len(item.candidatos_sugeridos) > 0


class _IndiceFake:
    def __init__(self, afinidade):
        self._afinidade = afinidade

    def afinidade_centroide(self, _texto):
        return self._afinidade


class _ScorersFake:
    def __init__(self, afinidade):
        self.indice = _IndiceFake(afinidade)


def _rec_com_candidatos(*scores):
    cands = tuple(Candidato(f"S{i}", s, "tfidf") for i, s in enumerate(scores))
    rec = _rec_incerto()
    return _replace(rec, candidatos=cands, status="decidido", sigla_sinal=cands[0].sigla)


def test_gap_decide_quando_diferenca_e_grande():
    d_disc = _rec_com_candidatos(0.90, 0.40)  # gap 0.50
    d_ana = _rec_com_candidatos(0.55, 0.50)   # gap 0.05
    vencedor = _desempatar_ambiguo(d_disc, d_ana, _ScorersFake(0.0), _ScorersFake(0.0), "x")
    assert vencedor is d_disc


def test_centroide_decide_quando_gap_empata():
    d_disc = _rec_com_candidatos(0.60, 0.55)  # gap 0.05
    d_ana = _rec_com_candidatos(0.62, 0.58)   # gap 0.04 — dentro da margem (0.03)
    vencedor = _desempatar_ambiguo(
        d_disc, d_ana, _ScorersFake(afinidade=0.2), _ScorersFake(afinidade=0.8), "x"
    )
    assert vencedor is d_ana


def test_permanece_ambiguo_quando_gap_e_centroide_empatam():
    d_disc = _rec_com_candidatos(0.60, 0.55)
    d_ana = _rec_com_candidatos(0.62, 0.58)
    vencedor = _desempatar_ambiguo(
        d_disc, d_ana, _ScorersFake(afinidade=0.5), _ScorersFake(afinidade=0.5), "x"
    )
    assert vencedor is None


def test_gap_de_candidato_unico_e_o_proprio_score():
    rec = _rec_com_candidatos(0.90)
    assert _gap(rec) == 0.90


# --- SP C2: inferência de equipamento por topologia visível ao scoring ------


def test_equipamento_inferido_fica_visivel_no_registro_decidido(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    """C2.3: equipamento_alvo inferido pela topologia (C1 resolve o módulo
    como Alimentador via GTD->AL) chega ao registro final decidido -- ou
    seja, o motor de regras (r_equipamento/r3_fase) já viu o equipamento
    inferido durante o scoring, não só depois."""
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.5, threshold_gap=0.05)
    inp = _input_sintetico(tmp_path)  # "FALHA COMUNICACAO" sem equipamento explícito
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="X", modo="nao-homogeneo",
    )
    fcom = next(r for r in resultado.lista.registros if r.sigla_sinal == "FCOM")
    assert fcom.modulo.tipo == "Alimentador"
    assert fcom.eletrico.equipamento_alvo == "Disjuntor"  # default da topologia
    assert fcom.eletrico.equipamento_inferido is True


# --- SP-C: equipamento_ambiguo não bloqueia mais a emissão -------------------


def _input_equip_ambiguo(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # "IB_1" classifica como módulo tipo Barra (sem equipamento default) --
    # sinais sem comando (sem seção "Comandos", direção fica Input/status).
    # 2 linhas: TF-IDF degenera com corpus de 1 doc só (IDF~0 -> mesmo score
    # pra qualquer candidato, igual o teste_equipamento_inferido evita).
    ws.title = "IB_1"
    ws.append(["", "", "SUBESTAÇÃO X", "", ""])
    ws.append(["IED", "Módulo", "Descrição do Ponto", "Tipo", "Endereço DNP3"])
    ws.append(["Digitais", "", "", "", ""])
    ws.append(["01F1", "IB_1", "ALARME BAIXA PRESSAO SF6", "Digital", "10"])
    ws.append(["01F1", "IB_1", "FALHA COMUNICACAO RELE PROTECAO", "Digital", "11"])
    p = tmp_path / "input.xlsx"
    wb.save(p)
    return p


def test_equipamento_ambiguo_sem_comando_vai_pra_tdt(
    tmp_path, template_dnp3_path, lista_padrao_path,
):
    """C1: sinal com sigla decidida, módulo de tipo sem equipamento default
    (família não inferida) e SEM comando -> entra na TDT (lista.registros),
    não vira revisão equipamento_ambiguo."""
    cfg = Config(peso_tfidf=1.0, peso_vetorial=0.0, threshold_pct=0.4, threshold_gap=0.05)
    inp = _input_equip_ambiguo(tmp_path)
    resultado, _ = executar(
        inp, template_dnp3_path, lista_padrao_path,
        config=cfg, encoder=_fake_encoder, subestacao="GTD", modo="nao-homogeneo",
    )
    motivos = {it.motivo for it in resultado.revisao}
    assert "equipamento_ambiguo" not in motivos
    assert any(r.sigla_sinal for r in resultado.lista.registros)


# --- Confiança normalizada [0,1] no boundary do pipeline ---------------------


def test_limitar_confianca_clampa_score_acima_de_1():
    from tdt.pipeline import _limitar_confianca

    rec = _rec_com_candidatos(1.25, 0.95, -0.10)
    norm = _limitar_confianca(rec)
    assert [c.score for c in norm.candidatos] == [1.0, 0.95, 0.0]
    # ordem e siglas preservadas
    assert [c.sigla for c in norm.candidatos] == [c.sigla for c in rec.candidatos]


def test_limitar_confianca_sem_estouro_devolve_o_mesmo_registro():
    from tdt.pipeline import _limitar_confianca

    rec = _rec_com_candidatos(0.90, 0.40)
    assert _limitar_confianca(rec) is rec


def test_limitar_confianca_item_revisao_clampa_registro_e_sugeridos():
    from tdt.contracts import ItemRevisao
    from tdt.pipeline import _limitar_confianca_item

    rec = _rec_com_candidatos(1.30, 0.80)
    item = ItemRevisao(rec, motivo="score_baixo", candidatos_sugeridos=rec.candidatos)
    norm = _limitar_confianca_item(item)
    assert norm.registro.candidatos[0].score == 1.0
    assert norm.candidatos_sugeridos[0].score == 1.0
    assert norm.motivo == "score_baixo"
