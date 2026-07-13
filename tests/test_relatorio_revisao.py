import openpyxl

from tdt.contracts import (
    Candidato, Descricoes, Diagnostico, Eletrico, Enderecamento, ItemRevisao,
    Modulo, SignalRecord, TipoSinal,
)
from tdt.relatorio_revisao import (
    CABECALHO, descricao_para_exibicao, gerar_relatorio_revisao,
)


def _rec(id_, sigla=None, candidatos=(), diagnostico=None, status="decidido", eletrico=None):
    return SignalRecord(
        id=id_, modulo=Modulo("M", "sheet_name"),
        tipo_sinal=TipoSinal("Discrete", "SingleBit", "Input"),
        enderecamento=Enderecamento("DNP3", (10,)),
        descricoes=Descricoes("Disjuntor Aberto", "disjuntor aberto"),
        sigla_sinal=sigla, candidatos=candidatos, status=status,
        diagnostico=diagnostico, eletrico=eletrico or Eletrico(),
    )


def _col(ws, nome_coluna: str, linha: int = 2):
    idx = CABECALHO.index(nome_coluna) + 1
    return ws.cell(linha, idx).value


def test_colunas_estendidas_presentes_no_cabecalho():
    esperado = [
        "Sheet Origem", "Desc Normalizada", "Desc Canônica", "Equip Alvo (N0)",
        "Nome Equip", "Barra", "Fase", "Estado Semântico", "Regras Aplicadas",
        "Gap", "Gap Exigido", "Etapa Decisora", "Endereço Bruto",
    ]
    for col in esperado:
        assert col in CABECALHO, f"coluna ausente: {col}"


def test_registro_decidido_popula_colunas_de_contexto_e_decisao(tmp_path):
    cands = (Candidato("DJF1", 0.91, "mesclado"), Candidato("SECC", 0.40, "mesclado"))
    rec = _rec(
        "SUB1:5", sigla="DJF1", candidatos=cands,
        eletrico=Eletrico(fase="A", equipamento_alvo="Disjuntor",
                          nome_equipamento="52-10", barra="Principal"),
    )
    diagnostico = {
        "SUB1:5": {
            "regras_aplicadas": "r3_fase: +0.10 (fase compatível)",
            "gap": 0.51,
            "gap_exigido": 0.08,
            "etapa_decisora": "quadrante",
            "endereco_bruto": "10",
        }
    }

    caminho = gerar_relatorio_revisao([rec], (), tmp_path, diagnostico=diagnostico)

    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    assert _col(ws, "Sheet Origem") == "SUB1"
    assert _col(ws, "Desc Canônica") == "disjuntor aberto"
    assert _col(ws, "Equip Alvo (N0)") == "Disjuntor"
    assert _col(ws, "Nome Equip") == "52-10"
    assert _col(ws, "Barra") == "Principal"
    assert _col(ws, "Fase") == "A"
    assert _col(ws, "Regras Aplicadas") == "r3_fase: +0.10 (fase compatível)"
    assert _col(ws, "Gap") == "0.510"
    assert _col(ws, "Gap Exigido") == "0.080"
    assert _col(ws, "Etapa Decisora") == "quadrante"
    assert _col(ws, "Endereço Bruto") == "10"


def test_registro_em_revisao_sem_diagnostico_nao_quebra_e_usa_placeholder(tmp_path):
    rec = _rec("SUB2:9", sigla=None, candidatos=(), status="revisao")
    revisao = (ItemRevisao(rec, motivo="score_baixo"),)

    caminho = gerar_relatorio_revisao([rec], revisao, tmp_path, diagnostico={})

    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    assert _col(ws, "Sheet Origem") == "SUB2"
    # sem diagnostico para este id -> colunas de decisão vazias, não fabricadas
    assert _col(ws, "Etapa Decisora") in (None, "")
    assert _col(ws, "Gap") in (None, "")
    assert _col(ws, "Regras Aplicadas") in (None, "")


def test_relatorio_duplicado_lista_sheets_de_origem_de_cada_registro(tmp_path):
    """itens de revisão por Custom ID duplicado -- cada linha deve exibir sua
    própria sheet de origem, para diferenciar os registros do grupo."""
    recs = [_rec("BC1:10"), _rec("BC2:12")]
    revisao = tuple(ItemRevisao(r, motivo="custom_id_duplicado") for r in recs)

    caminho = gerar_relatorio_revisao(recs, revisao, tmp_path)

    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    assert _col(ws, "Sheet Origem", linha=2) == "BC1"
    assert _col(ws, "Sheet Origem", linha=3) == "BC2"


def test_gerar_relatorio_sem_diagnostico_mantem_retrocompat(tmp_path):
    """Chamador antigo (sem o parâmetro novo) continua funcionando -- UI legada."""
    rec = _rec("SUB3:1", sigla="DJF1", candidatos=(Candidato("DJF1", 0.9, "mesclado"),))

    caminho = gerar_relatorio_revisao([rec], (), tmp_path)

    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    assert ws.cell(2, 1).value == "SUB3:1"


def test_gera_planilha_com_uma_linha_por_sinal(tmp_path):
    cands = (Candidato("DJ", 0.91, "mesclado"), Candidato("SC", 0.40, "mesclado"))
    diag = Diagnostico(scores_por_metodo={"DJ": {"tfidf": 0.9, "vetorial": 0.92, "fuzzy": 0.88}})
    registros = [_rec("S1:1", sigla="DJ", candidatos=cands, diagnostico=diag)]
    revisao = ()

    caminho = gerar_relatorio_revisao(registros, revisao, tmp_path)

    assert caminho.exists()
    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    assert ws.cell(1, 1).value == "ID Sinal"
    assert ws.cell(2, 1).value == "S1:1"
    assert ws.cell(2, 6).value == "DJ"  # Sigla Decidida
    assert ws.cell(2, 9).value == "DJ"  # Candidato 1


def test_sinal_sem_candidatos_nao_quebra(tmp_path):
    registros = [_rec("S1:2", sigla=None, candidatos=(), status="revisao")]
    revisao = (ItemRevisao(registros[0], motivo="score_baixo"),)

    caminho = gerar_relatorio_revisao(registros, revisao, tmp_path)

    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    assert ws.cell(2, 8).value == "score_baixo"  # Motivo Revisão
    assert ws.cell(2, 9).value in (None, "")  # sem candidato 1


def test_descricao_para_exibicao_usa_sigla_quando_bruta_e_codigo():
    assert descricao_para_exibicao("SND_LT67SAN_LT67SAN_79", "79") == "79"


def test_descricao_para_exibicao_mantem_descricao_real():
    assert descricao_para_exibicao("Disjuntor Aberto", "DJ") == "Disjuntor Aberto"


def test_descricao_para_exibicao_sem_sigla_mantem_bruta():
    assert descricao_para_exibicao("SND_LT67SAN_LT67SAN_79", None) == "SND_LT67SAN_LT67SAN_79"


def test_relatorio_usa_sigla_como_descricao_quando_bruta_e_codigo(tmp_path):
    from dataclasses import replace
    rec = replace(_rec("S1:3", sigla="79"), descricoes=Descricoes("SND_LT67SAN_LT67SAN_79", "x"))
    registros = [rec]

    caminho = gerar_relatorio_revisao(registros, (), tmp_path)

    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    assert ws.cell(2, 2).value == "79"  # Descrição Bruta -- sigla, não o código


def test_cabecalho_tem_formatacao_visual(tmp_path):
    cands = (Candidato("DJ", 0.91, "mesclado"),)
    registros = [_rec("S1:1", sigla="DJ", candidatos=cands)]

    caminho = gerar_relatorio_revisao(registros, (), tmp_path)

    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    for col in range(1, len(ws[1]) + 1):
        cell = ws.cell(1, col)
        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb == "004472C4"
        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FFFFFF"
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "thin"
        assert cell.alignment.horizontal == "center"


def test_largura_colunas_ajustada_pelo_conteudo(tmp_path):
    descricao_longa = "x" * 60
    registros = [
        SignalRecord(
            id="S1:1", modulo=Modulo("M", "sheet_name"),
            tipo_sinal=TipoSinal("Discrete", "SingleBit", "Input"),
            enderecamento=Enderecamento("DNP3", (10,)),
            descricoes=Descricoes(descricao_longa, "x"),
            sigla_sinal=None, candidatos=(), status="decidido",
            diagnostico=None,
        )
    ]

    caminho = gerar_relatorio_revisao(registros, (), tmp_path)

    wb = openpyxl.load_workbook(caminho)
    ws = wb["Auditoria"]
    # coluna "Descrição Bruta" (B) deve ter largura ajustada e capada em 40
    largura_b = ws.column_dimensions["B"].width
    assert largura_b is not None
    assert largura_b == 40  # capado: 60 chars + 2 > 40
    # nenhuma coluna deve ficar no default do openpyxl (None ou 8.43)
    for col_letra in ("A", "B", "C", "D", "E", "F", "G", "H"):
        largura = ws.column_dimensions[col_letra].width
        assert largura is not None
        assert largura != 8.43
