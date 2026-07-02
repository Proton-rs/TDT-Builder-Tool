import tempfile
from pathlib import Path

import openpyxl

from tdt.contracts import (
    Candidato, Descricoes, Diagnostico, Enderecamento, ItemRevisao,
    ListaHomogenea, Modulo, ResultadoPipeline, SignalRecord, TipoSinal,
)
from tdt.ui.exportar_analise import exportar_relatorio


def _rec(id_="s1", sigla="DJF1", candidatos=None, diagnostico=None, status="decidido"):
    return SignalRecord(
        id=id_,
        modulo=Modulo("M", "sheet_name"),
        tipo_sinal=TipoSinal("Discrete", "SingleBit", "Input"),
        enderecamento=Enderecamento("DNP3", (1,)),
        descricoes=Descricoes("DJ 52 FALHA", "DJ 52 FALHA"),
        sigla_sinal=sigla,
        status=status,
        candidatos=candidatos if candidatos is not None else (
            (Candidato(sigla, 0.85, "mesclado"),) if sigla else ()
        ),
        diagnostico=diagnostico,
        justificativa="ok",
    )


def _salvar_e_abrir(resultado):
    with tempfile.TemporaryDirectory() as tmp:
        destino = Path(tmp) / "relatorio.xlsx"
        exportar_relatorio(resultado, destino)
        wb = openpyxl.load_workbook(destino)
        ws1 = wb["Qualidade por Sinal"]
        ws2 = wb["Estatísticas"]
        dados1 = [list(row) for row in ws1.iter_rows(values_only=True)]
        dados2 = [list(row) for row in ws2.iter_rows(values_only=True)]
        wb.close()
        return dados1, dados2


def test_cria_xlsx_com_as_duas_sheets():
    resultado = ResultadoPipeline(
        lista=ListaHomogenea(subestacao="TEST", protocolo="DNP3", registros=(_rec(),)),
        revisao=(),
    )
    with tempfile.TemporaryDirectory() as tmp:
        destino = Path(tmp) / "relatorio.xlsx"
        exportar_relatorio(resultado, destino)
        wb = openpyxl.load_workbook(destino)
        assert "Qualidade por Sinal" in wb.sheetnames
        assert "Estatísticas" in wb.sheetnames
        wb.close()


def test_sheet_qualidade_tem_headers_esperados():
    resultado = ResultadoPipeline(
        lista=ListaHomogenea(subestacao="TEST", protocolo="DNP3", registros=(_rec(),)),
        revisao=(),
    )
    dados1, _ = _salvar_e_abrir(resultado)
    assert dados1[0] == [
        "ID", "Descrição", "Sigla Decidida", "Status",
        "Score Final", "TF-IDF", "Vetorial", "Fuzzy",
        "Gap", "Motivo Revisão",
    ]


def test_sheet_qualidade_uma_linha_por_registro_decidido():
    diag = Diagnostico({"DJF1": {"tfidf": 0.9, "vetorial": 0.8, "fuzzy": 0.7}})
    rec = _rec(id_="s1", sigla="DJF1", diagnostico=diag)
    resultado = ResultadoPipeline(
        lista=ListaHomogenea(subestacao="TEST", protocolo="DNP3", registros=(rec,)),
        revisao=(),
    )
    dados1, _ = _salvar_e_abrir(resultado)
    assert len(dados1) == 2  # header + 1 linha
    linha = dados1[1]
    assert linha[0] == "s1"
    assert linha[1] == "DJ 52 FALHA"
    assert linha[2] == "DJF1"
    assert linha[3] == "decidido"
    assert linha[4] == 0.85
    assert linha[5] == 0.9
    assert linha[6] == 0.8
    assert linha[7] == 0.7
    assert linha[9] is None  # openpyxl converte string vazia em None ao reabrir


def test_sheet_qualidade_nao_duplica_registros_em_revisao():
    rec_revisao = _rec(id_="s2", sigla=None, candidatos=(), status="revisao")
    item = ItemRevisao(registro=rec_revisao, motivo="score_baixo")
    resultado = ResultadoPipeline(
        lista=ListaHomogenea(subestacao="TEST", protocolo="DNP3", registros=(_rec(id_="s1"),)),
        revisao=(item,),
    )
    dados1, _ = _salvar_e_abrir(resultado)
    ids = [row[0] for row in dados1[1:]]
    assert sorted(ids) == ["s1", "s2"]
    assert len(ids) == len(set(ids))


def test_sheet_qualidade_motivo_revisao_aparece_na_linha_certa():
    rec_revisao = _rec(id_="s2", sigla=None, candidatos=(), status="revisao")
    item = ItemRevisao(registro=rec_revisao, motivo="score_baixo")
    resultado = ResultadoPipeline(
        lista=ListaHomogenea(subestacao="TEST", protocolo="DNP3", registros=()),
        revisao=(item,),
    )
    dados1, _ = _salvar_e_abrir(resultado)
    linha = dados1[1]
    assert linha[0] == "s2"
    assert linha[9] == "score_baixo"


def test_sheet_estatisticas_totais_e_taxa():
    rec_revisao = _rec(id_="s2", sigla=None, candidatos=(), status="revisao")
    item = ItemRevisao(registro=rec_revisao, motivo="score_baixo")
    resultado = ResultadoPipeline(
        lista=ListaHomogenea(subestacao="TEST", protocolo="DNP3", registros=(_rec(id_="s1"),)),
        revisao=(item,),
    )
    _, dados2 = _salvar_e_abrir(resultado)
    mapa = {row[0]: row[1] for row in dados2 if row[0] is not None}
    assert mapa["Total"] == 2
    assert mapa["Decididos"] == 1
    assert mapa["Revisão"] == 1
    assert mapa["Taxa de Decisão"] == "50.0%"


def test_sheet_estatisticas_motivos_ordenados_por_contagem_desc():
    rec1 = _rec(id_="s2", sigla=None, candidatos=(), status="revisao")
    rec2 = _rec(id_="s3", sigla=None, candidatos=(), status="revisao")
    rec3 = _rec(id_="s4", sigla=None, candidatos=(), status="revisao")
    revisao = (
        ItemRevisao(registro=rec1, motivo="score_baixo"),
        ItemRevisao(registro=rec2, motivo="score_baixo"),
        ItemRevisao(registro=rec3, motivo="sem_endereco"),
    )
    resultado = ResultadoPipeline(
        lista=ListaHomogenea(subestacao="TEST", protocolo="DNP3", registros=()),
        revisao=revisao,
    )
    _, dados2 = _salvar_e_abrir(resultado)
    inicio = dados2.index(["Motivo Revisão", "Contagem"]) + 1
    linhas_motivo = dados2[inicio:inicio + 2]
    assert linhas_motivo == [["score_baixo", 2], ["sem_endereco", 1]]


def test_sheet_estatisticas_taxa_sem_registros_eh_traco():
    resultado = ResultadoPipeline(
        lista=ListaHomogenea(subestacao="TEST", protocolo="DNP3", registros=()),
        revisao=(),
    )
    _, dados2 = _salvar_e_abrir(resultado)
    mapa = {row[0]: row[1] for row in dados2 if row[0] is not None}
    assert mapa["Total"] == 0
    assert mapa["Taxa de Decisão"] == "—"
