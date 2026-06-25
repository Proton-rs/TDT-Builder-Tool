from openpyxl import load_workbook
import re

# ANSI Code mappings: base_code -> (new_description_B, funcionamento_K)
ANSI_MAPPINGS = {
    20: ("20 - Proteção Diferencial de Corrente",
         "Proteção diferencial que monitora a diferença entre correntes de entrada e saída em transformadores e geradores. Detecta falhas internas comparando correntes nos terminais de um equipamento."),

    21: ("21 - Relé de Distância (Impedância)",
         "Proteção de linhas de transmissão baseada em medição de impedância/distância. Opera por zonas de proteção e oferece proteção escalonada. Sensível a variações de impedância da linha."),

    24: ("24 - Relé de Sobrecorrente com Reset de Sobrecarga",
         "Proteção de sobrecorrente com capacidade de reset (resetable). Monitora corrente contínua e oferece proteção contra sobrecargas com mecanismo de reinicialização automática."),

    25: ("25 - Relé de Sincronismo ou Verificação de Sincronismo",
         "Supervisiona sincronismo entre sistema e máquina síncrona. Controla condições de acoplamento para evitar conexão fora de fase. Monitora ângulo, magnitude e frequência."),

    26: ("26 - Proteção Térmica de Equipamento",
         "Monitora temperatura de equipamentos como transformadores e geradores. Detecta condições de sobretemperatura usando sensores térmicos ou modelos de elevação de temperatura."),

    27: ("27 - Relé de Subtensão",
         "Monitora tensão da rede e opera quando essa cai abaixo de valor pré-ajustado. Proteção contra perda de tensão e desconexão de circuitos em condições anormais."),

    32: ("32 - Relé de Potência Direcional",
         "Monitora fluxo de potência ativa em uma direção específica. Proteção contra fluxo reverso em geradores e alimentadores. Sensível à magnitude e direção da potência."),

    43: ("43 - Relé de Transferência Manual ou Seletor",
         "Controla transferência manual entre fontes de alimentação. Utilizado em operações de chaveamento e seleção de circuitos por comando do operador."),

    46: ("46 - Relé de Sequência Reversa ou Desbalanço de Fase",
         "Detecta sequência reversa de fases (RST em vez de STR) ou desbalanço entre fases. Proteção para motores e equipamentos sensíveis à sequência de fases."),

    49: ("49 - Relé Térmico",
         "Proteção térmica dos enrolamentos de máquinas. Monitora temperatura calculada ou medida e limita tempo de funcionamento em condições de sobrecarga térmica."),

    50: ("50 - Relé de Sobrecorrente Instantânea",
         "Detecta sobrecorrente com operação imediata (sem atraso de tempo). Proteção de alta velocidade contra curtos-circuitos. Usado para proteção primária rápida de circuitos."),

    51: ("51 - Relé de Tempo Inverso com Sobrecorrente AC",
         "Proteção de sobrecorrente com característica de tempo inverso. Atraso de tempo depende da magnitude da corrente. Proteção coordenada em cascata com relés downstream."),

    59: ("59 - Relé de Sobretensão",
         "Monitora tensão da rede e opera quando essa ultrapassa valor pré-ajustado. Proteção contra sobretensão e conexão indevida de circuitos em condições anormais."),

    61: ("61 - Relé de Retorno de Potência ou Potência Reversa",
         "Detecta inversão de fluxo de potência ativa. Proteção de geradores contra potência reversa indicando perda de atuação ou falha de máquina motriz."),

    62: ("62 - Relé de Atraso de Parada ou Desligamento",
         "Fornece atraso de tempo programável para operações de parada ou desligamento. Controla sequência temporizadas de eventos de proteção."),

    63: ("63 - Relé de Pressão",
         "Monitora pressão de gases ou líquidos em equipamentos (ex: transformadores com conservador). Proteção contra aumento anormal de pressão indicando falhas internas."),

    67: ("67 - Relé de Sobrecorrente Direcional AC",
         "Proteção de sobrecorrente direccionada que opera apenas para correntes acima do ajuste em direção específica. Permite coordenação bidirecional em malhas de distribuição."),

    71: ("71 - Relé de Nível de Líquido ou Gás",
         "Monitora nível de líquido ou gás em reservatórios e conservadores de equipamentos. Alerta ou desliga equipamentos quando nível está fora dos limites normais."),

    78: ("78 - Relé de Medição de Ângulo de Fase",
         "Mede e supervisiona ângulo de fase entre tensões ou correntes. Proteção contra desvios de fase e desincronismo. Usado em sistemas de proteção de estabilidade."),

    79: ("79 - Relé de Religa Automática AC",
         "Comanda religamento automático após operação de proteção. Restaura circuito após falhas transitórias. Inclui atrasos e contadores de ciclos de religa."),

    81: ("81 - Relé de Frequência",
         "Monitora frequência da rede. Detecta sub/sobfrequência indicando desbalanceamento gerador-carga. Proteção contra instabilidade do sistema e falhas de suprimento."),

    85: ("85 - Relé de Portadora ou Fio Piloto",
         "Comunica status de proteção entre terminais de linhas via portadora ou fio piloto. Coordenação de proteção de alta velocidade entre subestações."),

    86: ("86 - Relé de Bloqueio (Trip-Free)",
         "Bloqueia comandos de trip após operação. Mantém estado de travamento até reset manual. Proteção contra religamentos intempestivos."),

    87: ("87 - Relé de Proteção Diferencial",
         "Compara correntes em ambos os lados de um equipamento (transformador, gerador, linha). Detecta falhas internas por desbalanço de corrente diferencial."),

    90: ("90 - Dispositivo Regulador",
         "Controla e regula parâmetros de sistema como tensão, potência ou frequência. Mantém valores dentro de faixa pré-ajustada através de atuação contínua."),

    94: ("94 - Relé de Disparo ou Disparo Livre",
         "Relé auxiliar que comanda disparo final de disjuntores. Amplifica sinal de proteção e oferece garantia de abertura independente de outras proteções (trip-free).")
}

# Load workbook
wb = load_workbook('docs/Pontos Padrao ADMS_v3.xlsx')
ws = wb['DiscreteSignals']

# Track which base codes got K column updates (only first occurrence)
base_code_updated_K = set()

# Iterate through rows starting from row 2 (skip header)
for row in range(2, ws.max_row + 1):
    cell_a = ws.cell(row, 1).value  # SINAL column (ANSI code like "50F", "50BF")

    if cell_a is None:
        continue

    # Extract base ANSI code (first 1-2 digits)
    match = re.match(r'(\d{2})', str(cell_a))
    if not match:
        continue

    base_code = int(match.group(1))

    # Update if mapping exists
    if base_code in ANSI_MAPPINGS:
        desc_b, desc_k = ANSI_MAPPINGS[base_code]

        # Always update column B (DESCRIÇÃO NOVA)
        ws.cell(row, 2).value = desc_b

        # Update column K (Funcionamento) only for first occurrence of each base code
        if base_code not in base_code_updated_K:
            ws.cell(row, 11).value = desc_k
            base_code_updated_K.add(base_code)

# Save the workbook
wb.save('docs/Pontos Padrao ADMS_v3.xlsx')
print(f"[OK] Updated {len(base_code_updated_K)} ANSI codes")
print(f"[OK] Updated column B (DESCRICAO NOVA) for all occurrences")
print(f"[OK] Updated column K (Funcionamento) for primary entries: {sorted(base_code_updated_K)}")
print(f"[OK] File saved: docs/Pontos Padrao ADMS_v3.xlsx")
